# -*- coding: utf-8 -*-
"""
逐 SKU 导出店铺价格（支持批量）
- 单文件：保持 generate_price_excel(brand, input_dir, output_path) 原行为不变（取 input_dir 内“最近修改”的一个Excel）
- 批量：新增 generate_price_excels_bulk(brand, input_dir, output_dir, suffix)
  * 处理 input_dir 下所有 *.xlsx / *.xls
  * 输出文件名 = 输入文件名（不含扩展名） + suffix + ".xlsx"
  * 每个文件独立查价、独立导出，单个失败不影响其它文件
- 列识别：宝贝ID(=item_id)、商家编码(=product_code)、skuID(=skuid)
- 价格来源：品牌表中的 taobao_store_price（优先 product_code；兜底 product_name）
- 输出三列：宝贝id | skuid | 调整后价格
"""
from pathlib import Path
import unicodedata
import re
from typing import List, Dict, Optional, Iterable, Tuple
import pandas as pd
import psycopg2
import math

from config import BRAND_CONFIG

# ===== 列名别名 =====
COL_ALIASES = {
    "item_id": {"item_id","itemid","ITEM_ID","宝贝id","宝贝ID","宝贝Id","宝贝"},
    "product_code": {
        "product_code","productcode","code",
        "商品编码","商品Code","产品编码","编码","货号",
        "商家编码",
        "商家货号","外部编码","外部代码",
        "outer_id","outerid","outer code","outercode"
    },
    "skuid": {
        "skuid","sku_id","SkuId","SKU_ID","SKUID",
        "skuID",
        "渠道货品ID","渠道skuid","货品id","货品ID"
    },
}

_SPLIT = re.compile(r"[,\uFF0C;；\s\r\n]+")


import pandas as pd

import pandas as pd
from pathlib import Path

def _load_blacklist_codes(blacklist_excel_file: str | None) -> set[str]:
    """
    从 blacklist_excel_file 读取黑名单商品编码列表。
    如果 blacklist_excel_file 是 None 或文件不存在，则返回空集合，表示不启用黑名单。
    """
    if not blacklist_excel_file:
        # 没传，等于不启用黑名单
        print("[BLACKLIST] 未提供黑名单文件，黑名单为空")
        return set()

    path_obj = Path(blacklist_excel_file)
    if not path_obj.exists():
        print(f"[BLACKLIST] 黑名单文件不存在: {blacklist_excel_file}，黑名单为空")
        return set()

    df_blk = pd.read_excel(path_obj)

    candidate_cols = [
        "商品编码",
        "product_code",
        "商家编码",
        "外部商品编码",
        "商家货号",
        "货号",
        "编码",
    ]

    col_found = None
    for c in candidate_cols:
        if c in df_blk.columns:
            col_found = c
            break

    if col_found is None:
        print("[BLACKLIST] 未找到可识别的黑名单列，黑名单为空")
        return set()

    codes = (
        df_blk[col_found]
        .astype(str)
        .str.strip()
        .replace({"": pd.NA})
        .dropna()
        .unique()
        .tolist()
    )

    blacklist = {c.upper() for c in codes}
    print(f"[BLACKLIST] 读取到 {len(blacklist)} 个黑名单商品编码")
    return blacklist

import logging
from pathlib import Path

def _get_logger(name: str = "price_export"):
    logger = logging.getLogger(name)
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        h = logging.StreamHandler()
        fmt = logging.Formatter("[%(levelname)s] %(message)s")
        h.setFormatter(fmt)
        logger.addHandler(h)
    return logger

from datetime import datetime
import pandas as pd
from pathlib import Path

def _flush_filtered(output_dir: Path, input_file: str, logger=None):
    global filtered_rows
    if not filtered_rows:
        return
    temp_dir = Path(output_dir) / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)

    stem = Path(input_file).stem
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = temp_dir / f"{stem}_filtered_reasons_{ts}.xlsx"

    df_log = pd.DataFrame(filtered_rows)
    if "filter_reason" in df_log.columns:
        df_log = df_log.sort_values(["filter_reason"])

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
        df_log.to_excel(w, index=False, sheet_name="filtered_rows")
        summary = df_log.groupby("filter_reason").size().reset_index(name="count")
        summary.to_excel(w, index=False, sheet_name="summary")

    if logger:
        logger.info(f"过滤原因明细已写入: {out_xlsx}")
    else:
        print(f"[LOG] 过滤原因明细已写入: {out_xlsx}")



import psycopg2
from psycopg2.extras import execute_values

from psycopg2.extras import execute_values

def _fetch_prices_by_code_and_size_bulk(
    conn,
    table: str,
    code_size_pairs: list[tuple[str, str]]
) -> pd.DataFrame:
    """
    批量从 {table} 里拿 (product_code, size) 对应的 taobao_store_price。
    关键点：
    - 不再使用 execute_values（之前只拿到最后一批，导致严重丢价）
    - 手动分批，每批用 WITH wanted AS (VALUES ...) JOIN，逐批执行 cur.execute()
    - 汇总所有批次返回，去重后输出 DataFrame
    """

    # 1. 先把输入规格清洗成稳定字符串键，防止 '16 ' / None 等脏值
    cleaned_pairs = []
    for code, sz in code_size_pairs:
        if code is None or sz is None:
            continue
        c = str(code).strip()
        s = str(sz).strip()
        if c and s:
            cleaned_pairs.append((c, s))

    # 去重，避免同一个 (code,size) 重复查
    unique_pairs = sorted(set(cleaned_pairs))
    if not unique_pairs:
        return pd.DataFrame(columns=["product_code", "size", "taobao_store_price"])

    print(f"[PRICEDEBUG] 待查询组合总数: {len(cleaned_pairs)}，去重后: {len(unique_pairs)}")

    # 2. 分批，每批 50 个 (code,size)
    CHUNK_SIZE = 50
    all_rows: list[tuple[str, str, float]] = []

    with conn.cursor() as cur:
        total_batches = (len(unique_pairs) + CHUNK_SIZE - 1) // CHUNK_SIZE

        for batch_idx, start in enumerate(range(0, len(unique_pairs), CHUNK_SIZE), start=1):
            batch = unique_pairs[start:start + CHUNK_SIZE]

            # 用 mogrify 把每条 (code,size) 变成安全的 SQL ('CODE','SIZE') 片段
            values_sql_list = [
                cur.mogrify("(%s,%s)", (code, size)).decode("utf-8")
                for code, size in batch
            ]
            values_sql_block = ",".join(values_sql_list)

            # 构造本批 SQL
            query_sql = (
                f"WITH wanted(code,size) AS (VALUES {values_sql_block}) "
                f"SELECT t.product_code, t.size, t.taobao_store_price "
                f"FROM {table} t "
                "JOIN wanted w ON t.product_code = w.code AND t.size = w.size"
            )

            # Debug: 打印本批的关键信息（不会太吵）
            print("\n[DEBUG SQL BATCH]")
            print(f"  批次 {batch_idx}/{total_batches}, 本批组合数: {len(batch)}")
            print(f"  示例VALUES前5个: {', '.join(values_sql_list[:5])}")

            # 真正执行
            cur.execute(query_sql)
            batch_rows = cur.fetchall()
            print(f"  -> 本批返回 {len(batch_rows)} 行")

            all_rows.extend(batch_rows)

    print(f"[PRICEDEBUG] 全部分批合计返回 {len(all_rows)} 行")

    # 3. 汇总并去重 (防止同一个 (product_code,size) 多次出现)
    merged_unique = {}
    for (product_code, size, price_val) in all_rows:
        key = (str(product_code).strip(), str(size).strip())
        if key not in merged_unique:
            merged_unique[key] = price_val  # 第一条为准

    rows_for_df = [
        {
            "product_code": pc,
            "size": sz,
            "taobao_store_price": price_val,
        }
        for (pc, sz), price_val in merged_unique.items()
    ]

    df_price = pd.DataFrame(
        rows_for_df,
        columns=["product_code", "size", "taobao_store_price"]
    )

    print(f"[PRICEDEBUG] 去重后最终价格行数: {len(df_price)}")
    return df_price




def _canon(s: str) -> str:
    if s is None: return ""
    s = unicodedata.normalize("NFKC", str(s)).strip()
    s = s.replace("\u00A0", " ").replace("\u200B", "")
    return s.lower()

def _normalize_col(df: pd.DataFrame, want: str) -> str:
    canon2raw = {_canon(c): c for c in df.columns}
    for alias in COL_ALIASES[want]:
        key = _canon(alias)
        if key in canon2raw:
            return canon2raw[key]
    raise KeyError(f"Excel中缺少必要列：{want}（可用别名：{COL_ALIASES[want]}），当前表头：{list(df.columns)}")

def _list_excels(input_dir: Path) -> List[Path]:
    files = list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xls"))
    return sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)

def _find_latest_excel(input_dir: Path) -> Path:
    files = _list_excels(input_dir)
    if not files:
        raise FileNotFoundError(f"目录没有找到 Excel：{input_dir}")
    return files[0]

def _split_skuids(val) -> List[str]:
    if pd.isna(val): return []
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        val = str(int(val))
    s = str(val).strip()
    if not s: return []
    parts = [p.strip() for p in _SPLIT.split(s) if p.strip()]
    parts = [re.sub(r"[^\w\-]", "", p) for p in parts if p]
    return [p for p in parts if p]

def _chunked(it: Iterable, size: int):
    buf = []
    for x in it:
        buf.append(x)
        if len(buf) >= size:
            yield buf; buf = []
    if buf: yield buf

def _fetch_prices(conn, table: str, codes: List[str]) -> Dict[str, Optional[float]]:
    prices: Dict[str, Optional[float]] = {}
    missing = list(dict.fromkeys(codes))
    with conn.cursor() as cur:
        try:
            for chunk in _chunked(missing, 1000):
                cur.execute(
                    f"SELECT product_code, taobao_store_price FROM {table} "
                    f"WHERE product_code = ANY(%s)", (chunk,))
                for code, price in cur.fetchall():
                    prices[str(code)] = None if price is None else float(price)
            missing = [c for c in missing if c not in prices]
        except Exception:
            pass
        if missing:
            try:
                for chunk in _chunked(missing, 1000):
                    cur.execute(
                        f"SELECT product_name, taobao_store_price FROM {table} "
                        f"WHERE product_name = ANY(%s)", (chunk,))
                    for code, price in cur.fetchall():
                        prices[str(code)] = None if price is None else float(price)
            except Exception:
                pass
    return prices

def _to_text(v: object) -> str:
    if v is None or (isinstance(v, float) and (math.isnan(v))):
        return ""
    if isinstance(v, (int,)):
        return str(v)
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return f"{v:.15g}"
    if isinstance(v, pd._libs.missing.NAType) or (isinstance(v, str) and v.lower() == "nan"):
        return ""
    return str(v)

# ===== 单文件：保持兼容 =====
def generate_price_excel(
    brand: str,
    input_dir: str | Path,
    output_path: str | Path,
    drop_rows_without_price: bool = True
) -> Path:
    brand = brand.lower().strip()
    if brand not in BRAND_CONFIG:
        raise ValueError(f"未知品牌：{brand}")
    cfg = BRAND_CONFIG[brand]
    table = cfg["TABLE_NAME"]
    pg    = cfg["PGSQL_CONFIG"]

    input_dir = Path(input_dir)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    excel_file = _find_latest_excel(input_dir)
    print(f"📄 使用输入文件：{excel_file}")
    return _generate_price_excel_from_file(brand, excel_file, output_path, drop_rows_without_price, table, pg)

# ===== 批量：处理 input_dir 下所有 Excel =====
def generate_price_excels_bulk(
    brand: str,
    input_dir: str,
    output_dir: str,
    suffix: str = "_价格",
    drop_rows_without_price: bool = True,
    blacklist_excel_file: str | None = None,
):
    """
    从 input_dir 中批量读取店铺 Excel（每个代表一个店铺），
    调用 _generate_price_excel_from_file() 生成对应价格表，
    并输出到 output_dir。

    新增参数:
        blacklist_excel_file: 黑名单 Excel 的绝对路径。
                              如果提供，则会过滤掉黑名单商品编码。
                              如果为 None，则不启用黑名单。
    """

    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 数据库配置
    from config import BRAND_CONFIG
    pgsql_config = BRAND_CONFIG[brand]["PGSQL_CONFIG"]
    table_name = BRAND_CONFIG[brand]["TABLE_NAME"]

    # 找出目录下所有 Excel 文件
    excel_files = list(input_dir.glob("*.xlsx"))
    if not excel_files:
        print(f"⚠ 没找到任何 Excel 文件: {input_dir}")
        return

    print(f"[INFO] 共发现 {len(excel_files)} 个输入文件，将生成价格表...")
    print(f"[INFO] 品牌: {brand}")
    print(f"[INFO] 黑名单文件: {blacklist_excel_file or '未启用'}")

    for f in excel_files:
        try:
            print(f"\n[START] 处理文件: {f.name}")
            _generate_price_excel_from_file(
                file_path=str(f),
                output_dir=str(output_dir),
                brand=brand,
                pgsql_config=pgsql_config,
                blacklist_excel_file=blacklist_excel_file,
                table_name=table_name,
            )
        except Exception as e:
            print(f"[ERROR] 处理 {f.name} 失败: {e}")

    print(f"\n✅ 所有文件处理完成。输出路径: {output_dir}")


def normalize_size_for_taobao(size_raw):
    """
    将淘宝导出的尺码格式统一成数据库里的格式：
    - '38-39' -> '38'
    - '34码'  -> '34'
    - 其他不含 '码'、'-' 的值原样返回（如 '39'）
    """
    if size_raw is None:
        return None

    if not isinstance(size_raw, str):
        size_str = str(size_raw)
    else:
        size_str = size_raw

    size_str = size_str.strip()

    # 1) 去掉“码”
    size_str = size_str.replace("码", "").strip()

    # 2) 如果是 "38-39" 这种，取前半段
    if "-" in size_str:
        size_str = size_str.split("-", 1)[0].strip()

    # 3) 空串视为 None
    return size_str or None



import pandas as pd
from pathlib import Path

def _normalize_input_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    我们需要拿到三列：
    - item_id   (宝贝id)
    - skuid     (SKU唯一标识，渠道货品ID)
    - sku_spec  (sku规格，比如 'MWX0339OL71,M')

    这个函数的作用是：从各种可能的列名里抽取并标准化成这三列。
    """

    # 可能的列名映射
    item_id_candidates = ["宝贝id", "宝贝ID", "item_id", "itemid", "itemId", "商品ID", "item id"]
    skuid_candidates   = ["skuid", "SKU ID", "skuID", "SKUId", "渠道货品ID", "渠道货品id", "货品id", "sku id"]
    spec_candidates    = ["sku规格", "SKU规格", "规格", "sku spec", "销售属性"]

    colmap = {}

    # 找 item_id
    for c in item_id_candidates:
        if c in df.columns:
            colmap["item_id"] = c
            break
    # 找 skuid
    for c in skuid_candidates:
        if c in df.columns:
            colmap["skuid"] = c
            break
    # 找 sku_spec
    for c in spec_candidates:
        if c in df.columns:
            colmap["sku_spec"] = c
            break

    missing = [name for name in ["item_id","skuid","sku_spec"] if name not in colmap]
    if missing:
        print("⚠ 输入Excel缺少必要列:", missing)
        # 我们还是尽量返回一些列，后面会 dropna
    # 复制一份只保留我们关心的列，并重命名
    df2 = df.copy()
    out = pd.DataFrame()
    out["item_id"] = df2[colmap["item_id"]] if "item_id" in colmap else pd.NA
    out["skuid"]   = df2[colmap["skuid"]]   if "skuid"   in colmap else pd.NA
    out["sku_spec"]= df2[colmap["sku_spec"]]if "sku_spec"in colmap else pd.NA

    return out


def _split_spec_value(spec_val: str) -> tuple[str|None, str|None]:
    """
    把 'MWX0339OL71,M' 这种值拆成 ('MWX0339OL71','M')
    要求：第一个逗号前 = 商品编码
          第一个逗号后 = 尺码
    如果没有逗号，或为空，返回 (None, None)
    """
    if not isinstance(spec_val, str):
        return (None, None)
    raw = spec_val.strip()
    if raw == "":
        return (None, None)
    parts = [p.strip() for p in raw.split(",")]
    if len(parts) < 2:
        # 没有尺码信息，视为无效
        return (None, None)
    code_part = parts[0].strip().upper()
    size_part = parts[1].strip()
    if code_part == "" or size_part == "":
        return (None, None)
    return (code_part, size_part)





def _clean_spec_key(spec_val):
    if not isinstance(spec_val, str):
        return None
    v = spec_val.strip()
    return v if v else None


from datetime import datetime

# === 过滤原因记录器 ===
filtered_rows = []  # list[dict]

def _add_filtered(df_part, reason: str, extra: dict | None = None, limit: int | None = None):
    """
    df_part: 被过滤掉的 DataFrame（或一行转成 dict 也行）
    reason: 过滤原因
    extra: 额外字段
    limit: 限制记录条数，避免极端情况下文件太大（None=不限制）
    """
    if df_part is None:
        return
    if isinstance(df_part, dict):
        row = dict(df_part)
        row["filter_reason"] = reason
        if extra: row.update(extra)
        filtered_rows.append(row)
        return

    if df_part.empty:
        return

    cols_keep = [c for c in ["item_id", "skuid", "sku_spec", "code_part", "size_part", "spec_key"] if c in df_part.columns]
    sub = df_part[cols_keep].copy()
    sub["filter_reason"] = reason
    if extra:
        for k, v in extra.items():
            sub[k] = v
    if limit is not None:
        sub = sub.head(limit)
    filtered_rows.extend(sub.to_dict("records"))






def _generate_price_excel_from_file(
    file_path: str,
    output_dir: str,
    brand: str,
    pgsql_config: dict,
    blacklist_excel_file: str | None,
    table_name: str = "barbour_inventory",
):
    """
    生成价格导入表：宝贝id | skuid | 调整后价格
    额外能力：
      - 用 logger 打关键统计
      - 过滤掉的行写入 output_dir/temp/*_filtered_reasons_*.xlsx 方便排查
    """
    import pandas as pd
    import psycopg2
    from pathlib import Path

    logger = _get_logger(f"price_export.{brand}")
    logger.info(f"开始处理: {file_path}")

    # 每个文件重新清空过滤记录（你原本是全局列表）
    global filtered_rows
    filtered_rows = []

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1) 黑名单
    blacklist_codes = _load_blacklist_codes(blacklist_excel_file)
    logger.info(f"黑名单启用: {'是' if blacklist_codes else '否'} | count={len(blacklist_codes)}")

    # 2) 读输入 Excel + 标准化列
    df_raw = pd.read_excel(file_path, dtype=object)
    logger.info(f"原始行数 df_raw: {len(df_raw)}")

    df_norm = _normalize_input_columns(df_raw)   # -> item_id, skuid, sku_spec
    logger.info(f"标准化后 df_norm: {len(df_norm)} | columns={list(df_norm.columns)}")

    # 2.1 宝贝id 向下填充（淘宝导出常见结构）
    df_norm["item_id"] = df_norm["item_id"].ffill()

    # 2.2 item_id 字符串化（防止科学计数）
    def _item_id_to_str(val):
        if pd.isna(val):
            return ""
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return str(int(val))
        return str(val).strip()

    df_norm["item_id"] = df_norm["item_id"].map(_item_id_to_str)

    # 3) sku_spec -> code_part / size_part
    pairs = df_norm["sku_spec"].map(_split_spec_value)
    df_norm["code_part"] = [p[0] for p in pairs]
    df_norm["size_part"] = [normalize_size_for_taobao(p[1]) for p in pairs]

    # 3.1 drop 掉 code/size 缺失行（并记录原因）
    before = len(df_norm)
    dropped = df_norm[df_norm[["code_part", "size_part"]].isna().any(axis=1)].copy()
    _add_filtered(dropped, "dropna_code_or_size")
    df_norm = df_norm.dropna(subset=["code_part", "size_part"])
    logger.info(f"dropna(code/size): {before} -> {len(df_norm)}")

    # 4) 黑名单过滤（整款过滤）
    if blacklist_codes:
        before = len(df_norm)
        blocked = df_norm[df_norm["code_part"].isin(blacklist_codes)].copy()
        _add_filtered(blocked, "blacklist_filtered")
        df_norm = df_norm[~df_norm["code_part"].isin(blacklist_codes)]
        logger.info(f"blacklist过滤: {before} -> {len(df_norm)}")

    # 5) 展开 skuid（可能是一格多个）
    rows = []
    empty_sku_cnt = 0

    for _, r in df_norm.iterrows():
        item_id = r["item_id"]
        code = r["code_part"]
        size = r["size_part"]
        skus_raw = r["skuid"]
        sku_list = _split_skuids(skus_raw)

        if not sku_list:
            empty_sku_cnt += 1
            _add_filtered(
                {
                    "item_id": item_id,
                    "skuid": skus_raw,
                    "sku_spec": r.get("sku_spec", None),
                    "code_part": code,
                    "size_part": size,
                },
                "empty_sku_list"
            )
            continue

        for sid in sku_list:
            if sid:
                rows.append({
                    "item_id": item_id,
                    "skuid": sid,
                    "code_part": code,
                    "size_part": size,
                })

    df_expanded = pd.DataFrame(rows, columns=["item_id", "skuid", "code_part", "size_part"])
    logger.info(f"展开SKU后 df_expanded: {len(df_expanded)} | empty_sku_rows={empty_sku_cnt}")

    if df_expanded.empty:
        logger.warning("展开后无有效SKU行，终止该文件。")
        _flush_filtered(out_dir, file_path, logger=logger)
        return

    # 6) 批量查价
    pairs_for_price = list(zip(df_expanded["code_part"], df_expanded["size_part"]))
    uniq_pairs = len(set(pairs_for_price))
    logger.info(f"待查价组合: total={len(pairs_for_price)} unique={uniq_pairs} | table={table_name}")

    conn = psycopg2.connect(**pgsql_config)
    try:
        df_price = _fetch_prices_by_code_and_size_bulk(conn, table_name, pairs_for_price)
    finally:
        conn.close()

    logger.info(f"数据库返回 df_price 行数: {len(df_price)}")
    if df_price.empty:
        # 全都没价格，直接输出过滤原因
        _add_filtered(df_expanded, "missing_db_price_all_empty_price_table")
        _flush_filtered(out_dir, file_path, logger=logger)
        logger.warning("df_price 为空：该文件无法生成价格表。")
        return

    # 7) 合并价格
    df_merged = df_expanded.merge(
        df_price,
        left_on=["code_part", "size_part"],
        right_on=["product_code", "size"],
        how="left"
    )
    logger.info(f"合并后 df_merged: {len(df_merged)}")

    # 7.1 记录缺价行
    miss = df_merged[df_merged["taobao_store_price"].isna()].copy()
    _add_filtered(miss, "missing_db_price")

    df_merged = df_merged.dropna(subset=["taobao_store_price"])
    logger.info(f"丢掉无价后 df_merged: {len(df_merged)}")

    if df_merged.empty:
        _flush_filtered(out_dir, file_path, logger=logger)
        logger.warning("全部行都缺价，未生成价格Excel。")
        return

    # 8) 导出三列
    df_final = df_merged[["item_id", "skuid", "taobao_store_price"]].copy()
    df_final = df_final.rename(columns={
        "item_id": "宝贝id",
        "skuid": "skuid",
        "taobao_store_price": "调整后价格",
    })

    # 9) 去重 skuid
    before = len(df_final)
    df_final = df_final.drop_duplicates(subset=["skuid"], keep="first")
    logger.info(f"去重skuid: {before} -> {len(df_final)}")

    # 10) 宝贝id 强制字符串
    df_final["宝贝id"] = df_final["宝贝id"].astype(str)

    # 11) 写 Excel
    src_name = Path(file_path).stem
    out_path = out_dir / f"{src_name}_price.xlsx"
    df_final.to_excel(out_path, index=False)

    logger.info(f"生成完成: {out_path} | rows={len(df_final)}")

    # 12) 输出过滤原因明细（有就写）
    _flush_filtered(out_dir, file_path, logger=logger)











# ===== 新增：批量比较一口价 vs 数据库 taobao_store_price =====

def _fetch_product_prices_by_code(conn, table: str, codes: list[str]) -> pd.DataFrame:
    """
    按 product_code 批量查询两个价格字段（同款多尺码聚合为一行）：
    - taobao_store_price  取 AVG（均值作为代表价）
    - jingya_untaxed_price 取 MAX（保守取最高，与 check_taobao_item_price_diff 一致）
    返回 DataFrame: product_code | taobao_store_price | jingya_untaxed_price
    """
    empty = pd.DataFrame(columns=["product_code", "taobao_store_price", "jingya_untaxed_price"])
    if not codes:
        return empty
    unique_codes = list(dict.fromkeys(str(c).strip() for c in codes if c and str(c).strip()))
    if not unique_codes:
        return empty

    rows: list[dict] = []
    CHUNK_SIZE = 500
    with conn.cursor() as cur:
        for start in range(0, len(unique_codes), CHUNK_SIZE):
            chunk = unique_codes[start:start + CHUNK_SIZE]
            cur.execute(
                f"SELECT product_code, AVG(taobao_store_price), MAX(jingya_untaxed_price) "
                f"FROM {table} "
                f"WHERE product_code = ANY(%s) AND taobao_store_price IS NOT NULL "
                f"GROUP BY product_code",
                (chunk,)
            )
            for code, avg_store, max_jingya in cur.fetchall():
                try:
                    rows.append({
                        "product_code": str(code).strip(),
                        "taobao_store_price":   float(avg_store)  if avg_store  is not None else None,
                        "jingya_untaxed_price": float(max_jingya) if max_jingya is not None else None,
                    })
                except (ValueError, TypeError):
                    pass
    return pd.DataFrame(rows) if rows else empty


def _write_diff_excel_with_highlight(df_out: pd.DataFrame, out_path: "Path") -> None:
    """写 Excel 并对倒挂行（倒挂安全比例 ≤ 0）高亮红色。"""
    df_out.to_excel(out_path, index=False)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font

        wb = load_workbook(out_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        inv_col_idx = next(
            (i + 1 for i, h in enumerate(headers) if str(h) == "倒挂安全比例"),
            None,
        )
        if inv_col_idx:
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            bold_red = Font(bold=True, color="CC0000")
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=inv_col_idx).value
                try:
                    if float(v) <= 0:
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=r, column=c).fill = red_fill
                        ws.cell(row=r, column=inv_col_idx).font = bold_red
                except (TypeError, ValueError):
                    pass
        wb.save(out_path)
    except Exception:
        pass


def _compare_price_vs_db_single(
    file_path: str,
    output_dir: "Path",
    table: str,
    pgsql_config: dict,
    suffix: str,
    price_diff_threshold: float,
    blacklist_codes: "set[str]",
    taobao_discount: float,
    tax_factor: float,
    min_profit: float,
    logger,
):
    """
    单文件：比较一口价 vs 数据库价格，输出差价 + 倒挂安全报告。

    倒挂计算：
      到手价         = 一口价 × taobao_discount × tax_factor
      安全线         = jingya_untaxed_price × (1 + min_profit)
      倒挂安全比例   = (到手价 - 安全线) / 安全线   （> 0 才能卖）

    输出行条件（满足其一即输出）：
      - |差异百分比| > price_diff_threshold
      - 倒挂安全比例 ≤ 0
      - jingya_untaxed_price 在 DB 中缺失（无法验证利润）
    """
    df_raw = pd.read_excel(file_path, dtype=object)
    logger.info(f"原始行数: {len(df_raw)} | 列: {list(df_raw.columns)}")

    # 1) 识别列名
    def _find_col(df, candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    item_id_col = _find_col(df_raw, ["宝贝ID", "宝贝id", "宝贝Id", "item_id", "itemid", "商品ID"])
    title_col   = _find_col(df_raw, ["宝贝标题", "标题", "商品标题", "title"])
    price_col   = _find_col(df_raw, ["一口价", "售价", "销售价", "price"])
    spec_col    = _find_col(df_raw, ["sku规格", "SKU规格", "规格", "sku spec", "销售属性"])

    missing = [n for n, c in [("宝贝ID", item_id_col), ("一口价", price_col), ("sku规格", spec_col)] if c is None]
    if missing:
        raise ValueError(f"Excel缺少必要列: {missing}，当前列: {list(df_raw.columns)}")

    # 2) 向下填充（一款对应多 SKU 行）
    df = df_raw.copy()
    df[item_id_col] = df[item_id_col].ffill()
    df[price_col]   = df[price_col].ffill()
    if title_col:
        df[title_col] = df[title_col].ffill()

    # 3) 从 sku规格 提取 product_code（格式：CODE,SIZE[,]）
    df["_product_code"] = df[spec_col].map(
        lambda v: _split_spec_value(str(v) if v is not None else "")[0]
    )

    def _to_float_safe(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    df["_item_id_str"] = df[item_id_col].map(_to_text)
    df["_price_float"] = df[price_col].map(_to_float_safe)

    # 4) 聚合到商品级（一款一行）
    product_rows = []
    for item_id, gdf in df.groupby("_item_id_str", sort=False):
        if not item_id:
            continue
        codes = gdf["_product_code"].dropna()
        if codes.empty:
            continue
        prices = gdf["_price_float"].dropna()
        title  = gdf[title_col].dropna().iloc[0] if (title_col and not gdf[title_col].dropna().empty) else None
        product_rows.append({
            "channel_product_id": item_id,
            "product_code":       str(codes.iloc[0]),
            "宝贝标题":            title,
            "一口价":              float(prices.iloc[0]) if not prices.empty else None,
        })

    df_products = pd.DataFrame(product_rows)
    logger.info(f"识别到商品数: {len(df_products)}")

    if df_products.empty:
        logger.warning("无有效商品，跳过该文件。")
        return

    # 5) 黑名单过滤
    if blacklist_codes:
        before = len(df_products)
        df_products = df_products[~df_products["product_code"].str.upper().isin(blacklist_codes)]
        logger.info(f"黑名单过滤: {before} -> {len(df_products)}")

    # 6) 过滤一口价无效行
    df_products = df_products.dropna(subset=["一口价"])
    df_products = df_products[df_products["一口价"] > 0]
    logger.info(f"有效一口价商品数: {len(df_products)}")
    if df_products.empty:
        logger.warning("无有效一口价，跳过输出。")
        return

    # 7) 批量查 DB 价格（taobao_store_price + jingya_untaxed_price）
    conn = psycopg2.connect(**pgsql_config)
    try:
        df_db = _fetch_product_prices_by_code(conn, table, df_products["product_code"].tolist())
    finally:
        conn.close()
    logger.info(f"DB返回价格行数: {len(df_db)}")

    # 8) 合并 DB 价格（左连接，保留 DB 缺失的商品以便标记）
    df_cmp = df_products.merge(df_db, on="product_code", how="left")

    # 9) 计算差异百分比：(一口价 - taobao_store_price) / taobao_store_price × 100
    mask_has_store = df_cmp["taobao_store_price"].notna() & (df_cmp["taobao_store_price"] > 0)
    df_cmp["差异百分比"] = pd.NA
    df_cmp.loc[mask_has_store, "差异百分比"] = (
        (df_cmp.loc[mask_has_store, "一口价"] - df_cmp.loc[mask_has_store, "taobao_store_price"])
        / df_cmp.loc[mask_has_store, "taobao_store_price"] * 100
    ).round(2)

    # 10) 计算倒挂安全比例
    #     到手价 = 一口价 × taobao_discount × tax_factor
    #     安全线 = jingya_untaxed_price × (1 + min_profit)
    #     倒挂安全比例 = (到手价 - 安全线) / 安全线
    mask_has_jingya = df_cmp["jingya_untaxed_price"].notna() & (df_cmp["jingya_untaxed_price"] > 0)
    df_cmp["倒挂安全比例"] = pd.NA
    net_price  = df_cmp.loc[mask_has_jingya, "一口价"] * taobao_discount * tax_factor
    safe_line  = df_cmp.loc[mask_has_jingya, "jingya_untaxed_price"] * (1 + min_profit)
    df_cmp.loc[mask_has_jingya, "倒挂安全比例"] = ((net_price - safe_line) / safe_line).round(4)

    # 11) 筛选需要输出的行
    has_diff       = df_cmp["差异百分比"].notna() & (df_cmp["差异百分比"].abs() > price_diff_threshold)
    is_inverted    = df_cmp["倒挂安全比例"].notna() & (df_cmp["倒挂安全比例"] <= 0)
    no_jingya      = df_cmp["jingya_untaxed_price"].isna()  # 无法验证利润
    df_diff = df_cmp[has_diff | is_inverted | no_jingya].copy()

    logger.info(
        f"输出行数: {len(df_diff)} "
        f"（价格偏差: {has_diff.sum()}, 倒挂: {is_inverted.sum()}, DB缺失未税价: {no_jingya.sum()}）"
    )
    if df_diff.empty:
        logger.info("无问题商品，跳过输出。")
        return

    # 12) 排序：倒挂行优先，其次按差异百分比绝对值降序
    df_diff["_inv_flag"]  = (~df_diff["倒挂安全比例"].isna()) & (df_diff["倒挂安全比例"] <= 0)
    df_diff["_abs_diff"]  = df_diff["差异百分比"].fillna(0).abs()
    df_diff = df_diff.sort_values(["_inv_flag", "_abs_diff"], ascending=[False, False])
    df_diff = df_diff.drop(columns=["_inv_flag", "_abs_diff"])

    # 13) 整理输出列
    out_cols = [
        "宝贝标题", "channel_product_id", "product_code",
        "一口价", "taobao_store_price", "jingya_untaxed_price",
        "差异百分比", "倒挂安全比例",
    ]
    if not title_col:
        out_cols = [c for c in out_cols if c != "宝贝标题"]
    df_out = df_diff[[c for c in out_cols if c in df_diff.columns]].copy()
    df_out["channel_product_id"] = df_out["channel_product_id"].astype(str)

    # 14) 写 Excel + 高亮倒挂行
    out_path = output_dir / f"{Path(file_path).stem}{suffix}.xlsx"
    _write_diff_excel_with_highlight(df_out, out_path)
    logger.info(f"差价报告已生成: {out_path} | rows={len(df_out)}")


def compare_price_vs_db_bulk(
    brand: str,
    input_dir: str,
    output_dir: str,
    suffix: str = "_差价",
    price_diff_threshold: float = 0.0,
    blacklist_excel_file: str | None = None,
    taobao_discount: float = 0.85,
    tax_factor: float = 0.9,
    min_profit: float = 0.03,
):
    """
    批量比较店铺 Excel 的一口价 与数据库价格的差异，同时检查倒挂风险。

    Excel 格式（淘宝导出）：
      - 宝贝标题/宝贝ID/一口价 只在每款的第一行出现，其余行为空（向下填充）
      - sku规格 格式为 "product_code,size[,]"，每 SKU 一行

    输出列（每个输入文件对应一个输出 Excel）：
      宝贝标题 | channel_product_id | product_code | 一口价
      | taobao_store_price | jingya_untaxed_price | 差异百分比 | 倒挂安全比例

    输出行条件（满足其一即输出）：
      - |差异百分比| > price_diff_threshold
      - 倒挂安全比例 ≤ 0（到手价已低于未税成本线）
      - DB 中缺少 jingya_untaxed_price（无法验证利润）

    排序：倒挂行排最前，其次按差异百分比绝对值降序。
    倒挂行高亮红色。

    参数：
      brand                  品牌 key（小写）
      input_dir              存放店铺导出 Excel 的目录
      output_dir             输出报告的目录
      suffix                 输出文件名后缀，默认 "_差价"
      price_diff_threshold   差异绝对值超过该百分比才输出（默认 0.0 = 全部差异）
      blacklist_excel_file   黑名单 Excel 路径（可选）
      taobao_discount        淘宝平台扣点（默认 0.85）
      tax_factor             关税/税费系数（默认 0.9）
      min_profit             最低利润率要求（默认 0.03 = 3%）
    """
    brand = brand.lower().strip()
    if brand not in BRAND_CONFIG:
        raise ValueError(f"未知品牌：{brand}")
    cfg   = BRAND_CONFIG[brand]
    table = cfg["TABLE_NAME"]
    pg    = cfg["PGSQL_CONFIG"]

    input_dir  = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    excel_files = list(input_dir.glob("*.xlsx"))
    if not excel_files:
        print(f"⚠ 没找到任何 Excel 文件: {input_dir}")
        return

    blacklist_codes = _load_blacklist_codes(blacklist_excel_file)
    logger = _get_logger(f"price_compare.{brand}")
    logger.info(f"共发现 {len(excel_files)} 个输入文件 | 品牌: {brand} | 表: {table}")
    logger.info(
        f"差异阈值: {price_diff_threshold}% | 黑名单: {len(blacklist_codes)} 条 | "
        f"折扣系数: {taobao_discount} × {tax_factor} | 最低利润: {min_profit*100:.0f}%"
    )

    for f in excel_files:
        try:
            logger.info(f"\n[START] 处理文件: {f.name}")
            _compare_price_vs_db_single(
                file_path=str(f),
                output_dir=output_dir,
                table=table,
                pgsql_config=pg,
                suffix=suffix,
                price_diff_threshold=price_diff_threshold,
                blacklist_codes=blacklist_codes,
                taobao_discount=taobao_discount,
                tax_factor=tax_factor,
                min_profit=min_profit,
                logger=logger,
            )
        except Exception as e:
            logger.error(f"处理 {f.name} 失败: {e}")
            import traceback
            traceback.print_exc()

    logger.info(f"\n✅ 所有文件处理完成。输出路径: {output_dir}")


# ===== 新增：批量导出 SKU 库存 =====
def generate_stock_excels_bulk(
    brand: str,
    input_dir: str | Path,
    output_dir: str | Path,
    suffix: str = "_库存",
    in_stock_qty: int = 3,
    out_stock_qty: int = 0,
):
    """
    批量根据 input_dir 下的店铺导出表生成“SKUID | 调整后库存”的 Excel。
    规则与 generate_price_excels_bulk 保持一致：仅按 product_code 合并，全款同值。
    - 数据来源：<TABLE_NAME> 中的 taobao_store_price 字段：
        * '有货'  -> in_stock_qty（默认3）
        * 其他/空 -> out_stock_qty（默认0）
    - 输出：<输入文件名+suffix>.xlsx，只有两列：SKUID, 调整后库存
    """
    import pandas as pd
    import psycopg2
    from pathlib import Path

    def _to_text(v):
        if v is None:
            return ""
        s = str(v).strip()
        return s

    def _list_excels(input_dir: Path):
        return (list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xls")))

    def status_to_qty_from_price(val: str) -> int:
        # 按你要求：如果 taobao_store_price 字段文本为“有货” => 3，否则 => 0
        #（注意：如果该字段在某些品牌是“数字价格”，此映射会失真，需要改回 stock_status）
        s = _to_text(val)
        return in_stock_qty if s == "有货" else out_stock_qty

    brand = brand.lower().strip()
    if brand not in BRAND_CONFIG:
        raise ValueError(f"未知品牌：{brand}")
    cfg   = BRAND_CONFIG[brand]
    table = cfg["TABLE_NAME"]
    pg    = cfg["PGSQL_CONFIG"]

    input_dir  = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 1) 拉取“可售状态”并做映射（仅按 product_code）
    conn = psycopg2.connect(**pg)
    try:
        df_flag = pd.read_sql(f'SELECT product_code, taobao_store_price FROM {table}', conn)
    finally:
        conn.close()

    if "product_code" not in df_flag.columns:
        raise RuntimeError(f"{table} 缺少 product_code 列")

    df_flag["product_code"] = df_flag["product_code"].astype(str).str.strip()
    # 统一映射为数量
    df_flag["调整后库存"] = df_flag["taobao_store_price"].map(status_to_qty_from_price)
    # 按款聚合（若同一款多行，取最大——等价“只要有一行有货则有货”）
    qty_by_code = (
        df_flag.groupby("product_code")["调整后库存"].max()
               .reset_index()
    )

    # 2) 扫描输入目录
    files = _list_excels(input_dir)
    if not files:
        raise FileNotFoundError(f"目录没有找到 Excel：{input_dir}")

    results = []
    for f in sorted(files, key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            df0 = pd.read_excel(f, dtype=object)

            # 复用你现有的列名别名解析
            col_item = _normalize_col(df0, "item_id")        # 宝贝id
            col_code = _normalize_col(df0, "product_code")   # 商家编码
            col_sku  = _normalize_col(df0, "skuid")          # skuID

            # 展开 SKU 行（与价格导出一致：只按 product_code 合并，不看尺码）
            rows = []
            for _, r in df0.iterrows():
                item_id = _to_text(r.get(col_item))
                code    = _to_text(r.get(col_code))
                skus    = _split_skuids(r.get(col_sku))
                for sid in skus:
                    sid = _to_text(sid)
                    if sid:
                        rows.append((item_id, code, sid))

            if not rows:
                raise ValueError(f"{f.name} 无有效 SKU 记录（检查宝贝ID/商家编码/skuID）。")

            df_expanded = pd.DataFrame(rows, columns=["宝贝id", "product_code", "skuid"])
            df_expanded["product_code"] = df_expanded["product_code"].astype(str).str.strip()

            # 仅按 product_code 合并库存数量（与价格导出相同合并粒度）
            df_tmp = df_expanded.merge(qty_by_code, on="product_code", how="left")
            df_tmp["调整后库存"] = df_tmp["调整后库存"].fillna(out_stock_qty).astype(int)

            # 输出两列
            out_df = df_tmp[["skuid", "调整后库存"]]
            out_name = f.stem + (suffix or "")
            if not out_name.endswith(".xlsx"):
                out_name += ".xlsx"
            out_path = output_dir / out_name
            out_df.to_excel(out_path, index=False)

            print(f"✅ {f.name} -> {out_name} (rows={len(out_df)})")
            results.append((str(f), str(out_path), len(out_df), None))
        except Exception as e:
            print(f"❌ 处理失败：{f} | 错误：{e}")
            results.append((str(f), None, 0, str(e)))

    return results



if __name__ == "__main__":
    # 用法一：单文件（保持兼容）
    #   python this_script.py <brand> <input_dir> <output_excel_path>
    # 用法二：批量（推荐你现在的场景）
    #   python this_script.py <brand> <input_dir> <output_dir> --bulk [--suffix "_价格"]
    import sys, traceback
    try:
        if len(sys.argv) >= 5 and sys.argv[4] == "--bulk":
            brand = sys.argv[1]
            input_dir = sys.argv[2]
            output_dir = sys.argv[3]
            suffix = "_价格"
            if len(sys.argv) >= 7 and sys.argv[5] == "--suffix":
                suffix = sys.argv[6]
            results = generate_price_excels_bulk(brand, input_dir, output_dir, suffix=suffix, drop_rows_without_price=True)
            ok = [r for r in results if r[1] is not None]
            bad = [r for r in results if r[1] is None]
            print(f"📦 批量完成：成功 {len(ok)} 个，失败 {len(bad)} 个")
            if bad:
                print("失败清单：")
                for f, _, err in bad:
                    print(f" - {f} -> 错误：{err}")
        elif len(sys.argv) >= 4:
            generate_price_excel(sys.argv[1], sys.argv[2], sys.argv[3])
        else:
            print("用法：")
            print("  单文件：python this_script.py <brand> <input_dir> <output_excel_path>")
            print('  批  量：python this_script.py <brand> <input_dir> <output_dir> --bulk [--suffix "_价格"]')
    except Exception as e:
        print("❌ 失败：", e)
        traceback.print_exc()

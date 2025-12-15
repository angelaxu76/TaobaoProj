# generate_discount_excel.py
# -*- coding: utf-8 -*-

import os
import re
from typing import List, Tuple, Set

import openpyxl
import psycopg2

from config import BRAND_CONFIG, PGSQL_CONFIG


# =========================
# 固定模板参数
# =========================
TEMPLATE_SHEET_NAME = "Sheet1"
REQUIRED_HEADERS = ["商品Id", "商家编码", "打折", "减钱", "促销价"]


# =========================
# 工具函数
# =========================
def extract_codes_from_text(text) -> Set[str]:
    if text is None:
        return set()
    s = str(text).strip()
    if not s:
        return set()

    codes = set()
    codes |= set(re.findall(r"\b\d{5,}\b", s))
    for m in re.findall(r"\b[A-Za-z0-9][A-Za-z0-9_-]{4,}\b", s):
        if re.fullmatch(r"[A-Za-z]{5,}", m):
            continue
        codes.add(m)
    return codes


def load_blacklist_codes(blacklist_excel_path: str) -> Set[str]:
    if not blacklist_excel_path:
        return set()
    if not os.path.exists(blacklist_excel_path):
        raise FileNotFoundError(f"blacklist_excel 不存在: {blacklist_excel_path}")

    wb = openpyxl.load_workbook(blacklist_excel_path, data_only=True)
    out: Set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                out |= extract_codes_from_text(v)
    return out


def connect_pg():
    return psycopg2.connect(**PGSQL_CONFIG)


def fetch_inventory_rows(
    conn,
    table_name: str,
    product_code_col: str,
    price_col: str,
) -> List[Tuple[str, float]]:
    cur = conn.cursor()

    sql_list = [
        f"""
        SELECT {product_code_col}, {price_col}
        FROM {table_name}
        WHERE {product_code_col} IS NOT NULL
          AND {price_col} IS NOT NULL
        ORDER BY {product_code_col}, updated_at DESC, id DESC
        """,
        f"""
        SELECT {product_code_col}, {price_col}
        FROM {table_name}
        WHERE {product_code_col} IS NOT NULL
          AND {price_col} IS NOT NULL
        ORDER BY {product_code_col}
        """,
    ]

    last_err = None
    for sql in sql_list:
        try:
            cur.execute(sql)
            rows = cur.fetchall()
            cur.close()
            return rows
        except Exception as e:
            last_err = e
            conn.rollback()

    raise RuntimeError(f"查询 inventory 失败: {last_err}")


def dedup_keep_first(rows: List[Tuple[str, float]]) -> List[Tuple[str, float]]:
    seen = set()
    out = []
    for code, price in rows:
        if not code:
            continue
        code = str(code).strip()
        if code in seen:
            continue
        seen.add(code)
        out.append((code, price))
    return out


def locate_columns(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip()] = c

    for h in REQUIRED_HEADERS:
        if h not in headers:
            raise RuntimeError(f"模板表头缺失: {h}")

    return headers["商家编码"], headers["促销价"]


def clear_data_area(ws, max_cols=5, start_row=2):
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_cols + 1):
            ws.cell(row=r, column=c).value = None


def write_excel(output_path: str, data: List[Tuple[str, float]]):
    """
    不依赖模板文件，直接生成符合“打折导入宝贝模板”结构的 Excel：
    Sheet1
    表头：商品Id | 商家编码 | 打折 | 减钱 | 促销价
    仅填：商家编码、促销价，其它列留空
    """
    wb = openpyxl.Workbook()

    # openpyxl 默认会创建一个活动 sheet
    ws = wb.active
    ws.title = TEMPLATE_SHEET_NAME  # 严格按模板 sheet name

    # 写表头（严格按模板 title）
    ws.append(REQUIRED_HEADERS)

    # 写数据（只填两列：商家编码、促销价）
    # 列顺序：商品Id(空), 商家编码(填), 打折(空), 减钱(空), 促销价(填)
    for code, price in data:
        ws.append([None, str(code).strip(), None, None, float(price)])

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)



# =========================
# ✅ 对外唯一调用函数
# =========================
def generate_discount_excel(
    brand: str,
    output_excel_path: str,
    blacklist_excel_file: str = "",
) -> str:
    """
    从 inventory 表生成【打折导入】Excel
    - 只填：商家编码 / 促销价
    - 数据来源：config.BRAND_CONFIG
    """

    if brand not in BRAND_CONFIG:
        raise ValueError(f"未知品牌: {brand}")

    brand_cfg = BRAND_CONFIG[brand]

    table_name = brand_cfg["TABLE_NAME"]
    fields = brand_cfg["FIELDS"]

    product_code_col = fields["product_code"]
    price_col = "taobao_store_price"  # ✅ 业务字段，明确写死

    blacklist = load_blacklist_codes(blacklist_excel_file)

    conn = connect_pg()
    try:
        rows = fetch_inventory_rows(
            conn,
            table_name=table_name,
            product_code_col=product_code_col,
            price_col=price_col,
        )
    finally:
        conn.close()

    rows = dedup_keep_first(rows)

    if blacklist:
        rows = [(c, p) for c, p in rows if c not in blacklist]

    write_excel(output_excel_path, rows)
    return output_excel_path

# -*- coding: utf-8 -*-
"""
对齐后的“宝贝维度”价格校验（不需要 sku规格 / 尺码）

输入（淘宝导出 Excel）：
- 宝贝标题
- 宝贝ID
- 商家编码
- 一口价

数据库（*_inventory 表）：
- product_code
- jingya_untaxed_price
- taobao_store_price

输出（只导出有差异/有风险的）：
- 宝贝标题
- 宝贝id
- 商品编码
- 一口价(淘宝)
- taobao_store_price
- jingya_untaxed_price
- 价格偏离比例（淘宝 vs 系统）
- 倒挂安全比例（>0 才能卖）

倒挂规则：
到手价 = 一口价 * 0.85 * 0.9
要求：到手价 >= jingya_untaxed_price * 1.03
倒挂安全比例 = (到手价 - jingya_untaxed_price*1.03) / (jingya_untaxed_price*1.03)
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional, Tuple, List

import pandas as pd
import psycopg2

from config import BRAND_CONFIG, PGSQL_CONFIG


# -----------------------------
# Excel column resolver
# -----------------------------
def _find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = [str(c).strip() for c in df.columns]
    lower_map = {c.lower(): c for c in cols}
    for cand in candidates:
        c = str(cand).strip()
        if c in cols:
            return c
        lc = c.lower()
        if lc in lower_map:
            return lower_map[lc]
    return None


def _read_taobao_item_excel(taobao_excel_path: str) -> pd.DataFrame:
    df = pd.read_excel(taobao_excel_path)

    # ✅ 已按你提供的 Camper 店铺 Excel 对齐：宝贝标题/宝贝ID/商家编码/一口价
    col_title = _find_col(df, ["宝贝标题", "商品标题", "标题", "宝贝名称"])
    col_itemid = _find_col(df, ["宝贝ID", "宝贝id", "商品ID", "item_id"])
    col_code = _find_col(df, ["商家编码", "商品编码", "product_code", "Product Code", "货号"])
    col_price = _find_col(df, ["一口价", "一口价(淘宝)", "宝贝价格", "价格", "售价", "售卖价"])

    missing = []
    if not col_title: missing.append("宝贝标题")
    if not col_itemid: missing.append("宝贝ID")
    if not col_code: missing.append("商家编码/商品编码")
    if not col_price: missing.append("一口价")
    if missing:
        raise KeyError(f"淘宝 Excel 缺少列（或列名不在候选中）: {', '.join(missing)}")

    out = df[[col_title, col_itemid, col_code, col_price]].copy()
    out.columns = ["宝贝标题", "宝贝id", "商品编码", "一口价(淘宝)"]

    out["商品编码"] = out["商品编码"].fillna("").astype(str).str.strip()
    out = out[out["商品编码"] != ""].copy()

    out["一口价(淘宝)"] = pd.to_numeric(out["一口价(淘宝)"], errors="coerce")
    out = out.dropna(subset=["一口价(淘宝)"]).copy()

    out["宝贝标题"] = out["宝贝标题"].fillna("").astype(str).str.strip()
    out["宝贝id"] = out["宝贝id"].fillna("").astype(str).str.strip()

    return out


# -----------------------------
# DB access
# -----------------------------
def _get_pg_conn_and_table(brand: str) -> Tuple["psycopg2.extensions.connection", str]:
    brand_l = brand.lower().strip()
    if brand_l not in BRAND_CONFIG:
        raise ValueError(f"未知品牌: {brand}（BRAND_CONFIG 里没有该 key）")

    cfg = BRAND_CONFIG[brand_l]
    table_name = cfg["TABLE_NAME"]
    pgcfg = cfg.get("PGSQL_CONFIG", PGSQL_CONFIG)

    # 兼容 database/dbname 两种写法
    dbname = pgcfg.get("dbname") or pgcfg.get("database")
    if not dbname:
        raise KeyError(f"PGSQL_CONFIG 缺少 dbname/database（brand={brand}）")

    conn = psycopg2.connect(
        host=pgcfg["host"],
        port=pgcfg["port"],
        user=pgcfg["user"],
        password=pgcfg["password"],
        dbname=dbname,
    )
    return conn, table_name


def _fetch_item_level_prices_from_db(brand: str) -> pd.DataFrame:
    """
    只取商品编码维度（不按尺码/SKU）
    - jingya_untaxed_price/taobao_store_price：按 product_code 聚合取 MAX（更保守）
    """
    conn, table_name = _get_pg_conn_and_table(brand)

    sql = f"""
        SELECT
            product_code,
            jingya_untaxed_price,
            taobao_store_price
        FROM {table_name}
        WHERE product_code IS NOT NULL
          AND TRIM(product_code) <> ''
    """
    df = pd.read_sql(sql, conn)
    conn.close()

    df["product_code"] = df["product_code"].astype(str).str.strip()
    df["jingya_untaxed_price"] = pd.to_numeric(df["jingya_untaxed_price"], errors="coerce")
    df["taobao_store_price"] = pd.to_numeric(df["taobao_store_price"], errors="coerce")

    agg = (
        df.groupby("product_code", as_index=False)
          .agg(
              jingya_untaxed_price=("jingya_untaxed_price", "max"),
              taobao_store_price=("taobao_store_price", "max"),
          )
    )
    return agg


# -----------------------------
# Excel styling (optional highlight)
# -----------------------------
def _export_with_highlight(df: pd.DataFrame, output_path: str) -> str:
    output_path = str(Path(output_path))
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)

    # 倒挂行高亮（不增加额外列）
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font

        wb = load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        inv_col = None
        for i, h in enumerate(headers, start=1):
            if str(h).strip() == "倒挂安全比例（>0 才能卖）":
                inv_col = i
                break

        if inv_col:
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            bold = Font(bold=True)
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=inv_col).value
                try:
                    v_num = float(v)
                except Exception:
                    continue
                if v_num <= 0:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill
                    ws.cell(row=r, column=inv_col).font = bold

        wb.save(output_path)
    except Exception:
        pass

    return output_path


# -----------------------------
# Public API
# -----------------------------
def check_taobao_item_price_diff(
    brand: str,
    taobao_excel_path: str,
    output_report_path: str,
    taobao_discount: float = 0.85,
    tax_factor: float = 0.9,
    min_profit: float = 0.03,
    deviation_tolerance: float = 0.001,
) -> str:
    """
    只导出“有问题”的记录：
    - DB 缺失（编码在 DB 找不到 或 DB 价为空）
    - 价格偏离比例超出容差
    - 倒挂安全比例 <= 0
    """
    df_tb = _read_taobao_item_excel(taobao_excel_path)
    df_db = _fetch_item_level_prices_from_db(brand)

    merged = df_tb.merge(df_db, left_on="商品编码", right_on="product_code", how="left")
    merged.drop(columns=["product_code"], inplace=True)

    # 主指标1：价格偏离比例（淘宝 vs 系统）= (一口价 - store) / store
    merged["价格偏离比例（淘宝 vs 系统）"] = (
        (merged["一口价(淘宝)"] - merged["taobao_store_price"]) / merged["taobao_store_price"]
    )

    # 到手价
    net = merged["一口价(淘宝)"] * taobao_discount * tax_factor
    safe_line = merged["jingya_untaxed_price"] * (1 + min_profit)

    # 主指标2：倒挂安全比例（>0 才能卖）
    merged["倒挂安全比例（>0 才能卖）"] = (net - safe_line) / safe_line

    # 过滤：只导出有差异/风险的
    db_missing = merged["jingya_untaxed_price"].isna() | merged["taobao_store_price"].isna()
    deviation_bad = merged["价格偏离比例（淘宝 vs 系统）"].abs() > deviation_tolerance
    inverted = merged["倒挂安全比例（>0 才能卖）"] <= 0

    out = merged[db_missing | deviation_bad | inverted].copy()

    # 最终输出列（你指定 + 两个主指标）
    out = out[
        [
            "宝贝标题",
            "宝贝id",
            "商品编码",
            "一口价(淘宝)",
            "taobao_store_price",
            "jingya_untaxed_price",
            "价格偏离比例（淘宝 vs 系统）",
            "倒挂安全比例（>0 才能卖）",
        ]
    ].copy()

    return _export_with_highlight(out, output_report_path)

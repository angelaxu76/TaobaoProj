# export_channel_price_excel_jingya.py
"""
通用导出：鲸芽 渠道价格更新 Excel（不读模板；固定 sheet 与表头；分文件）
- 函数签名不变：export_jiangya_channel_prices(brand: str, output_dir: Optional[str] = None) -> str
- 仅导出 channel_product_id 非空；每个渠道商品一行（按 channel_product_id 聚合）
- Base Price = min(original_price_gbp, discount_price_gbp)（存在者择其一）× BRAND_DISCOUNT[brand]（默认 1.0）
- 跳过下架/无价：Base Price 非法（NaN/<=0）不写入 Excel
- 定价：price_utils.calculate_jingya_prices(base_price, delivery_cost=7, exchange_rate=9.7)
    渠道价格(未税)(元)(必填) ← untaxed
    最低建议零售价(元)       ← retail
    最高建议零售价(元)       ← retail
- SKU ID 固定写 0
- 分包写出：每个文件最多 480 条数据行（不含表头），文件名末尾附 part 序号
"""

from pathlib import Path
from typing import Optional, List, Tuple
import math

import pandas as pd
import openpyxl
import psycopg2

from config import BRAND_CONFIG,BRAND_DISCOUNT
try:
    from config import PGSQL_CONFIG  # 兜底
except Exception:
    PGSQL_CONFIG = {}

# 价格工具
try:
    from common.pricing.price_utils import calculate_jingya_prices
except Exception:
    # 若你的工程里在其它路径，可替换为实际导入
    from common.pricing.price_utils import calculate_jingya_prices  # type: ignore


# 固定 sheet 与表头
SHEET_NAME = "sheet1"
HEADERS = [
    "渠道产品ID",
    "SKU ID(不存在或者设置品价格时,sku填写0)",
    "渠道价格(未税)(元)(必填)",
    "最低建议零售价(元)",
    "最高建议零售价(元)",
]


def _brand_discount(brand: str) -> float:
    return float(BRAND_DISCOUNT.get(brand.lower().strip(), 1.0))

def _to_float_safe(x) -> float:
    try:
        v = float(x)
        if math.isnan(v) or math.isinf(v):
            return 0.0
        return v
    except Exception:
        return 0.0

def _compute_base_price(row: pd.Series, brand: str) -> float:
    """Base = min(original, discount)（存在者择其一）× 品牌折扣；无值则 0"""
    o = _to_float_safe(row.get("original_price_gbp"))
    d = _to_float_safe(row.get("discount_price_gbp"))
    if o > 0 and d > 0:
        base_raw = min(o, d)
    else:
        base_raw = d if d > 0 else o
    return base_raw * _brand_discount(brand)

def _is_valid_price(x) -> bool:
    try:
        v = float(x)
        return (not math.isnan(v)) and (not math.isinf(v)) and (v > 0)
    except Exception:
        return False

def _write_one_excel(df_chunk: pd.DataFrame, file_path: Path):
    """按固定 sheet 和表头写一个 Excel（不依赖模板）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    # 表头
    for c_idx, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c_idx, value=h)
    # 数据
    for r_idx, row in enumerate(df_chunk.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file_path)
    wb.close()

def export_jiangya_channel_prices(
    brand: str,
    output_dir: Optional[str] = None,
    exclude_excel_file: Optional[str] = None,
    chunk_size: int = 400,
    exchange_rate: float = 9.6,
    delivery_cost: float = 7,
    untaxed_margin: float = 1.13,
    retail_margin: float = 1.43,
) -> str:
    """
    通用鲸芽渠道价格导出。
    新增功能：
    - 可传入 exclude_excel_file 指定Excel路径，读取商品编码列做排除。
    其他逻辑完全保持不变。
    """
    brand_l = brand.lower().strip()
    if brand_l not in BRAND_CONFIG:
        raise ValueError(f"未知品牌：{brand}")
    cfg = BRAND_CONFIG[brand_l]
    table = cfg["TABLE_NAME"]
    pgcfg = cfg.get("PGSQL_CONFIG", PGSQL_CONFIG)
    if not pgcfg:
        raise RuntimeError("PGSQL配置缺失")

    conn = psycopg2.connect(
        host=pgcfg["host"], port=pgcfg["port"],
        user=pgcfg["user"], password=pgcfg["password"], dbname=pgcfg["dbname"],
    )
    sql = f"""
        SELECT
            channel_product_id,
            product_code,
            original_price_gbp,
            discount_price_gbp,
            jingya_untaxed_price,
            taobao_store_price
        FROM {table}
        WHERE channel_product_id IS NOT NULL
          AND TRIM(channel_product_id) <> ''
    """
    df = pd.read_sql(sql, conn)

    # ── 诊断：统计数据库中商品的 channel_product_id 覆盖情况 ──
    sql_total = f"SELECT COUNT(DISTINCT product_code) FROM {table}"
    sql_with_id = f"""
        SELECT COUNT(DISTINCT product_code) FROM {table}
        WHERE channel_product_id IS NOT NULL AND TRIM(channel_product_id) <> ''
    """
    with conn.cursor() as cur:
        cur.execute(sql_total);   total_codes   = cur.fetchone()[0]
        cur.execute(sql_with_id); codes_with_id = cur.fetchone()[0]
    print(f"[诊断] 数据库中总商品编码数：{total_codes}")
    print(f"[诊断] 有 channel_product_id 的商品编码数：{codes_with_id}")
    print(f"[诊断] 缺少 channel_product_id（未导入鲸芽ID）：{total_codes - codes_with_id}")

    conn.close()

    out_dir = Path(output_dir) if output_dir else Path(cfg["OUTPUT_DIR"])
    out_dir.mkdir(parents=True, exist_ok=True)
    if df.empty:
        out_file = out_dir / f"{brand_l}_jiangya_price_update_part1_of1.xlsx"
        _write_one_excel(pd.DataFrame(columns=HEADERS), out_file)
        print("[INFO] 无可导出的记录。")
        return str(out_file)

    df_grp = df.groupby("channel_product_id", dropna=False).agg({
        "product_code": "first",
        "original_price_gbp": "first",
        "discount_price_gbp": "first",
        "jingya_untaxed_price": "first",
        "taobao_store_price": "first",
    }).reset_index()

    # Excel黑名单过滤（可选参数）
    blacklist_codes = set()
    if exclude_excel_file:
        blacklist_codes = set(_load_exclude_codes_from_excel(Path(exclude_excel_file)))
    if blacklist_codes:
        before = len(df_grp)
        df_grp = df_grp[~df_grp["product_code"].astype(str).str.strip().str.upper().isin(blacklist_codes)]
        print(f"[INFO] Excel排除清单过滤 {before - len(df_grp)} 行。")

    def _valid_num(x):
        try:
            v = float(x)
            return v > 0 and not math.isnan(v)
        except:
            return False

    def _pick_prices(row):
        db_untaxed, db_retail = row.get("jingya_untaxed_price"), row.get("taobao_store_price")
        if _valid_num(db_untaxed) and _valid_num(db_retail):
            return float(db_untaxed), float(db_retail)
        o, d = _to_float_safe(row.get("original_price_gbp")), _to_float_safe(row.get("discount_price_gbp"))
        base_raw = min(o, d) if o > 0 and d > 0 else (d if d > 0 else o)
        base = base_raw * _brand_discount(brand_l)
        if not _is_valid_price(base):
            return 0.0, 0.0
        return calculate_jingya_prices(base, delivery_cost=delivery_cost, exchange_rate=exchange_rate, untaxed_margin=untaxed_margin, retail_margin=retail_margin)

    prices = df_grp.apply(_pick_prices, axis=1)
    expanded = prices.apply(pd.Series)
    expanded.columns = ["untaxed", "retail"]
    mask_valid = expanded.apply(lambda r: _valid_num(r["untaxed"]) and _valid_num(r["retail"]), axis=1)
    invalid_count = (~mask_valid).sum()
    if invalid_count:
        print(f"[诊断] 因价格无效被剔除：{invalid_count} 条（价格为0或NaN）")
    df_grp, expanded = df_grp[mask_valid], expanded[mask_valid]
    print(f"[诊断] 最终可导出渠道产品数：{len(df_grp)}")

    out_df = pd.DataFrame({
        "渠道产品ID": df_grp["channel_product_id"],
        "SKU ID(不存在或者设置品价格时,sku填写0)": 0,
        "渠道价格(未税)(元)(必填)": expanded["untaxed"].astype(int),
        "最低建议零售价(元)": expanded["retail"].astype(int),
        "最高建议零售价(元)": expanded["retail"].astype(int),
    })[HEADERS]

    num_parts = (len(out_df) + chunk_size - 1) // chunk_size
    for i in range(num_parts):
        start, end = i * chunk_size, min((i + 1) * chunk_size, len(out_df))
        part = out_df.iloc[start:end]
        out_file = out_dir / f"{brand_l}_jiangya_price_update_part{i+1}_of_{num_parts}.xlsx"
        _write_one_excel(part, out_file)
        print(f"[OK] 写出 {out_file}")
    return str(out_file)







SHEET_NAME_PRICE = "sheet1"
HEADERS_PRICE = ["渠道产品ID(必填)", "skuID", "渠道价格(未税)(元)(必填)", "最低建议零售价(元)", "最高建议零售价(元)"]

def _write_simple_excel(df: pd.DataFrame, file_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME_PRICE
    # 表头
    for c, h in enumerate(HEADERS_PRICE, start=1):
        ws.cell(row=1, column=c, value=h)
    # 数据
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file_path)
    wb.close()

def _load_exclude_codes(file_path: Path) -> List[str]:
    codes = []
    if not file_path.exists():
        print(f"[WARN] 排除清单未找到：{file_path}（将不做排除）")
        return codes
    for line in file_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        codes.append(s.upper())
    print(f"[INFO] 已加载排除编码 {len(codes)} 条。")
    return codes

def _load_exclude_codes_from_excel(excel_path: Path) -> List[str]:
    """
    从 Excel 读取要排除的商品编码列表。
    逻辑：
    - 找一列列名里包含 "code"（不区分大小写）或包含 "编码"
      例: "product_code", "Product Code", "商品编码", "编码"
    - 读取这一列的值，去前后空格
    - 转成大写，过滤空值 / NaN / 空字符串
    - 返回list[str]
    如果 Excel 不存在或无法解析，返回空列表。
    """
    codes: List[str] = []
    if not excel_path.exists():
        print(f"[INFO] 未找到Excel排除清单文件：{excel_path}")
        return codes

    try:
        df_exc = pd.read_excel(excel_path)
    except Exception as e:
        print(f"[WARN] 无法读取Excel排除清单 {excel_path}: {e}")
        return codes

    # 自动识别“商品编码列”
    candidate_cols = []
    for col in df_exc.columns:
        col_str = str(col)
        low = col_str.lower().strip()
        if ("code" in low) or ("编码" in col_str):
            candidate_cols.append(col)

    if not candidate_cols:
        print(f"[WARN] Excel文件中未找到包含 'code' 或 '编码' 的列，跳过。列名={list(df_exc.columns)}")
        return codes

    col_use = candidate_cols[0]
    raw_vals = (
        df_exc[col_use]
        .astype(str)
        .map(lambda s: s.strip())
        .tolist()
    )

    for v in raw_vals:
        up = v.upper()
        if up and up != "NAN":
            codes.append(up)

    print(f"[INFO] 从Excel排除清单读取 {len(codes)} 条商品编码（原始 {len(raw_vals)} 条）。")
    return codes


import os
import math
import pandas as pd
from pathlib import Path
from sqlalchemy import create_engine, text
from config import PGSQL_CONFIG, BRAND_CONFIG

def export_channel_price_by_sku(
    brand: str,
    output_excel_path: str,
    exclude_excel_file: str = None,
    chunk_size: int = 200,
    filter_txt_file: str = None,
):
    """
    导出用于鲸芽SKU级别价格更新的Excel

    功能：
    1. 读取数据库中当前品牌(brand)的库存/价格明细（按SKU）。
    2. 支持用 exclude_excel_file (黑名单Excel) 排除不想导出的商品编码。
    3. ✅ 新增：支持用 filter_txt_file (TXT列表) 只导出指定的 channel_product_id。
    4. 结果按 chunk_size 分批写入多个 Excel 文件，方便导入鲸芽。

    参数：
    - brand: 例如 "barbour"
    - output_excel_path: 输出文件路径前缀或目录
        例: r"D:\TB\Products\barbour\repulibcation\publication_sku_prices\sku_level_prices"
        会生成 sku_level_prices_part1.xlsx, sku_level_prices_part2.xlsx ...
    - exclude_excel_file: Excel黑名单路径，可选
        里面包含需要排除的商品编码(我们会尝试读取 _code / product_code / 商品编码 这些列)
    - chunk_size: 每个Excel最多写入多少行
    - filter_txt_file: TXT白名单路径，可选
        TXT每一行是一个 channel_product_id，只导出这些ID的记录
    """

    # ============================================================
    # 0. 准备输出目录
    # ============================================================
    out_path = Path(output_excel_path)
    out_dir = out_path.parent if out_path.suffix == "" else out_path.parent
    if not out_dir.exists():
        out_dir.mkdir(parents=True, exist_ok=True)

    # 最终文件名前缀（不带扩展名）
    if out_path.suffix.lower() in [".xlsx", ".xls"]:
        base_name = out_path.with_suffix("").name
    else:
        base_name = out_path.name  # 例如 "sku_level_prices"

    # ============================================================
    # 1. 从数据库读取SKU级别信息
    #    这里的 SQL 需要跟你的 barbour_inventory / offers 结构一致
    #    下面 SQL 是示例，请和你线上已有的SQL对齐
    # ============================================================
    conn_url = (
        f"postgresql+psycopg2://{PGSQL_CONFIG['user']}:"
        f"{PGSQL_CONFIG['password']}@{PGSQL_CONFIG['host']}:"
        f"{PGSQL_CONFIG['port']}/{PGSQL_CONFIG['database']}"
    )
    engine = create_engine(conn_url)

    # 这个SQL假设 barbour_inventory 里已经 merge 了供货商和价格信息
    # 如果你原始实现不同，请把你原来的 SELECT 整段粘回来覆盖这里就行，
    # 其他过滤逻辑可以保持不变
    sql = f"""
        SELECT
            channel_product_id,
            product_name       AS product_code,
            size               AS sku_size,
            supplier_name,
            stock_qty,
            jingya_untaxed_price,
            taobao_store_price
        FROM {brand}_inventory
        WHERE channel_product_id IS NOT NULL
          AND TRIM(channel_product_id) <> ''
    """

    df = pd.read_sql(text(sql), engine)

    # 正常来说我们希望统一列名用于输出
    df = df.rename(
        columns={
            "product_name": "product_code",  # 双保险
            "size": "sku_size",
        }
    )

    # ============================================================
    # 2. 读取并应用黑名单 exclude_excel_file
    #    把这些商品编码(product_code) 完全排除
    # ============================================================
    if exclude_excel_file and os.path.exists(exclude_excel_file):
        try:
            excl_df = pd.read_excel(exclude_excel_file)
        except Exception:
            excl_df = pd.DataFrame()

        # 支持多种列名，尽量兼容
        exclude_codes = set()

        for col_candidate in ["_code", "product_code", "商品编码", "货品ID", "货品id", "货品Id"]:
            if col_candidate in excl_df.columns:
                tmp = (
                    excl_df[col_candidate]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )
                exclude_codes.update(tmp.tolist())

        if exclude_codes:
            before_rows = len(df)
            df = df[~df["product_code"].astype(str).isin(exclude_codes)]
            after_rows = len(df)
            print(f"⛔ 黑名单过滤: {before_rows} → {after_rows} 行 (排除了 {before_rows - after_rows} 行)")

    # ============================================================
    # 3. ✅ 读取并应用白名单 filter_txt_file
    #    只保留这些 channel_product_id
    # ============================================================
    if filter_txt_file and os.path.exists(filter_txt_file):
        with open(filter_txt_file, "r", encoding="utf-8") as f:
            wanted_ids = [line.strip() for line in f if line.strip()]
        wanted_set = set(wanted_ids)

        before_rows = len(df)
        df = df[df["channel_product_id"].astype(str).isin(wanted_set)]
        after_rows = len(df)
        print(f"📋 白名单过滤: {before_rows} → {after_rows} 行 (仅保留 {after_rows} 行, {len(wanted_set)} 个ID)")

    # ============================================================
    # 4. 排序、清洗输出列
    # ============================================================
    # 我们定义输出列顺序，方便鲸芽导入/校对
    output_cols = [
        "channel_product_id",
        "product_code",
        "sku_size",
        "supplier_name",
        "stock_qty",
        "jingya_untaxed_price",
        "taobao_store_price",
    ]

    # 只保留我们需要的列，如果缺列就先加空列
    for col in output_cols:
        if col not in df.columns:
            df[col] = ""

    df = df[output_cols].copy()

    # 去掉明显空的一些SKU，避免写一堆空行
    df = df[df["product_code"].fillna("").astype(str).str.strip() != ""].copy()

    # 给出总数
    total_rows = len(df)
    print(f"📦 最终可导出行数: {total_rows}")

    if total_rows == 0:
        print("⚠ 没有可导出的数据，已跳过写Excel。")
        return

    # ============================================================
    # 5. 分chunk写多个 Excel
    # ============================================================
    num_parts = math.ceil(total_rows / chunk_size)

    for part_idx in range(num_parts):
        start = part_idx * chunk_size
        end = min(start + chunk_size, total_rows)
        chunk_df = df.iloc[start:end].copy()

        part_no = part_idx + 1
        out_file = out_dir / f"{base_name}_part{part_no}.xlsx"

        # 写 Excel
        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            sheet_name = f"{brand}_sku_price"
            chunk_df.to_excel(writer, index=False, sheet_name=sheet_name)

        print(f"✅ 写入: {out_file} [{start}:{end}] 共 {len(chunk_df)} 行")

    print("🎉 SKU价格导出完成。")




# CLI（可选）
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="导出鲸芽渠道价格更新 Excel（固定表头/分文件）")
    parser.add_argument("--brand", required=True, help="品牌名，例如 camper / clarks / geox / barbour")
    parser.add_argument("--output-dir", default=None, help="可选，导出目录（默认 BRAND_CONFIG[brand]['OUTPUT_DIR']）")
    args = parser.parse_args()
    path = export_jiangya_channel_prices(args.brand, args.output_dir)
    print("[OK]", path)

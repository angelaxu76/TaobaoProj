# jiangya_export_stockcount_to_excel.py
"""
å¯¼å‡ºé²¸èŠ½åº“å­˜æ›´æ–° Excelï¼ˆå›ºå®š sheet/è¡¨å¤´ï¼›åˆ†æ–‡ä»¶ 490 è¡Œ/æ–‡ä»¶ï¼‰
- å‡½æ•°ç­¾åä¿æŒä¸å˜ï¼šexport_stock_excel(brand: str, output_dir: Optional[str] = None, filename: Optional[str] = None) -> str
- ä»…å¯¼å‡º channel_product_id éç©ºï¼ˆæ’é™¤ NULL/ç©ºä¸²ï¼‰
- å­—æ®µï¼šchannel_product_id, skuid, stock_count
- sheet åå›ºå®šï¼šsheet1
- è¡¨å¤´å›ºå®šï¼š['æ¸ é“äº§å“ID(å¿…å¡«)', 'skuID', 'åº“å­˜(å¿…å¡«)']
- åˆ†åŒ…ï¼šæ¯ä¸ªæ–‡ä»¶æœ€å¤š 490 æ¡â€œæ•°æ®è¡Œâ€ï¼ˆä¸å«è¡¨å¤´ï¼‰

ä¾èµ–ï¼špandas openpyxl psycopg2-binary
"""

from pathlib import Path
from typing import Optional, List

import pandas as pd
import openpyxl
import psycopg2

from config import BRAND_CONFIG
try:
    from config import PGSQL_CONFIG  # å…œåº•
except Exception:
    PGSQL_CONFIG = {}

SHEET_NAME = "sheet1"
HEADERS = ["æ¸ é“äº§å“ID(å¿…å¡«)", "skuID", "åº“å­˜(å¿…å¡«)"]
CHUNK_SIZE = 490  # æ¯ä¸ªæ–‡ä»¶æœ€å¤šçš„æ•°æ®è¡Œæ•°ï¼ˆä¸å«è¡¨å¤´ï¼‰

def _write_one_excel(df_chunk: pd.DataFrame, file_path: Path):
    """æŒ‰å›ºå®š sheet ä¸è¡¨å¤´å†™ä¸€ä¸ª Excelï¼ˆä¸ä¾èµ–æ¨¡æ¿ï¼‰"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    # è¡¨å¤´
    for c_idx, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c_idx, value=h)
    # æ•°æ®
    for r_idx, row in enumerate(df_chunk.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file_path)
    wb.close()

def export_stock_excel(brand: str, output_dir: Optional[str] = None, filename: Optional[str] = None) -> str:
    brand_l = brand.lower().strip()
    if brand_l not in BRAND_CONFIG:
        raise ValueError(f"æœªçŸ¥å“ç‰Œï¼š{brand}ã€‚å¯ç”¨ï¼š{', '.join(sorted(BRAND_CONFIG.keys()))}")

    cfg = BRAND_CONFIG[brand_l]
    table = cfg["TABLE_NAME"]
    pgcfg = cfg.get("PGSQL_CONFIG", PGSQL_CONFIG)
    if not pgcfg:
        raise RuntimeError("PGSQL è¿æ¥é…ç½®ç¼ºå¤±ï¼Œè¯·åœ¨ config.py ä¸­æä¾› PGSQL_CONFIG æˆ–å“ç‰Œçº§ PGSQL_CONFIGã€‚")

    # 1) è¯»å–æ•°æ®ï¼šä»… channel_product_id éç©º/éç©ºä¸²
    conn = psycopg2.connect(
        host=pgcfg["host"], port=pgcfg["port"],
        user=pgcfg["user"], password=pgcfg["password"], dbname=pgcfg["dbname"],
    )
    sql = f"""
        SELECT channel_product_id, skuid, stock_count
        FROM {table}
        WHERE channel_product_id IS NOT NULL AND TRIM(channel_product_id) <> ''
    """
    df = pd.read_sql(sql, conn)
    conn.close()

    # 2) æ¸…æ´—
    if df.empty:
        # ä»ç”Ÿæˆä¸€ä¸ªåªæœ‰è¡¨å¤´çš„ç©ºæ–‡ä»¶
        out_dir = Path(output_dir) if output_dir else Path(cfg["OUTPUT_DIR"])
        out_dir.mkdir(parents=True, exist_ok=True)
        base = Path(filename).stem if filename else f"{brand_l}_jiangya_stock_update"
        out_file = out_dir / f"{base}_part1_of1.xlsx"
        empty = pd.DataFrame(columns=HEADERS)
        _write_one_excel(empty, out_file)
        print(f"[INFO] æ— å¯å¯¼å‡ºçš„è®°å½•ï¼Œç”Ÿæˆç©ºè¡¨ï¼š{out_file}")
        return str(out_file)

    df["channel_product_id"] = df["channel_product_id"].astype(str).str.strip()
    df["skuid"] = df["skuid"].astype(str).fillna("0").replace({"": "0"})
    df["stock_count"] = pd.to_numeric(df["stock_count"], errors="coerce").fillna(0).astype(int)

    out_df = pd.DataFrame({
        "æ¸ é“äº§å“ID(å¿…å¡«)": df["channel_product_id"],
        "skuID": df["skuid"],
        "åº“å­˜(å¿…å¡«)": df["stock_count"],
    })[HEADERS]

    # ğŸ‘‰ åœ¨è¿™é‡ŒåŠ æ’åº
    out_df = out_df.sort_values(by=["æ¸ é“äº§å“ID(å¿…å¡«)", "skuID"]).reset_index(drop=True)

    # 3) åˆ†åŒ…å†™å‡ºï¼šæ¯æ–‡ä»¶æœ€å¤š 490 æ¡æ•°æ®è¡Œï¼ˆä¸å«è¡¨å¤´ï¼‰
    out_dir = Path(output_dir) if output_dir else Path(cfg["OUTPUT_DIR"])
    out_dir.mkdir(parents=True, exist_ok=True)

    n = len(out_df)
    num_parts = (n + CHUNK_SIZE - 1) // CHUNK_SIZE if n > 0 else 1
    created_files: List[Path] = []

    # è¾“å‡ºæ–‡ä»¶ååŸºåº•ï¼šå¦‚æœè°ƒç”¨è€…ä¼ äº† filename ä¸”åªæœ‰ 1 ä¸ªåˆ†åŒ…ï¼Œå°±ç”¨å®ƒï¼›å¦åˆ™ç»Ÿä¸€æ·»åŠ  part åºå·
    base = Path(filename).stem if filename else f"{brand_l}_jiangya_stock_update"

    if n == 0:
        out_file = out_dir / f"{base}_part1_of1.xlsx"
        _write_one_excel(out_df, out_file)
        print(f"[INFO] æ— æœ‰æ•ˆè®°å½•ï¼Œå·²ç”Ÿæˆç©ºè¡¨ï¼š{out_file}")
        return str(out_file)

    for i in range(num_parts):
        start = i * CHUNK_SIZE
        end = min(start + CHUNK_SIZE, n)
        df_chunk = out_df.iloc[start:end].reset_index(drop=True)

        if num_parts == 1 and filename:
            out_file = out_dir / (base + ".xlsx")
        else:
            out_file = out_dir / f"{base}_part{i+1}_of_{num_parts}.xlsx"

        _write_one_excel(df_chunk, out_file)
        created_files.append(out_file)
        print(f"[OK] å†™å‡ºï¼š{out_file}ï¼ˆè¡Œæ•°ï¼š{len(df_chunk)}ï¼‰")

    # è¿”å›ç¬¬ä¸€ä¸ªæ–‡ä»¶è·¯å¾„ï¼Œä¿æŒè¿”å›ç±»å‹ä¸å˜
    return str(created_files[0])


# CLI å…¥å£ï¼šbrand å¿…å¡«ï¼›--output-dir/--filename å¯é€‰
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="å¯¼å‡ºé²¸èŠ½åº“å­˜æ›´æ–° Excelï¼ˆå›ºå®š sheet/è¡¨å¤´ï¼›åˆ†æ–‡ä»¶ 490 è¡Œï¼‰")
    parser.add_argument("--brand", required=True, help="å“ç‰Œåï¼Œä¾‹å¦‚ camper / clarks_jingya / geox / barbour")
    parser.add_argument("--output-dir", default=None, help="å¯é€‰ï¼ŒæŒ‡å®šå¯¼å‡ºç›®å½•ï¼ˆé»˜è®¤ BRAND_CONFIG[brand]['OUTPUT_DIR']ï¼‰")
    parser.add_argument("--filename", default=None, help="å¯é€‰ï¼Œè¾“å‡ºæ–‡ä»¶ååŸºåº•ï¼›å•æ–‡ä»¶æ—¶å¯ç”¨ç²¾å‡†æ–‡ä»¶å")
    args = parser.parse_args()
    path = export_stock_excel(args.brand, args.output_dir, args.filename)
    print("[OK]", path)

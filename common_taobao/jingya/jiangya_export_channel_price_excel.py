# jiangya_export_channel_price_excel.py
"""
通用导出：鲸芽 渠道价格更新 Excel（不读模板；固定 sheet 与表头；分文件）
- 函数签名不变：export_jiangya_channel_prices(brand: str, output_dir: Optional[str] = None) -> str
- 仅导出 channel_product_id 非空；每个渠道商品一行（按 channel_product_id 聚合）
- Base Price = min(original_price_gbp, discount_price_gbp)（存在者择其一）× BRAND_DISCOUNT[brand]（默认 1.0）
- 跳过下架/无价：Base Price 非法（NaN/<=0）不写入 Excel
- 定价：price_utils.calculate_jingya_prices(base_price, delivery_cost=7, exchange_rate=9.7)
    渠道价格(未税)(元)(必填) ← untaxed
    最低建议零售价(元)       ← retail
    最高建议零售价(元)       ← retail
- SKU ID 固定写 0
- 分包写出：每个文件最多 480 条数据行（不含表头），文件名末尾附 part 序号
"""

from pathlib import Path
from typing import Optional, List, Tuple
import math

import pandas as pd
import openpyxl
import psycopg2

from config import BRAND_CONFIG
try:
    from config import PGSQL_CONFIG  # 兜底
except Exception:
    PGSQL_CONFIG = {}

# 价格工具
try:
    from common_taobao.core.price_utils import calculate_jingya_prices
except Exception:
    # 若你的工程里在其它路径，可替换为实际导入
    from common_taobao.core.price_utils import calculate_jingya_prices  # type: ignore


# 固定 sheet 与表头
SHEET_NAME = "sheet1"
HEADERS = [
    "渠道产品ID",
    "SKU ID(不存在或者设置品价格时,sku填写0)",
    "渠道价格(未税)(元)(必填)",
    "最低建议零售价(元)",
    "最高建议零售价(元)",
]

# 品牌默认折扣
BRAND_DISCOUNT = {
    "camper": 0.71,
    "geox": 0.98,
    "clarks_jingya": 1.0,
    # 其它品牌默认 1.0
}

def _brand_discount(brand: str) -> float:
    return float(BRAND_DISCOUNT.get(brand.lower().strip(), 1.0))

def _to_float_safe(x) -> float:
    try:
        v = float(x)
        if math.isnan(v) or math.isinf(v):
            return 0.0
        return v
    except Exception:
        return 0.0

def _compute_base_price(row: pd.Series, brand: str) -> float:
    """Base = min(original, discount)（存在者择其一）× 品牌折扣；无值则 0"""
    o = _to_float_safe(row.get("original_price_gbp"))
    d = _to_float_safe(row.get("discount_price_gbp"))
    if o > 0 and d > 0:
        base_raw = min(o, d)
    else:
        base_raw = d if d > 0 else o
    return base_raw * _brand_discount(brand)

def _is_valid_price(x) -> bool:
    try:
        v = float(x)
        return (not math.isnan(v)) and (not math.isinf(v)) and (v > 0)
    except Exception:
        return False

def _write_one_excel(df_chunk: pd.DataFrame, file_path: Path):
    """按固定 sheet 和表头写一个 Excel（不依赖模板）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    # 表头
    for c_idx, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c_idx, value=h)
    # 数据
    for r_idx, row in enumerate(df_chunk.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file_path)
    wb.close()

def export_jiangya_channel_prices(
    brand: str,
    output_dir: Optional[str] = None,
    exclude_excel_file: Optional[str] = None,  # ⭐ 新增参数
) -> str:
    """
    通用鲸芽渠道价格导出。
    新增功能：
    - 可传入 exclude_excel_file 指定Excel路径，读取商品编码列做排除。
    其他逻辑完全保持不变。
    """
    brand_l = brand.lower().strip()
    if brand_l not in BRAND_CONFIG:
        raise ValueError(f"未知品牌：{brand}")
    cfg = BRAND_CONFIG[brand_l]
    table = cfg["TABLE_NAME"]
    pgcfg = cfg.get("PGSQL_CONFIG", PGSQL_CONFIG)
    if not pgcfg:
        raise RuntimeError("PGSQL配置缺失")

    conn = psycopg2.connect(
        host=pgcfg["host"], port=pgcfg["port"],
        user=pgcfg["user"], password=pgcfg["password"], dbname=pgcfg["dbname"],
    )
    sql = f"""
        SELECT
            channel_product_id,
            product_code,
            original_price_gbp,
            discount_price_gbp,
            jingya_untaxed_price,
            taobao_store_price
        FROM {table}
        WHERE channel_product_id IS NOT NULL
          AND TRIM(channel_product_id) <> ''
    """
    df = pd.read_sql(sql, conn)
    conn.close()

    out_dir = Path(output_dir) if output_dir else Path(cfg["OUTPUT_DIR"])
    out_dir.mkdir(parents=True, exist_ok=True)
    if df.empty:
        out_file = out_dir / f"{brand_l}_jiangya_price_update_part1_of1.xlsx"
        _write_one_excel(pd.DataFrame(columns=HEADERS), out_file)
        print("[INFO] 无可导出的记录。")
        return str(out_file)

    df_grp = df.groupby("channel_product_id", dropna=False).agg({
        "product_code": "first",
        "original_price_gbp": "first",
        "discount_price_gbp": "first",
        "jingya_untaxed_price": "first",
        "taobao_store_price": "first",
    }).reset_index()

    # Excel黑名单过滤（可选参数）
    blacklist_codes = set()
    if exclude_excel_file:
        blacklist_codes = set(_load_exclude_codes_from_excel(Path(exclude_excel_file)))
    if blacklist_codes:
        before = len(df_grp)
        df_grp = df_grp[~df_grp["product_code"].astype(str).str.strip().str.upper().isin(blacklist_codes)]
        print(f"[INFO] Excel排除清单过滤 {before - len(df_grp)} 行。")

    def _valid_num(x):
        try:
            v = float(x)
            return v > 0 and not math.isnan(v)
        except:
            return False

    def _pick_prices(row):
        db_untaxed, db_retail = row.get("jingya_untaxed_price"), row.get("taobao_store_price")
        if _valid_num(db_untaxed) and _valid_num(db_retail):
            return float(db_untaxed), float(db_retail)
        o, d = _to_float_safe(row.get("original_price_gbp")), _to_float_safe(row.get("discount_price_gbp"))
        base_raw = min(o, d) if o > 0 and d > 0 else (d if d > 0 else o)
        base = base_raw * _brand_discount(brand_l)
        if not _is_valid_price(base):
            return 0.0, 0.0
        return calculate_jingya_prices(base, delivery_cost=7, exchange_rate=9.6)

    prices = df_grp.apply(_pick_prices, axis=1)
    expanded = prices.apply(pd.Series)
    expanded.columns = ["untaxed", "retail"]
    mask_valid = expanded.apply(lambda r: _valid_num(r["untaxed"]) and _valid_num(r["retail"]), axis=1)
    df_grp, expanded = df_grp[mask_valid], expanded[mask_valid]

    out_df = pd.DataFrame({
        "渠道产品ID": df_grp["channel_product_id"],
        "SKU ID(不存在或者设置品价格时,sku填写0)": 0,
        "渠道价格(未税)(元)(必填)": expanded["untaxed"].astype(int),
        "最低建议零售价(元)": expanded["retail"].astype(int),
        "最高建议零售价(元)": expanded["retail"].astype(int),
    })[HEADERS]

    chunk_size = 1000
    num_parts = (len(out_df) + chunk_size - 1) // chunk_size
    for i in range(num_parts):
        start, end = i * chunk_size, min((i + 1) * chunk_size, len(out_df))
        part = out_df.iloc[start:end]
        out_file = out_dir / f"{brand_l}_jiangya_price_update_part{i+1}_of_{num_parts}.xlsx"
        _write_one_excel(part, out_file)
        print(f"[OK] 写出 {out_file}")
    return str(out_file)







SHEET_NAME_PRICE = "sheet1"
HEADERS_PRICE = ["渠道产品ID(必填)", "skuID", "渠道价格(未税)(元)(必填)", "最低建议零售价(元)", "最高建议零售价(元)"]

def _write_simple_excel(df: pd.DataFrame, file_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME_PRICE
    # 表头
    for c, h in enumerate(HEADERS_PRICE, start=1):
        ws.cell(row=1, column=c, value=h)
    # 数据
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file_path)
    wb.close()

def _load_exclude_codes(file_path: Path) -> List[str]:
    codes = []
    if not file_path.exists():
        print(f"[WARN] 排除清单未找到：{file_path}（将不做排除）")
        return codes
    for line in file_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        codes.append(s.upper())
    print(f"[INFO] 已加载排除编码 {len(codes)} 条。")
    return codes

def _load_exclude_codes_from_excel(excel_path: Path) -> List[str]:
    """
    从 Excel 读取要排除的商品编码列表。
    逻辑：
    - 找一列列名里包含 "code"（不区分大小写）或包含 "编码"
      例: "product_code", "Product Code", "商品编码", "编码"
    - 读取这一列的值，去前后空格
    - 转成大写，过滤空值 / NaN / 空字符串
    - 返回list[str]
    如果 Excel 不存在或无法解析，返回空列表。
    """
    codes: List[str] = []
    if not excel_path.exists():
        print(f"[INFO] 未找到Excel排除清单文件：{excel_path}")
        return codes

    try:
        df_exc = pd.read_excel(excel_path)
    except Exception as e:
        print(f"[WARN] 无法读取Excel排除清单 {excel_path}: {e}")
        return codes

    # 自动识别“商品编码列”
    candidate_cols = []
    for col in df_exc.columns:
        col_str = str(col)
        low = col_str.lower().strip()
        if ("code" in low) or ("编码" in col_str):
            candidate_cols.append(col)

    if not candidate_cols:
        print(f"[WARN] Excel文件中未找到包含 'code' 或 '编码' 的列，跳过。列名={list(df_exc.columns)}")
        return codes

    col_use = candidate_cols[0]
    raw_vals = (
        df_exc[col_use]
        .astype(str)
        .map(lambda s: s.strip())
        .tolist()
    )

    for v in raw_vals:
        up = v.upper()
        if up and up != "NAN":
            codes.append(up)

    print(f"[INFO] 从Excel排除清单读取 {len(codes)} 条商品编码（原始 {len(raw_vals)} 条）。")
    return codes


def export_barbour_channel_price_by_sku(
    brand: str = "barbour",
    output_excel_path: str = "D:/TB/Products/barbour_channel_price.xlsx",
    exclude_excel_file: Optional[str] = None,
    chunk_size: Optional[int] = None,   # ⭐ 新增可选参数：行数分包上限
) -> str:
    """
    Barbour SKU 级价格导出（最终版）
    - 参数最简化，仅保留：
        brand：品牌名
        output_excel_path：输出文件的完整路径（绝对路径）
        exclude_excel_file：Excel黑名单路径（排除商品编码）
        chunk_size：可选参数；若传入，则按每 chunk_size 行切分为多个文件
    - 保留所有原有价格与SKU汇总逻辑
    """
    brand_l = brand.lower().strip()
    if brand_l not in BRAND_CONFIG:
        raise ValueError(f"未知品牌：{brand}。")

    cfg = BRAND_CONFIG[brand_l]
    table = cfg["TABLE_NAME"]
    pgcfg = cfg.get("PGSQL_CONFIG", PGSQL_CONFIG)
    if not pgcfg:
        raise RuntimeError("PGSQL 连接配置缺失，请检查 config.py。")

    # === 载入 Excel 黑名单 ===
    exclude_codes = set()
    if exclude_excel_file:
        exclude_codes = set(_load_exclude_codes_from_excel(Path(exclude_excel_file)))
        print(f"[INFO] 从Excel排除清单加载 {len(exclude_codes)} 条商品编码。")

    # === 查询数据库 ===
    conn = psycopg2.connect(
        host=pgcfg["host"], port=pgcfg["port"],
        user=pgcfg["user"], password=pgcfg["password"], dbname=pgcfg["dbname"],
    )
    sql = f"""
        SELECT
            channel_product_id,
            skuid,
            product_code,
            jingya_untaxed_price,
            taobao_store_price
        FROM {table}
        WHERE channel_product_id IS NOT NULL
          AND TRIM(channel_product_id) <> ''
    """
    df = pd.read_sql(sql, conn)
    conn.close()

    if df.empty:
        print(f"[INFO] 数据表 {table} 无可导出记录。")
        _write_simple_excel(pd.DataFrame(columns=HEADERS_PRICE), Path(output_excel_path))
        return output_excel_path

    # === 标准化字段 ===
    for col in ("channel_product_id", "skuid", "product_code"):
        df[col] = df[col].astype(str).str.strip()

    # === 黑名单过滤 ===
    if exclude_codes:
        before = len(df)
        df = df[~df["product_code"].str.upper().isin(exclude_codes)].reset_index(drop=True)
        print(f"[INFO] 已按Excel黑名单过滤 {before - len(df)} 行。")

    # === 转换数值类型 ===
    def to_num(s):
        try:
            return float(s)
        except Exception:
            return None
    df["jingya_untaxed_price"] = df["jingya_untaxed_price"].apply(to_num)
    df["taobao_store_price"] = df["taobao_store_price"].apply(to_num)

    # === 构造 skuID=0 的品价格行 ===
    df_use = df[["channel_product_id", "skuid", "jingya_untaxed_price", "taobao_store_price"]].copy()
    agg_max = (
        df_use.groupby("channel_product_id", as_index=False)
              .agg(max_j=("jingya_untaxed_price", "max"),
                   max_t=("taobao_store_price", "max"))
    )
    zero_rows = agg_max.assign(
        skuid="0",
        jingya_untaxed_price=lambda x: x["max_j"],
        taobao_store_price=lambda x: x["max_t"]
    )[["channel_product_id", "skuid", "jingya_untaxed_price", "taobao_store_price"]]
    detail_rows = df_use[df_use["skuid"] != "0"]

    df_final = pd.concat([zero_rows, detail_rows], ignore_index=True)
    df_final["zero_first"] = (df_final["skuid"] != "0").astype(int)
    df_final = df_final.sort_values(
        by=["channel_product_id", "zero_first", "skuid"]
    ).drop(columns=["zero_first"]).reset_index(drop=True)

    # === 生成输出表 ===
    out_df = pd.DataFrame({
        "渠道产品ID(必填)": df_final["channel_product_id"],
        "skuID": df_final["skuid"],
        "渠道价格(未税)(元)(必填)": df_final["jingya_untaxed_price"].round(2),
        "最低建议零售价(元)": df_final["taobao_store_price"].round(2),
        "最高建议零售价(元)": df_final["taobao_store_price"].round(2),
    })[HEADERS_PRICE]

    # === 输出 ===
    output_path = Path(output_excel_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ⭐ 根据 chunk_size 决定是否分包
    if chunk_size and len(out_df) > chunk_size:
        num_parts = (len(out_df) + chunk_size - 1) // chunk_size
        for i in range(num_parts):
            start, end = i * chunk_size, min((i + 1) * chunk_size, len(out_df))
            chunk = out_df.iloc[start:end]
            part_path = output_path.parent / f"{output_path.stem}_part{i+1}_of_{num_parts}.xlsx"
            _write_simple_excel(chunk, part_path)
            print(f"[OK] 写出分包文件：{part_path}")
        return str(output_path.parent)
    else:
        _write_simple_excel(out_df, output_path)
        print(f"[OK] 已生成完整文件：{output_path}（共 {len(out_df)} 行）")
        return str(output_path)




# CLI（可选）
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="导出鲸芽渠道价格更新 Excel（固定表头/分文件）")
    parser.add_argument("--brand", required=True, help="品牌名，例如 camper / clarks_jingya / geox / barbour")
    parser.add_argument("--output-dir", default=None, help="可选，导出目录（默认 BRAND_CONFIG[brand]['OUTPUT_DIR']）")
    args = parser.parse_args()
    path = export_jiangya_channel_prices(args.brand, args.output_dir)
    print("[OK]", path)

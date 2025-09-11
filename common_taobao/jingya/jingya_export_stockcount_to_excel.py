# jiangya_export_stockcount_to_excel.py
"""
导出鲸芽库存更新 Excel（固定 sheet/表头；分文件 490 行/文件）
- 函数签名保持不变：export_stock_excel(brand: str, output_dir: Optional[str] = None, filename: Optional[str] = None) -> str
- 仅导出 channel_product_id 非空（排除 NULL/空串）
- 字段：channel_product_id, skuid, stock_count
- sheet 名固定：sheet1
- 表头固定：['渠道产品ID(必填)', 'skuID', '库存(必填)']
- 分包：每个文件最多 490 条“数据行”（不含表头）

依赖：pandas openpyxl psycopg2-binary
"""

from pathlib import Path
from typing import Optional, List

import pandas as pd
import openpyxl
import psycopg2

from config import BRAND_CONFIG
try:
    from config import PGSQL_CONFIG  # 兜底
except Exception:
    PGSQL_CONFIG = {}

SHEET_NAME = "sheet1"
HEADERS = ["渠道产品ID(必填)", "skuID", "库存(必填)"]
CHUNK_SIZE = 490  # 每个文件最多的数据行数（不含表头）

def _write_one_excel(df_chunk: pd.DataFrame, file_path: Path):
    """按固定 sheet 与表头写一个 Excel（不依赖模板）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    # 表头
    for c_idx, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c_idx, value=h)
    # 数据
    for r_idx, row in enumerate(df_chunk.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file_path)
    wb.close()

def export_stock_excel(brand: str, output_dir: Optional[str] = None, filename: Optional[str] = None) -> str:
    brand_l = brand.lower().strip()
    if brand_l not in BRAND_CONFIG:
        raise ValueError(f"未知品牌：{brand}。可用：{', '.join(sorted(BRAND_CONFIG.keys()))}")

    cfg = BRAND_CONFIG[brand_l]
    table = cfg["TABLE_NAME"]
    pgcfg = cfg.get("PGSQL_CONFIG", PGSQL_CONFIG)
    if not pgcfg:
        raise RuntimeError("PGSQL 连接配置缺失，请在 config.py 中提供 PGSQL_CONFIG 或品牌级 PGSQL_CONFIG。")

    # 1) 读取数据：仅 channel_product_id 非空/非空串
    conn = psycopg2.connect(
        host=pgcfg["host"], port=pgcfg["port"],
        user=pgcfg["user"], password=pgcfg["password"], dbname=pgcfg["dbname"],
    )
    sql = f"""
        SELECT channel_product_id, skuid, stock_count
        FROM {table}
        WHERE channel_product_id IS NOT NULL AND TRIM(channel_product_id) <> ''
    """
    df = pd.read_sql(sql, conn)
    conn.close()

    # 2) 清洗
    if df.empty:
        # 仍生成一个只有表头的空文件
        out_dir = Path(output_dir) if output_dir else Path(cfg["OUTPUT_DIR"])
        out_dir.mkdir(parents=True, exist_ok=True)
        base = Path(filename).stem if filename else f"{brand_l}_jiangya_stock_update"
        out_file = out_dir / f"{base}_part1_of1.xlsx"
        empty = pd.DataFrame(columns=HEADERS)
        _write_one_excel(empty, out_file)
        print(f"[INFO] 无可导出的记录，生成空表：{out_file}")
        return str(out_file)

    df["channel_product_id"] = df["channel_product_id"].astype(str).str.strip()
    df["skuid"] = df["skuid"].astype(str).fillna("0").replace({"": "0"})
    df["stock_count"] = pd.to_numeric(df["stock_count"], errors="coerce").fillna(0).astype(int)

    out_df = pd.DataFrame({
        "渠道产品ID(必填)": df["channel_product_id"],
        "skuID": df["skuid"],
        "库存(必填)": df["stock_count"],
    })[HEADERS]

    # 3) 分包写出：每文件最多 490 条数据行（不含表头）
    out_dir = Path(output_dir) if output_dir else Path(cfg["OUTPUT_DIR"])
    out_dir.mkdir(parents=True, exist_ok=True)

    n = len(out_df)
    num_parts = (n + CHUNK_SIZE - 1) // CHUNK_SIZE if n > 0 else 1
    created_files: List[Path] = []

    # 输出文件名基底：如果调用者传了 filename 且只有 1 个分包，就用它；否则统一添加 part 序号
    base = Path(filename).stem if filename else f"{brand_l}_jiangya_stock_update"

    if n == 0:
        out_file = out_dir / f"{base}_part1_of1.xlsx"
        _write_one_excel(out_df, out_file)
        print(f"[INFO] 无有效记录，已生成空表：{out_file}")
        return str(out_file)

    for i in range(num_parts):
        start = i * CHUNK_SIZE
        end = min(start + CHUNK_SIZE, n)
        df_chunk = out_df.iloc[start:end].reset_index(drop=True)

        if num_parts == 1 and filename:
            out_file = out_dir / (base + ".xlsx")
        else:
            out_file = out_dir / f"{base}_part{i+1}_of_{num_parts}.xlsx"

        _write_one_excel(df_chunk, out_file)
        created_files.append(out_file)
        print(f"[OK] 写出：{out_file}（行数：{len(df_chunk)}）")

    # 返回第一个文件路径，保持返回类型不变
    return str(created_files[0])


# CLI 入口：brand 必填；--output-dir/--filename 可选
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="导出鲸芽库存更新 Excel（固定 sheet/表头；分文件 490 行）")
    parser.add_argument("--brand", required=True, help="品牌名，例如 camper / clarks_jingya / geox / barbour")
    parser.add_argument("--output-dir", default=None, help="可选，指定导出目录（默认 BRAND_CONFIG[brand]['OUTPUT_DIR']）")
    parser.add_argument("--filename", default=None, help="可选，输出文件名基底；单文件时可用精准文件名")
    args = parser.parse_args()
    path = export_stock_excel(args.brand, args.output_dir, args.filename)
    print("[OK]", path)

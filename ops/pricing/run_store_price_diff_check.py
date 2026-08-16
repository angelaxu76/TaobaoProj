# -*- coding: utf-8 -*-
# 同品牌、不同店铺间的一口价差异检查。
# 输入：GEI_SHARED/{品牌}/store_prices/ 下的所有店铺 xlsx（每个文件 = 一个店铺）
# 输出：桌面 store_price_diff_report/ 下的一个 xlsx，每个品牌一个 sheet
#
# 逻辑：
#   1. 每个店铺文件按 product_code 聚合出一个代表价（一口价）
#   2. 同一 product_code 在不同店铺间取 最高价 / 最低价
#   3. 差异百分比 = (最高价 - 最低价) / 最低价 * 100
#   4. 只输出差异百分比 > THRESHOLD 的商品，按差异百分比降序排列
#
# 修改下面 [运行参数] 部分后直接运行即可。

from pathlib import Path
from datetime import date

import pandas as pd

from cfg.paths import DESKTOP_DIR, GEI_EXPORT_BASE

# ========== 运行参数（按需修改）==========

GEI_SHARED = GEI_EXPORT_BASE

# 要处理的品牌（对应 GEI_SHARED/<品牌>/store_prices 目录）
BRANDS = ["barbour", "camper", "clarks", "ecco", "geox"]

# 差异百分比阈值，超过此值才输出
THRESHOLD = 2.0

# 输出报告路径（每个品牌一个 sheet）
OUTPUT_PATH = DESKTOP_DIR / "store_price_diff_report" / f"店铺一口价差异_{date.today():%Y%m%d}.xlsx"

# ========== 解析单个店铺 Excel ==========


def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _split_product_code(spec_val: object) -> str | None:
    """把 'MWX0339OL71,M' 这种 sku规格 值拆出 product_code（第一个逗号前的部分）。"""
    if not isinstance(spec_val, str):
        return None
    raw = spec_val.strip()
    if raw == "":
        return None
    parts = [p.strip() for p in raw.split(",")]
    if len(parts) < 2 or parts[0] == "" or parts[1] == "":
        return None
    return parts[0].upper()


def _to_text(v: object) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


def _to_float_safe(v: object) -> float | None:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_store_file(file_path: Path) -> pd.DataFrame:
    """解析单个店铺 Excel，返回 product_code 级别一口价：product_code | 宝贝标题 | 一口价"""
    df_raw = pd.read_excel(file_path, dtype=object)

    item_id_col = _find_col(df_raw, ["宝贝ID", "宝贝id", "宝贝Id", "item_id", "itemid", "商品ID"])
    title_col   = _find_col(df_raw, ["宝贝标题", "标题", "商品标题", "title"])
    price_col   = _find_col(df_raw, ["一口价", "售价", "销售价", "price"])
    spec_col    = _find_col(df_raw, ["sku规格", "SKU规格", "规格", "sku spec", "销售属性"])

    missing = [n for n, c in [("宝贝ID", item_id_col), ("一口价", price_col), ("sku规格", spec_col)] if c is None]
    if missing:
        raise ValueError(f"{file_path.name} 缺少必要列: {missing}，当前列: {list(df_raw.columns)}")

    df = df_raw.copy()
    df[item_id_col] = df[item_id_col].ffill()
    df[price_col]   = df[price_col].ffill()
    if title_col:
        df[title_col] = df[title_col].ffill()

    df["_product_code"] = df[spec_col].map(_split_product_code)
    df["_item_id_str"]  = df[item_id_col].map(_to_text)
    df["_price_float"]  = df[price_col].map(_to_float_safe)

    rows = []
    for item_id, gdf in df.groupby("_item_id_str", sort=False):
        if not item_id:
            continue
        codes = gdf["_product_code"].dropna()
        prices = gdf["_price_float"].dropna()
        if codes.empty or prices.empty:
            continue
        title = gdf[title_col].dropna().iloc[0] if (title_col and not gdf[title_col].dropna().empty) else None
        rows.append({
            "product_code": str(codes.iloc[0]).strip(),
            "宝贝标题": title,
            "一口价": float(prices.iloc[0]),
        })

    if not rows:
        return pd.DataFrame(columns=["product_code", "宝贝标题", "一口价"])

    df_products = pd.DataFrame(rows)
    # 同一店铺内同一编码出现多次（多个链接），取均价
    return df_products.groupby("product_code", as_index=False).agg({"宝贝标题": "first", "一口价": "mean"})


# ========== 单品牌：跨店铺比较 ==========


def _compare_brand(brand: str) -> pd.DataFrame | None:
    input_dir = GEI_SHARED / brand / "store_prices"
    if not input_dir.exists():
        print(f"  [SKIP] 输入目录不存在: {input_dir}")
        return None

    files = sorted(
        f for f in list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xls"))
        if not f.name.startswith("~$")
    )
    if len(files) < 2:
        print(f"  [SKIP] 店铺文件数量不足（{len(files)}个），跳过比较")
        return None

    store_series: dict[str, pd.Series] = {}
    titles: dict[str, object] = {}
    for f in files:
        store_name = f.stem
        try:
            df_p = _parse_store_file(f)
        except Exception as e:
            print(f"  [ERROR] 解析失败: {f.name} -> {e}")
            continue
        if df_p.empty:
            continue
        df_p = df_p.set_index("product_code")
        store_series[store_name] = df_p["一口价"]
        for code, title in df_p["宝贝标题"].items():
            if code not in titles and pd.notna(title):
                titles[code] = title
        print(f"  {store_name}: {len(df_p)} 个商品编码")

    if len(store_series) < 2:
        print(f"  [SKIP] 有效店铺不足 2 个，跳过比较")
        return None

    store_cols = list(store_series.keys())
    df_wide = pd.DataFrame(store_series)

    df_wide["有效店铺数"] = df_wide[store_cols].notna().sum(axis=1)
    df_wide = df_wide[df_wide["有效店铺数"] >= 2].copy()
    if df_wide.empty:
        return None

    df_wide["最高价"] = df_wide[store_cols].max(axis=1, skipna=True)
    df_wide["最低价"] = df_wide[store_cols].min(axis=1, skipna=True)
    df_wide["差异百分比"] = ((df_wide["最高价"] - df_wide["最低价"]) / df_wide["最低价"] * 100).round(2)

    df_diff = df_wide[df_wide["差异百分比"] > THRESHOLD].copy()
    if df_diff.empty:
        return None

    df_diff = df_diff.sort_values("差异百分比", ascending=False)
    df_diff.insert(0, "product_code", df_diff.index)
    df_diff.insert(1, "宝贝标题", df_diff["product_code"].map(titles))
    df_diff = df_diff.drop(columns=["有效店铺数"]).reset_index(drop=True)

    out_cols = ["product_code", "宝贝标题"] + store_cols + ["最高价", "最低价", "差异百分比"]
    return df_diff[out_cols]


# ========== 写多 sheet Excel ==========


def _write_report(brand_reports: dict[str, pd.DataFrame], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for brand, df in brand_reports.items():
            df.to_excel(writer, sheet_name=brand[:31], index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(out_path)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    warn_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws.freeze_panes = "A2"

        diff_col_idx = next((i + 1 for i, c in enumerate(ws[1]) if c.value == "差异百分比"), None)
        if diff_col_idx:
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=diff_col_idx).value
                try:
                    if float(v) >= 10:
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=r, column=c).fill = warn_fill
                except (TypeError, ValueError):
                    pass

        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    wb.save(out_path)


def run_all():
    print("=" * 60)
    print(f"  店铺间一口价差异检查（阈值 > {THRESHOLD}%）")
    print("=" * 60)

    brand_reports: dict[str, pd.DataFrame] = {}
    for brand in BRANDS:
        print(f"\n{'─'*50}")
        print(f"  品牌: {brand.upper()}")
        print(f"{'─'*50}")
        df = _compare_brand(brand)
        if df is None or df.empty:
            print(f"  无差异超过 {THRESHOLD}% 的商品")
            continue
        brand_reports[brand] = df
        print(f"  输出 {len(df)} 个商品编码")

    if not brand_reports:
        print("\n所有品牌均无差异超过阈值的商品，未生成报告。")
        return

    _write_report(brand_reports, OUTPUT_PATH)
    print(f"\n报告已生成: {OUTPUT_PATH}")


if __name__ == "__main__":
    run_all()

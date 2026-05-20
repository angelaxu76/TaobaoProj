"""
根据图片目录生成 codes.xlsx。

从指定目录读取所有 .jpg 文件，提取唯一商品编码（去掉 _front_N / _back_N 等后缀），
去重后写入 codes.xlsx 第一列。

支持两种命名格式：
  T420011D_front_1.jpg          → 编码 T420011D
  T420013D_BLACK_front_1.jpg    → 编码 T420013D_BLACK

用法：
  直接运行：python ops/linkfox/gen_codes_excel_from_images.py
  输出路径、来源目录在下方"运行参数"区块修改，或直接读 _session_config.py 默认值。
"""
import os
import re
import sys
from pathlib import Path

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(os.path.dirname(_HERE)))
sys.path.insert(0, _HERE)

from _session_config import BRAND_ROOT, CODES_EXCEL

import openpyxl

# ============================================================
# 运行参数（按需修改）
# ============================================================

# 图片来源目录
SOURCE_DIR = BRAND_ROOT / "classify" / "target"

# 输出 Excel 路径（默认与 run_linkfox_faceswap.py 共用同一文件）
OUTPUT_FILE = CODES_EXCEL

# 表头文字（设为 None 则不写表头）
HEADER = "商品编码"

# ============================================================

# 后缀关键词：文件名中 _{keyword}_N 视为"拍摄角度后缀"，前面的部分为编码
_SHOT_KEYWORDS = {"front", "back", "detail", "flat", "side", "alt", "hero", "model", "pair"}
_SUFFIX_RE = re.compile(
    r"_(" + "|".join(_SHOT_KEYWORDS) + r")_\d+$",
    flags=re.IGNORECASE,
)


def extract_code(filename: str) -> str | None:
    """从文件名（含或不含扩展名）提取商品编码。"""
    stem = Path(filename).stem
    m = _SUFFIX_RE.search(stem)
    if m:
        return stem[: m.start()]
    return None  # 文件名不符合预期格式，跳过


def collect_codes(source_dir: Path) -> list[str]:
    """读目录内所有 .jpg，提取唯一编码，保留首次出现顺序。"""
    seen: set[str] = set()
    codes: list[str] = []
    for f in sorted(source_dir.iterdir()):
        if f.suffix.lower() != ".jpg":
            continue
        code = extract_code(f.name)
        if code and code not in seen:
            seen.add(code)
            codes.append(code)
    return codes


def write_codes_excel(codes: list[str], output_path: Path, header: str | None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "codes"
    row = 1
    if header:
        ws.cell(row=row, column=1, value=header)
        row += 1
    for code in codes:
        ws.cell(row=row, column=1, value=code)
        row += 1
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def main() -> None:
    source = Path(SOURCE_DIR)
    if not source.exists():
        print(f"[ERROR] source dir not found: {source}")
        return

    codes = collect_codes(source)
    if not codes:
        print(f"[WARN] no matching .jpg files in {source}")
        return

    output = Path(OUTPUT_FILE)
    write_codes_excel(codes, output, HEADER)

    print(f"Done: {len(codes)} unique codes -> {output}")
    for c in codes:
        print(f"  {c}")


if __name__ == "__main__":
    main()

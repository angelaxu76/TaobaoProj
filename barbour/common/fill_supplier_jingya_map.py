# -*- coding: utf-8 -*-
from __future__ import annotations
import csv
from pathlib import Path
from typing import Dict, List, Optional, Set

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Connection

from config import BRAND_CONFIG, BARBOUR
from barbour.core.site_utils import canonical_site

PUBLICATION_DIR = Path(BARBOUR["PUBLICATION_DIR"])
PATTERN = "barbour_publication_*.xlsx"
TABLE = "barbour_supplier_map"

SQL_CREATE = text(f"""
CREATE TABLE IF NOT EXISTS {TABLE} (
  product_code VARCHAR(50) PRIMARY KEY,
  site_name    VARCHAR(100) NOT NULL
);
""")

SQL_PUBLISHED_CODES = text("""
SELECT DISTINCT product_code
FROM barbour_inventory
WHERE is_published = TRUE
  AND product_code IS NOT NULL
""")  # is_published/product_code 来自 inventory。:contentReference[oaicite:0]{index=0}

SQL_EXISTING_MAP = text(f"SELECT product_code FROM {TABLE}")

SQL_UPSERT = text(f"""
INSERT INTO {TABLE} (product_code, site_name)
VALUES (:code, :site)
ON CONFLICT (product_code) DO UPDATE SET site_name = EXCLUDED.site_name
""")

SQL_LOWEST_SITE = text("""
WITH agg AS (
  SELECT
    site_name,
    -- 有货尺码数
    SUM(CASE WHEN COALESCE(stock_count,0) > 0 THEN 1 ELSE 0 END) AS sizes_in_stock,
    -- 仅在有货尺码中取最低价
    MIN(COALESCE(NULLIF(price_gbp,0), original_price_gbp))
      FILTER (WHERE COALESCE(stock_count,0) > 0)                 AS min_price,
    MAX(last_checked)                                            AS latest
  FROM barbour_offers
  WHERE product_code = :code
    AND is_active = TRUE
  GROUP BY site_name
),
eligible AS (
  SELECT * FROM agg WHERE sizes_in_stock >= 3
)
SELECT site_name
FROM eligible
ORDER BY
  min_price ASC NULLS LAST,
  sizes_in_stock DESC,
  latest DESC
LIMIT 1
""")  # offers 字段参考：site_name/price_gbp/original_price_gbp/stock_count/last_checked。:contentReference[oaicite:1]{index=1}


def _load_publication_mappings(pub_dir: Path) -> Dict[str, str]:
    """
    读取目录下所有 barbour_publication_*.xlsx，返回 {product_code -> canonical_site}
    后读的新文件覆盖旧文件（以“最新发布”为准）。
    兼容列名：
      编码：Product Code / 商品编码 / product_code / color_code / 编码
      站点：Supplier / 供应商 / Site / site / 站点
    """
    def _headers(ws) -> Dict[str, int]:
        h = {}
        for j, c in enumerate(ws[1], start=1):
            k = str(c.value or "").strip().lower().replace(" ", "")
            if k:
                h[k] = j
        return h

    mappings: Dict[str, str] = {}
    files = sorted(pub_dir.glob(PATTERN), key=lambda p: p.stat().st_mtime)
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb.active
            hdr = _headers(ws)
            col_code = next((hdr[k] for k in ("productcode","商品编码","product_code","color_code","编码") if k in hdr), None)
            col_site = next((hdr[k] for k in ("supplier","供应商","site","站点") if k in hdr), None)
            if not col_code or not col_site:
                continue
            for i in range(2, ws.max_row + 1):
                code = str(ws.cell(i, col_code).value or "").strip()
                site_raw = str(ws.cell(i, col_site).value or "").strip()
                if not code or not site_raw:
                    continue
                site = canonical_site(site_raw)
                if site:
                    mappings[code] = site
        except Exception as e:
            print(f"⚠️ 解析失败 {fp.name}: {e}")
    return mappings


def _pick_lowest_site(conn: Connection, code: str) -> Optional[str]:
    row = conn.execute(SQL_LOWEST_SITE, {"code": code}).fetchone()
    if not row:
        return None
    return canonical_site(row[0]) or row[0]


def fill_supplier_map(force_refresh: bool = False) -> None:
    cfg = BRAND_CONFIG["barbour"]["PGSQL_CONFIG"]
    engine = create_engine(
        f"postgresql+psycopg2://{cfg['user']}:{cfg['password']}@{cfg['host']}:{cfg['port']}/{cfg['dbname']}"
    )

    with engine.begin() as conn:
        # 0) 确保表存在（NOT NULL 约束保持）
        conn.execute(SQL_CREATE)

        if force_refresh:
            conn.execute(text(f"TRUNCATE TABLE {TABLE};"))
            print(f"⚠️ 已清空 {TABLE} 表。")       

        # 1) 取“已发布”的编码集合（不插入任何 NULL）
        published: Set[str] = {r[0] for r in conn.execute(SQL_PUBLISHED_CODES).fetchall()}
        print(f"📦 已发布编码：{len(published)} 个。")  # 来自 inventory。:contentReference[oaicite:2]{index=2}

        # 2) 已有映射（避免重复处理）
        try:
            existing: Set[str] = {r[0] for r in conn.execute(SQL_EXISTING_MAP).fetchall()}
        except Exception:
            existing = set()

        # 3) 从发布清单写入映射（最新覆盖旧值）
        pub_map = _load_publication_mappings(PUBLICATION_DIR)
        pub_hit: List[str] = []
        for code, site in pub_map.items():
            if code in published:
                conn.execute(SQL_UPSERT, {"code": code, "site": site})
                pub_hit.append(code)
        print(f"✅ 按发布文件更新：{len(pub_hit)} 条。")
        # 打印命中编码（便于你区分来源）
        if pub_hit:
            print("→ 来自 publication 的编码：", ", ".join(pub_hit))

        # 4) 兜底：对“已发布但还未映射”的编码，用 offers 选最低价站点
        need = (published - set(pub_hit)) - existing
        offer_filled: List[str] = []
        for code in sorted(need):
            site = _pick_lowest_site(conn, code)
            if site:
                conn.execute(SQL_UPSERT, {"code": code, "site": site})
                offer_filled.append(code)
        print(f"✅ 按 offers 兜底：{len(offer_filled)} 条。")
        if offer_filled:
            print("→ 来自 offers 兜底的编码：", ", ".join(offer_filled))

        # 5) 统计收尾
        total_now = conn.execute(text(f"SELECT COUNT(*) FROM {TABLE}")).scalar_one()
        print(f"🎯 完成映射，总计 {total_now} 条。")

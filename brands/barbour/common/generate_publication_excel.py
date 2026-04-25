# -*- coding: utf-8 -*-
"""
生成 Barbour 外套夹克发布 Excel
"""

import re
import unicodedata
import openpyxl
from sqlalchemy import create_engine, text

from config import BRAND_CONFIG, BARBOUR, SETTINGS, EXCHANGE_RATE
from brands.barbour.common.generate_taobao_title_v2 import generate_barbour_taobao_title
from common.pricing.price_utils import calculate_jingya_prices
from datetime import datetime
from brands.barbour.core.site_utils import canonical_site

# ========== 路径 ==========
TXT_DIR = BARBOUR["TXT_DIR_ALL"]
CODES_FILE = BARBOUR["OUTPUT_DIR"] / "codes.txt"
CODES_XLSX = BARBOUR["OUTPUT_DIR"] / "codes.xlsx"   # ← 新增：支持 Excel 输入（商品编码+供应商）
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = BARBOUR["PUBLICATION_DIR"] / f"barbour_publication_{timestamp}.xlsx"

# ========== 正则规则 ==========
FIT_PAT = {
    "修身型": re.compile(r"\b(slim|tailored|trim)\b", re.I),
    "宽松型": re.compile(r"\b(relaxed|loose|oversized|boxy)\b", re.I),
}
NECK_PAT = [
    ("连帽", re.compile(r"\bhood(ed)?|detachable hood\b", re.I)),
    ("立领", re.compile(r"\bstand collar|funnel (neck|collar)|mock neck\b", re.I)),
    ("翻领", re.compile(r"\bcord(uroy)? collar|spread collar|point collar|shirt[- ]style collar\b", re.I)),
]
LEN_TXT = [
    ("短款", re.compile(r"\bshort(er)? length|cropped\b", re.I)),
    ("长款", re.compile(r"\blong(er)? length|longline\b", re.I)),
]
LEN_NUM = re.compile(r"(Back length|Sample length)[:：]?\s*([0-9]+(?:\.[0-9]+)?)\s*cm", re.I)


def _nfkc(s: str) -> str:
    return unicodedata.normalize("NFKC", (s or "")).strip()


def map_color(color_en: str) -> str:
    c = _nfkc(color_en)
    c = re.sub(r"^[\-\:\|•\.\s]+", "", c)
    c = c.split("/")[0].strip()
    cl = c.lower()
    COLOR_MAP = BARBOUR["BARBOUR_COLOR_MAP"]
    if cl in COLOR_MAP:
        return COLOR_MAP[cl]
    cl2 = re.sub(r"^(classic|washed|burnt|dark|light)\s+", "", cl).strip()
    if cl2 in COLOR_MAP:
        return COLOR_MAP[cl2]
    fallback = [
        ("sage", "鼠尾草绿"),
        ("olive", "橄榄绿"),
        ("navy", "海军蓝"),
        ("moss", "苔绿"),
        ("khaki", "卡其"),
        ("stone", "石色"),
        ("sand", "沙色"),
        ("beige", "米色"),
        ("burgundy", "酒红"),
        ("cobalt", "钴蓝"),
        ("denim", "丹宁蓝"),
        ("grey", "灰色"),
        ("gray", "灰色"),
    ]
    for key, zh in fallback:
        if key in cl:
            return zh
    return c


def infer_fit_neck_length(name: str, desc: str = "", feature: str = ""):
    text = " ".join([name or "", desc or "", feature or ""])
    fit = "标准"
    for zh, pat in FIT_PAT.items():
        if pat.search(text):
            fit = zh
            break
    neckline = "无"
    for zh, pat in NECK_PAT:
        if pat.search(text):
            neckline = zh
            break
    length = None
    m = LEN_NUM.search(text)
    if m:
        cm = float(m.group(2))
        if cm < 66:
            length = "短款"
        elif cm > 78:
            length = "长款"
        else:
            length = "常规"
    if not length:
        for zh, pat in LEN_TXT:
            if pat.search(text):
                length = zh
                break
    if not length:
        length = "常规"
    return fit, neckline, length


# ========== SQL ==========
SQL_PRODUCT = text("""
    SELECT DISTINCT style_name, color
    FROM barbour_products
    WHERE product_code = :code
    ORDER BY style_name
    LIMIT 1
""")

# 基础 SQL（按需拼接站点过滤）
SQL_OFFERS_ORDERABLE_BASE = """
    SELECT site_name, offer_url, price_gbp,
           stock_count, last_checked, size
    FROM barbour_offers
    WHERE product_code = :code
      AND (
        stock_count IS NULL
        OR stock_count > 0
      )
      AND price_gbp IS NOT NULL
"""

SQL_OFFERS_ORDER_BY = " ORDER BY price_gbp ASC"


def _in_stock_can_order(row: dict) -> bool:
    cnt = row.get("stock_count")
    try:
        cnt = int(cnt) if cnt is not None else None
    except Exception:
        cnt = None
    return bool(row.get("can_order")) and (cnt is None or cnt > 0)


def compute_rmb_price(min_gbp: float, exchange_rate: float):
    try:
        orig = float(min_gbp) if min_gbp is not None else 0.0
        disc = orig
        base_price = max(orig, disc)
        untaxed, retail = calculate_jingya_prices(
            base_price=base_price,
            delivery_cost=7,
            exchange_rate=exchange_rate
        )
        return untaxed, retail
    except Exception:
        return "", ""


# ===== 新增：读取 codes.xlsx（code + supplier），如无则回退 codes.txt（仅 code） =====
def load_codes_with_supplier():
    """
    返回列表 [(code, supplier or None), ...]
    - 优先读取 OUTPUT_DIR/codes.xlsx
      * 头部容错：['Product Code', '商品编码', 'product_code'] / ['Supplier', '供应商', 'site_name']
    - 否则回退 OUTPUT_DIR/codes.txt（每行一个 code）
    """
    pairs = []
    if CODES_XLSX.exists():
        wb = openpyxl.load_workbook(CODES_XLSX, data_only=True)
        ws = wb.active
        # 读取表头
        headers = {}
        for j, cell in enumerate(ws[1], start=1):
            key = _nfkc(str(cell.value)).lower()
            headers[key] = j
        # 映射列
        def _col(name_candidates):
            for k, col in headers.items():
                for cand in name_candidates:
                    if k == cand or k.replace(" ", "") == cand.replace(" ", ""):
                        return col
            return None

        col_code = _col(["product code", "商品编码", "product_code", "color_code", "编码"])
        col_sup  = _col(["supplier", "供应商", "site_name", "站点"])

        if not col_code:
            raise SystemExit("❌ codes.xlsx 缺少 'Product Code/商品编码' 列")

        for i in range(2, ws.max_row + 1):
            code = _nfkc(ws.cell(i, col_code).value or "")
            if not code:
                continue
            supplier_raw = _nfkc(ws.cell(i, col_sup).value) if col_sup else ""
            # 新增：统一站点名为配置键名；无法识别就置为 None
            supplier = canonical_site(supplier_raw) if supplier_raw else None
            pairs.append((code, supplier))
        if not pairs:
            raise SystemExit("❌ codes.xlsx 未读取到任何编码")
        print(f"🟢 已从 Excel 读取 {len(pairs)} 条编码（含供应商）")
        return pairs

    # 回退到 TXT
    if not CODES_FILE.exists():
        raise SystemExit(f"❌ 未找到 codes.xlsx 或 codes.txt（请在 {BARBOUR['OUTPUT_DIR']} 放置其中之一）")
    codes = [line.strip() for line in CODES_FILE.read_text(encoding="utf-8").splitlines() if line.strip()]
    pairs = [(c, None) for c in codes]
    print(f"🟡 未发现 codes.xlsx，回退使用 codes.txt（{len(pairs)} 条）")
    return pairs


def generate_publication_excel():
    cfg = BRAND_CONFIG["barbour"]
    PGSQL = cfg["PGSQL_CONFIG"]

    engine = create_engine(
        f"postgresql+psycopg2://{PGSQL['user']}:{PGSQL['password']}@{PGSQL['host']}:{PGSQL['port']}/{PGSQL['dbname']}"
    )

    # 读取 (code, supplier)
    code_pairs = load_codes_with_supplier()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header = [
        "商品编码", "Product Name EN", "颜色",
        "商品名称",
        "Min Price (GBP)", "鲸芽价格", "淘宝价格",
        "版型", "领口设计", "衣长",
        "Sizes (In Stock)",
        "Supplier",
        "Site", "Offer URL", "Stock Count", "Last Checked"
    ]
    ws.append(header)

    rate = EXCHANGE_RATE

    with engine.connect() as conn:
        for idx, (code, supplier) in enumerate(code_pairs, 1):
            product = conn.execute(SQL_PRODUCT, {"code": code}).mappings().first()
            if not product:
                print(f"[{idx}/{len(code_pairs)}] {code} ❌ 未找到产品信息")
                continue

            style_name = product["style_name"]
            color_en = product["color"]
            color_cn = map_color(color_en)

            # 构建 offers 查询（若提供 supplier，仅在该站点检索）
            sql = SQL_OFFERS_ORDERABLE_BASE
            params = {"code": code}
            if supplier:
                sql += " AND lower(site_name) = lower(:supplier)"
                params["supplier"] = supplier
            sql += SQL_OFFERS_ORDER_BY

            offers_rows = list(conn.execute(text(sql), params).mappings())

            if supplier and not offers_rows:
                print(f"[{idx}/{len(code_pairs)}] {code} ⚠️ 供应商={supplier} 无可下单报价，已跳过（请检查 codes.xlsx）")
                continue
            if not offers_rows:
                print(f"[{idx}/{len(code_pairs)}] {code} ⚠️ 未找到可下单报价")
                continue

            def _row_is_available(row: dict) -> bool:
                cnt = row.get("stock_count")
                try:
                    cnt = int(cnt) if cnt is not None else None
                except Exception:
                    cnt = None
                # 约定：NULL 视为未知（保留），>0 视为可下单，<=0 视为不可下单
                return (cnt is None) or (cnt > 0)

            sizes_in_stock = sorted({
                (row.get("size") or "").strip()
                for row in offers_rows
                if row.get("size") and _row_is_available(row)
            })
            sizes_str = ", ".join(sizes_in_stock) if sizes_in_stock else ""

            # 最低价记录（SQL 已按 price_gbp ASC）
            best = offers_rows[0]
            site_name = best["site_name"]
            offer_url = best["offer_url"]
            price_gbp = best["price_gbp"]
            stock_count = best["stock_count"]
            last_checked = best["last_checked"]

            # 中文标题
            title_info = generate_barbour_taobao_title(code, style_name, color_en)
            title_cn = title_info["Title"]

            # 售价计算
            untaxed, retail = compute_rmb_price(price_gbp, exchange_rate=rate)

            # 版型/领口/衣长
            fit, neckline, coat_len = infer_fit_neck_length(style_name)

            row = [
                code, style_name, color_cn,
                title_cn,
                float(price_gbp) if price_gbp is not None else "",
                untaxed, retail,
                fit, neckline, coat_len,
                sizes_str,
                supplier or "",   # ← 新增：把输入 supplier 原样写入（作为发布清单锚点）
                site_name, offer_url, stock_count, last_checked
            ]
            ws.append(row)

            print(f"[{idx}/{len(code_pairs)}] {code} ← supplier={supplier or 'N/A'} | site={site_name} | £{price_gbp} | 尺码[{sizes_str}] | {title_cn}")

    wb.save(OUTPUT_FILE)
    print(f"✅ Excel 已生成: {OUTPUT_FILE}")


if __name__ == "__main__":
    generate_publication_excel()

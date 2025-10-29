# -*- coding: utf-8 -*-
from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Optional, Set

import openpyxl
from sqlalchemy import text
from sqlalchemy.engine import Connection

from config import BRAND_CONFIG, BARBOUR
from brands.barbour.core.site_utils import canonical_site

PUBLICATION_DIR = Path(BARBOUR["PUBLICATION_DIR"])
PATTERN = "barbour_publication_*.xlsx"
TABLE = "barbour_supplier_map"

SQL_CREATE = text(f"""
CREATE TABLE IF NOT EXISTS {TABLE} (
  product_code VARCHAR(50) PRIMARY KEY,
  site_name    VARCHAR(100) NOT NULL
);
""")

SQL_PUBLISHED_CODES = text("""
SELECT DISTINCT product_code
FROM barbour_inventory
WHERE is_published = TRUE
  AND product_code IS NOT NULL
""")  # is_published/product_code 来自 inventory。:contentReference[oaicite:0]{index=0}

SQL_EXISTING_MAP = text(f"SELECT product_code FROM {TABLE}")

SQL_UPSERT = text(f"""
INSERT INTO {TABLE} (product_code, site_name)
VALUES (:code, :site)
ON CONFLICT (product_code) DO UPDATE SET site_name = EXCLUDED.site_name
""")

SQL_LOWEST_SITE = text("""
WITH agg AS (
  SELECT
    site_name,
    -- 有货尺码数
    SUM(CASE WHEN COALESCE(stock_count,0) > 0 THEN 1 ELSE 0 END) AS sizes_in_stock,
    -- 仅在有货尺码中取最低价
    MIN(COALESCE(NULLIF(price_gbp,0), original_price_gbp))
      FILTER (WHERE COALESCE(stock_count,0) > 0)                 AS min_price,
    MAX(last_checked)                                            AS latest
  FROM barbour_offers
  WHERE product_code = :code
    AND is_active = TRUE
  GROUP BY site_name
),
eligible AS (
  SELECT * FROM agg WHERE sizes_in_stock >= 3
)
SELECT site_name
FROM eligible
ORDER BY
  min_price ASC NULLS LAST,
  sizes_in_stock DESC,
  latest DESC
LIMIT 1
""")  # offers 字段参考：site_name/price_gbp/original_price_gbp/stock_count/last_checked。:contentReference[oaicite:1]{index=1}


def _load_publication_mappings(pub_dir: Path) -> Dict[str, str]:
    """
    读取目录下所有 barbour_publication_*.xlsx，返回 {product_code -> canonical_site}
    后读的新文件覆盖旧文件（以“最新发布”为准）。
    兼容列名：
      编码：Product Code / 商品编码 / product_code / color_code / 编码
      站点：Supplier / 供应商 / Site / site / 站点
    """
    def _headers(ws) -> Dict[str, int]:
        h = {}
        for j, c in enumerate(ws[1], start=1):
            k = str(c.value or "").strip().lower().replace(" ", "")
            if k:
                h[k] = j
        return h

    mappings: Dict[str, str] = {}
    files = sorted(pub_dir.glob(PATTERN), key=lambda p: p.stat().st_mtime)
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb.active
            hdr = _headers(ws)
            col_code = next((hdr[k] for k in ("productcode","商品编码","product_code","color_code","编码") if k in hdr), None)
            col_site = next((hdr[k] for k in ("supplier","供应商","site","站点") if k in hdr), None)
            if not col_code or not col_site:
                continue
            for i in range(2, ws.max_row + 1):
                code = str(ws.cell(i, col_code).value or "").strip()
                site_raw = str(ws.cell(i, col_site).value or "").strip()
                if not code or not site_raw:
                    continue
                site = canonical_site(site_raw)
                if site:
                    mappings[code] = site
        except Exception as e:
            print(f"⚠️ 解析失败 {fp.name}: {e}")
    return mappings


def _pick_lowest_site(conn: Connection, code: str) -> Optional[str]:
    row = conn.execute(SQL_LOWEST_SITE, {"code": code}).fetchone()
    if not row:
        return None
    return canonical_site(row[0]) or row[0]


def fill_supplier_map(force_refresh: bool = False) -> None:
    cfg = BRAND_CONFIG["barbour"]["PGSQL_CONFIG"]
    engine = create_engine(
        f"postgresql+psycopg2://{cfg['user']}:{cfg['password']}@{cfg['host']}:{cfg['port']}/{cfg['dbname']}"
    )

    with engine.begin() as conn:
        # 0) 确保表存在（NOT NULL 约束保持）
        conn.execute(SQL_CREATE)

        if force_refresh:
            conn.execute(text(f"TRUNCATE TABLE {TABLE};"))
            print(f"⚠️ 已清空 {TABLE} 表。")       

        # 1) 取“已发布”的编码集合（不插入任何 NULL）
        published: Set[str] = {r[0] for r in conn.execute(SQL_PUBLISHED_CODES).fetchall()}
        print(f"📦 已发布编码：{len(published)} 个。")  # 来自 inventory。:contentReference[oaicite:2]{index=2}

        # 2) 已有映射（避免重复处理）
        try:
            existing: Set[str] = {r[0] for r in conn.execute(SQL_EXISTING_MAP).fetchall()}
        except Exception:
            existing = set()

        # 3) 从发布清单写入映射（最新覆盖旧值）
        pub_map = _load_publication_mappings(PUBLICATION_DIR)
        pub_hit: List[str] = []
        for code, site in pub_map.items():
            if code in published:
                conn.execute(SQL_UPSERT, {"code": code, "site": site})
                pub_hit.append(code)
        print(f"✅ 按发布文件更新：{len(pub_hit)} 条。")
        # 打印命中编码（便于你区分来源）
        if pub_hit:
            print("→ 来自 publication 的编码：", ", ".join(pub_hit))

        # 4) 兜底：对“已发布但还未映射”的编码，用 offers 选最低价站点
        need = (published - set(pub_hit)) - existing
        offer_filled: List[str] = []
        for code in sorted(need):
            site = _pick_lowest_site(conn, code)
            if site:
                conn.execute(SQL_UPSERT, {"code": code, "site": site})
                offer_filled.append(code)
        print(f"✅ 按 offers 兜底：{len(offer_filled)} 条。")
        if offer_filled:
            print("→ 来自 offers 兜底的编码：", ", ".join(offer_filled))

        # 5) 统计收尾
        total_now = conn.execute(text(f"SELECT COUNT(*) FROM {TABLE}")).scalar_one()
        print(f"🎯 完成映射，总计 {total_now} 条。")


def reassign_low_stock_suppliers(size_threshold: int = 3, dry_run: bool = True) -> list[dict]:
    """
    找出当前映射站点“在售尺码数 < size_threshold”的商品；若存在其它站点满足(尺码≥阈值 & 最低价最低)，则建议/执行切换。
    - dry_run=True：只打印与返回建议，不改库
    - 返回：建议切换清单 [{code, old_site, old_sizes, new_site, new_sizes, old_min_price, new_min_price}]
    """
    cfg = BRAND_CONFIG["barbour"]["PGSQL_CONFIG"]
    eng = create_engine(f"postgresql+psycopg2://{cfg['user']}:{cfg['password']}@{cfg['host']}:{cfg['port']}/{cfg['dbname']}")
    suggest: list[dict] = []

    SQL_AGG = text("""
    WITH agg AS (
      SELECT
        product_code,
        site_name,
        SUM(CASE WHEN COALESCE(stock_count,0) > 0 THEN 1 ELSE 0 END) AS sizes_in_stock,
        MIN(COALESCE(NULLIF(sale_price_gbp,0), NULLIF(price_gbp,0), original_price_gbp))
            FILTER (WHERE COALESCE(stock_count,0) > 0)               AS min_eff_price,
        MAX(last_checked)                                            AS latest
      FROM barbour_offers
      WHERE is_active = TRUE
      GROUP BY product_code, site_name
    )
    SELECT * FROM agg
    """)

    with eng.begin() as conn:
        # 映射表
        map_df = pd.read_sql("SELECT product_code, site_name FROM barbour_supplier_map", conn)
        map_df["site_name"] = map_df["site_name"].map(lambda s: canonical_site(s) or s)

        # 全部站点聚合表现
        agg_df = pd.read_sql(SQL_AGG, conn)
        agg_df["site_name"] = agg_df["site_name"].map(lambda s: canonical_site(s) or s)

        # 合并拿到当前站点表现
        cur_df = map_df.merge(
            agg_df.rename(columns={
                "sizes_in_stock": "cur_sizes_in_stock",
                "min_eff_price": "cur_min_eff_price",
                "latest": "cur_latest"
            }),
            on=["product_code", "site_name"], how="left"
        )

        # 遍历：当前站点尺码不足阈值 → 找“候选最佳站点”
        for _, r in cur_df.iterrows():
            code = str(r["product_code"])
            old_site = r["site_name"]
            cur_sizes = (r.get("cur_sizes_in_stock") or 0)

            if cur_sizes is None or int(cur_sizes) >= int(size_threshold):
                continue  # 当前站点尺码数已满足，不处理

            # 候选站点：≥阈值的站点里选价格最低（若并列则尺码多、更新时间新）
            cand = (
                agg_df[(agg_df["product_code"] == code) & (agg_df["sizes_in_stock"] >= size_threshold) & agg_df["min_eff_price"].notna()]
                .sort_values(["min_eff_price", "sizes_in_stock", "latest"], ascending=[True, False, False])
                .head(1)
            )
            if cand.empty:
                continue

            new_site = cand.iloc[0]["site_name"]
            new_sizes = int(cand.iloc[0]["sizes_in_stock"])
            new_price = float(cand.iloc[0]["min_eff_price"] or 0.0)
            old_price = float(r.get("cur_min_eff_price") or 0.0)

            # 记录建议
            suggest.append({
                "product_code": code,
                "old_site": old_site,
                "old_sizes": int(cur_sizes or 0),
                "old_min_price": old_price,
                "new_site": new_site,
                "new_sizes": new_sizes,
                "new_min_price": new_price,
            })

        # 执行切换（可选）
        if suggest and not dry_run:
            rows = [{"code": s["product_code"], "site": s["new_site"]} for s in suggest]
            conn.execute(text("""
                INSERT INTO barbour_supplier_map(product_code, site_name)
                VALUES (:code, :site)
                ON CONFLICT (product_code) DO UPDATE SET site_name = EXCLUDED.site_name
            """), rows)

    # 打印预览
    if suggest:
        print(f"共{len(suggest)}条建议：")
        for s in suggest[:30]:
            print(f"- {s['product_code']}: {s['old_site']}({s['old_sizes']}尺) -> {s['new_site']}({s['new_sizes']}尺), "
                  f"价 {s['old_min_price']} -> {s['new_min_price']}")
        if len(suggest) > 30:
            print("...（其余省略）")
    else:
        print("未找到需要切换的商品（当前映射站点均满足尺码阈值或无更优候选）。")

    return suggest


def export_supplier_stock_price_report(min_sizes_ok: int = 1, output_path: str | None = None) -> str:
    """
    导出每个商品在各站点的“在售尺码数、最低有效价、最近更新时间”，并标注当前映射与推荐站点。
    - min_sizes_ok: 统计“有货尺码数”的阈值（默认>=1视为有货参与最低价评选）
    - output_path: 指定导出xlsx路径；不传则写到 BARBOUR['OUTPUT_DIR']/barbour_supplier_report.xlsx
    """
    cfg = BRAND_CONFIG["barbour"]["PGSQL_CONFIG"]
    out_dir = Path(BARBOUR["OUTPUT_DIR"])
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = Path(output_path) if output_path else (out_dir / "barbour_supplier_report.xlsx")

    eng = create_engine(f"postgresql+psycopg2://{cfg['user']}:{cfg['password']}@{cfg['host']}:{cfg['port']}/{cfg['dbname']}")
    with eng.begin() as conn:
        # 当前映射
        map_df = pd.read_sql("SELECT product_code, site_name FROM barbour_supplier_map", conn)

        # 聚合各站点表现（和 fill_supplier_jingya_map 内部 SQL 口径一致）
        sql = f"""
        WITH base AS (
          SELECT
            product_code,
            site_name,
            COALESCE(stock_count,0) AS stock_count,
            COALESCE(NULLIF(sale_price_gbp,0), NULLIF(price_gbp,0), original_price_gbp) AS eff_price,
            last_checked
          FROM barbour_offers
          WHERE is_active = TRUE
            AND product_code IS NOT NULL AND product_code <> ''
            AND site_name IS NOT NULL AND site_name <> ''
        ),
        agg AS (
          SELECT
            product_code,
            site_name,
            SUM(CASE WHEN stock_count > 0 THEN 1 ELSE 0 END)                       AS sizes_in_stock,
            MIN(eff_price) FILTER (WHERE stock_count > 0)                           AS min_eff_price,
            MAX(last_checked)                                                       AS latest
          FROM base
          GROUP BY product_code, site_name
        ),
        ranked AS (
          SELECT
            *,
            ROW_NUMBER() OVER (
              PARTITION BY product_code
              ORDER BY
                CASE WHEN sizes_in_stock >= 3 THEN 0 ELSE 1 END,   -- ≥3尺有货优先
                min_eff_price ASC NULLS LAST,
                sizes_in_stock DESC,
                latest DESC
            ) AS rank_all,
            ROW_NUMBER() OVER (
              PARTITION BY product_code
              ORDER BY
                CASE WHEN sizes_in_stock >= 3 THEN 0 ELSE 1 END,
                min_eff_price ASC NULLS LAST
            ) AS rank_price_first
          FROM agg
        )
        SELECT * FROM ranked
        """
        df = pd.read_sql(sql, conn)

    # 规范
    df["site_name"] = df["site_name"].map(lambda s: canonical_site(s) or s)
    map_df["site_name"] = map_df["site_name"].map(lambda s: canonical_site(s) or s)

    # 标注“当前映射”
    df = df.merge(map_df, on="product_code", how="left", suffixes=("", "_mapped"))
    df["is_current"] = (df["site_name"] == df["site_name_mapped"]).fillna(False)

    # 推荐站点（≥3尺有货 & 价格最低）
    best = (
        df[(df["sizes_in_stock"] >= 3) & df["min_eff_price"].notna()]
        .sort_values(["product_code", "min_eff_price", "sizes_in_stock", "latest"], ascending=[True, True, False, False])
        .drop_duplicates(["product_code"])
        .rename(columns={"site_name":"best_site", "min_eff_price":"best_min_eff_price", "sizes_in_stock":"best_sizes_in_stock"})
        [["product_code","best_site","best_min_eff_price","best_sizes_in_stock"]]
    )
    df = df.merge(best, on="product_code", how="left")

    # 只保留有意义的列
    out = df[[
        "product_code",
        "site_name",
        "is_current",
        "sizes_in_stock",
        "min_eff_price",
        "latest",
        "site_name_mapped",
        "best_site",
        "best_min_eff_price",
        "best_sizes_in_stock",
        "rank_all",
        "rank_price_first",
    ]].sort_values(["product_code", "is_current"], ascending=[True, False])

    out.to_excel(out_file, index=False)
    return str(out_file)


import pandas as pd
from sqlalchemy import create_engine, text

def apply_barbour_supplier_overrides(xlsx_path: str, dry_run: bool = False) -> None:
    """
    按 Excel 文件手动指定 Barbour 商品供货商。
    Excel 要求：
      - 第一列：商品编码（列名必须为 “商品编码”）
      - 第二列：供货商（列名必须为 “供货商”）
    会将对应关系写入 barbour_supplier_map（有则更新，无则插入）。
    dry_run=True 时只预览不写库。
    """
    cfg = BRAND_CONFIG["barbour"]["PGSQL_CONFIG"]
    engine = create_engine(
        f"postgresql+psycopg2://{cfg['user']}:{cfg['password']}@{cfg['host']}:{cfg['port']}/{cfg['dbname']}"
    )

    # 1) 读取 Excel（要求含表头）
    df = pd.read_excel(xlsx_path, dtype=str)
    required_cols = ["商品编码", "供货商"]
    if not all(c in df.columns for c in required_cols):
        raise ValueError(f"Excel 必须包含列：{required_cols}，当前表头：{list(df.columns)}")

    df = df[required_cols].rename(columns={"商品编码": "product_code", "供货商": "site_name"})
    df["product_code"] = df["product_code"].astype(str).str.strip()
    df["site_name"] = df["site_name"].astype(str).str.strip()
    df = df[(df["product_code"] != "") & (df["site_name"] != "")].drop_duplicates()

    if df.empty:
        print("[INFO] Excel 中无有效数据。")
        return

    # 站点归一化
    df["site_name"] = df["site_name"].map(lambda s: canonical_site(s) or s)

    print(f"[INFO] 读取 {len(df)} 条供货商指定记录。示例前5条：")
    for _, row in df.head(5).iterrows():
        print(f"  - {row['product_code']} -> {row['site_name']}")
    if len(df) > 5:
        print("  ...")

    # 2) 执行写入
    if dry_run:
        print("[DRY-RUN] 仅预览，不写入数据库。")
        return

    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS barbour_supplier_map (
              product_code VARCHAR(50) PRIMARY KEY,
              site_name    VARCHAR(100) NOT NULL
            )
        """))
        conn.execute(text("""
            INSERT INTO barbour_supplier_map(product_code, site_name)
            VALUES (:product_code, :site_name)
            ON CONFLICT (product_code) DO UPDATE SET site_name = EXCLUDED.site_name
        """), df.to_dict(orient="records"))
        print(f"[OK] 已成功更新 {len(df)} 条供货商映射。")

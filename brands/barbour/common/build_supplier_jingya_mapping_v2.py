# -*- coding: utf-8 -*-
"""
build_supplier_jingya_mapping v2
================================
根据供货商报价，为每个已发布商品选择最优供货商。

相比 v1 的改进：
- 策略模式：支持 LowestPriceStrategy / MostStockStrategy / PreferredSiteStrategy 切换
- 统一 import、统一引擎创建、统一日志
- 修复死代码与结构隐患
"""
from __future__ import annotations

import logging
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Dict, List, Optional, Set

import openpyxl
import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Connection, Engine

from config import BRAND_CONFIG, BARBOUR
from brands.barbour.core.site_utils import canonical_site
from brands.barbour.core.supplier_price_rules import (
    strategy_all_ratio,
    strategy_ratio_when_no_discount,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
#  常量
# ---------------------------------------------------------------------------
PUBLICATION_DIR = Path(BARBOUR["PUBLICATION_DIR"])
PATTERN = "barbour_publication_*.xlsx"
TABLE = "barbour_supplier_map"

# ---------------------------------------------------------------------------
#  SQL
# ---------------------------------------------------------------------------
SQL_CREATE = text(f"""
CREATE TABLE IF NOT EXISTS {TABLE} (
  product_code VARCHAR(50) PRIMARY KEY,
  site_name    VARCHAR(100) NOT NULL
);
""")

SQL_PUBLISHED_CODES = text("""
SELECT DISTINCT product_code
FROM barbour_inventory
WHERE is_published = TRUE
  AND product_code IS NOT NULL
""")

SQL_EXISTING_MAP = text(f"SELECT product_code FROM {TABLE}")

SQL_UPSERT = text(f"""
INSERT INTO {TABLE} (product_code, site_name)
VALUES (:code, :site)
ON CONFLICT (product_code) DO UPDATE SET site_name = EXCLUDED.site_name
""")

# 通用聚合：按 product_code + site_name 统计有货尺码数、最低有效价、最近检查时间
SQL_SITE_AGG = text("""
SELECT
  site_name,
  SUM(CASE WHEN COALESCE(stock_count, 0) > 0 THEN 1 ELSE 0 END) AS sizes_in_stock,
  MIN(COALESCE(NULLIF(sale_price_gbp, 0), NULLIF(price_gbp, 0), original_price_gbp))
    FILTER (WHERE COALESCE(stock_count, 0) > 0)                   AS min_price,
  MAX(last_checked)                                                AS latest
FROM barbour_offers
WHERE product_code = :code
  AND is_active = TRUE
GROUP BY site_name
""")

# 全量聚合（用于 reassign / report）
SQL_FULL_AGG = text("""
SELECT
  product_code,
  site_name,
  SUM(CASE WHEN COALESCE(stock_count, 0) > 0 THEN 1 ELSE 0 END) AS sizes_in_stock,
  MIN(COALESCE(NULLIF(sale_price_gbp, 0), NULLIF(price_gbp, 0), original_price_gbp))
    FILTER (WHERE COALESCE(stock_count, 0) > 0)                   AS min_eff_price,
  MAX(last_checked)                                                AS latest
FROM barbour_offers
WHERE is_active = TRUE
GROUP BY product_code, site_name
""")


# 按站点返回每条有货 offer 的原价 + 折扣价（TrueCostStrategy 用）
SQL_SITE_OFFERS = text("""
SELECT
  site_name,
  original_price_gbp,
  sale_price_gbp,
  COALESCE(stock_count, 0) AS stock_count
FROM barbour_offers
WHERE product_code = :code
  AND is_active = TRUE
""")

# 全量 offer 明细（TrueCostStrategy 批量场景用）
SQL_FULL_OFFERS = text("""
SELECT
  product_code,
  site_name,
  original_price_gbp,
  sale_price_gbp,
  COALESCE(stock_count, 0) AS stock_count,
  last_checked
FROM barbour_offers
WHERE is_active = TRUE
""")

# ---------------------------------------------------------------------------
#  供货商折扣规则工具
# ---------------------------------------------------------------------------
_DISCOUNT_RULES: Dict[str, dict] = BARBOUR.get("SUPPLIER_DISCOUNT_RULES", {})
_STRATEGY_FN = {
    "all_ratio": strategy_all_ratio,
    "ratio_when_no_discount": strategy_ratio_when_no_discount,
}


def _true_cost(
    site_name: str,
    original_price: Optional[float],
    sale_price: Optional[float],
) -> float:
    """
    根据 SUPPLIER_DISCOUNT_RULES 计算供货商的真实落地成本。
    应用 extra_ratio + shipping_fee，不同站点策略不同。
    """
    canon = canonical_site(site_name) or site_name
    rule = _DISCOUNT_RULES.get(canon, _DISCOUNT_RULES.get("__default__", {}))
    fn = _STRATEGY_FN.get(rule.get("strategy", ""), strategy_ratio_when_no_discount)
    cost = fn(
        original_price=original_price,
        discounted_price=sale_price,
        extra_ratio=rule.get("extra_ratio", 1.0),
        shipping_fee=rule.get("shipping_fee", 0.0),
    )
    return cost


# ---------------------------------------------------------------------------
#  引擎工厂
# ---------------------------------------------------------------------------
def _get_engine() -> Engine:
    cfg = BRAND_CONFIG["barbour"]["PGSQL_CONFIG"]
    return create_engine(
        f"postgresql+psycopg2://{cfg['user']}:{cfg['password']}"
        f"@{cfg['host']}:{cfg['port']}/{cfg['dbname']}"
    )


# ---------------------------------------------------------------------------
#  策略模式
# ---------------------------------------------------------------------------
class SupplierStrategy(ABC):
    """供货商选择策略基类"""

    @abstractmethod
    def select(self, conn: Connection, code: str) -> Optional[str]:
        """给定商品编码，返回选中的 site_name；无合适候选时返回 None。"""
        ...

    def select_from_df(
        self, candidates: pd.DataFrame, code: str
    ) -> Optional[str]:
        """
        从 DataFrame（含 site_name / sizes_in_stock / min_eff_price / latest 列）
        中选出最优站点。用于 reassign 等批量场景，避免逐条查库。
        默认实现委托子类覆写；若子类未覆写则回退到 None。
        """
        return None


class LowestPriceStrategy(SupplierStrategy):
    """现有默认逻辑：库存达标后选最低有效价。"""

    def __init__(self, min_sizes: int = 3):
        self.min_sizes = min_sizes

    def select(self, conn: Connection, code: str) -> Optional[str]:
        rows = conn.execute(SQL_SITE_AGG, {"code": code}).fetchall()
        eligible = [
            r for r in rows if (r.sizes_in_stock or 0) >= self.min_sizes
        ]
        if not eligible:
            return None
        # 排序：价格升序 → 尺码降序 → 最近检查降序
        eligible.sort(
            key=lambda r: (
                r.min_price if r.min_price is not None else float("inf"),
                -(r.sizes_in_stock or 0),
                -(r.latest.timestamp() if r.latest else 0),
            )
        )
        site = eligible[0].site_name
        return canonical_site(site) or site

    def select_from_df(
        self, candidates: pd.DataFrame, code: str
    ) -> Optional[str]:
        cand = candidates[
            (candidates["product_code"] == code)
            & (candidates["sizes_in_stock"] >= self.min_sizes)
            & candidates["min_eff_price"].notna()
        ]
        if cand.empty:
            return None
        best = cand.sort_values(
            ["min_eff_price", "sizes_in_stock", "latest"],
            ascending=[True, False, False],
        ).iloc[0]
        return best["site_name"]


class MostStockStrategy(SupplierStrategy):
    """优先选库存（有货尺码数）最多的供货商，相同尺码数时选最低价。"""

    def __init__(self, min_sizes: int = 1):
        self.min_sizes = min_sizes

    def select(self, conn: Connection, code: str) -> Optional[str]:
        rows = conn.execute(SQL_SITE_AGG, {"code": code}).fetchall()
        eligible = [
            r for r in rows if (r.sizes_in_stock or 0) >= self.min_sizes
        ]
        if not eligible:
            return None
        eligible.sort(
            key=lambda r: (
                -(r.sizes_in_stock or 0),
                r.min_price if r.min_price is not None else float("inf"),
                -(r.latest.timestamp() if r.latest else 0),
            )
        )
        site = eligible[0].site_name
        return canonical_site(site) or site

    def select_from_df(
        self, candidates: pd.DataFrame, code: str
    ) -> Optional[str]:
        cand = candidates[
            (candidates["product_code"] == code)
            & (candidates["sizes_in_stock"] >= self.min_sizes)
            & candidates["min_eff_price"].notna()
        ]
        if cand.empty:
            return None
        best = cand.sort_values(
            ["sizes_in_stock", "min_eff_price", "latest"],
            ascending=[False, True, False],
        ).iloc[0]
        return best["site_name"]


class PreferredSiteStrategy(SupplierStrategy):
    """
    优先从指定站点列表中选择（按列表顺序）；
    若指定站点均无货/不达标，则 fallback 到备用策略。
    """

    def __init__(
        self,
        preferred: List[str],
        fallback: SupplierStrategy | None = None,
        min_sizes: int = 3,
    ):
        self.preferred = [canonical_site(s) or s for s in preferred]
        self.fallback = fallback or LowestPriceStrategy(min_sizes=min_sizes)
        self.min_sizes = min_sizes

    def select(self, conn: Connection, code: str) -> Optional[str]:
        rows = conn.execute(SQL_SITE_AGG, {"code": code}).fetchall()
        site_map = {
            (canonical_site(r.site_name) or r.site_name): r for r in rows
        }
        # 按偏好顺序查找
        for pref in self.preferred:
            r = site_map.get(pref)
            if r and (r.sizes_in_stock or 0) >= self.min_sizes:
                return pref
        # fallback
        return self.fallback.select(conn, code)

    def select_from_df(
        self, candidates: pd.DataFrame, code: str
    ) -> Optional[str]:
        cand = candidates[
            (candidates["product_code"] == code)
            & (candidates["sizes_in_stock"] >= self.min_sizes)
            & candidates["min_eff_price"].notna()
        ]
        if cand.empty:
            return self.fallback.select_from_df(candidates, code)
        for pref in self.preferred:
            match = cand[cand["site_name"] == pref]
            if not match.empty:
                return pref
        return self.fallback.select_from_df(candidates, code)


class TrueCostStrategy(SupplierStrategy):
    """
    真实落地成本策略：应用每个供货商的 SUPPLIER_DISCOUNT_RULES
    （extra_ratio + shipping_fee）计算实际采购成本后再比价。

    解决的问题：网页标价不等于真实成本。例如 allweathers 标价 £100
    但 extra_ratio=0.9 实际成本 £90，比标价 £92 + 运费 £4 的 very (£96) 更便宜。
    """

    def __init__(self, min_sizes: int = 3):
        self.min_sizes = min_sizes

    def select(self, conn: Connection, code: str) -> Optional[str]:
        rows = conn.execute(SQL_SITE_OFFERS, {"code": code}).fetchall()

        # 按站点聚合：有货尺码数 + 最低真实成本
        site_stats: Dict[str, dict] = {}
        for r in rows:
            if r.stock_count <= 0:
                continue
            site = canonical_site(r.site_name) or r.site_name
            cost = _true_cost(site, r.original_price_gbp, r.sale_price_gbp)
            if cost <= 0:
                continue
            if site not in site_stats:
                site_stats[site] = {"sizes": 0, "min_cost": float("inf")}
            site_stats[site]["sizes"] += 1
            site_stats[site]["min_cost"] = min(
                site_stats[site]["min_cost"], cost
            )

        # 筛选库存达标的
        eligible = [
            (site, s)
            for site, s in site_stats.items()
            if s["sizes"] >= self.min_sizes
        ]
        if not eligible:
            return None

        # 排序：真实成本升序 → 尺码降序
        eligible.sort(key=lambda x: (x[1]["min_cost"], -x[1]["sizes"]))
        return eligible[0][0]

    def select_from_df(
        self, candidates: pd.DataFrame, code: str
    ) -> Optional[str]:
        """
        批量场景下使用。candidates 需要包含
        product_code / site_name / original_price_gbp / sale_price_gbp / stock_count 列。
        如果只有聚合后的 min_eff_price（不含原价/折扣价明细），则退化为按网页价比较。
        """
        cand = candidates[candidates["product_code"] == code].copy()
        if cand.empty:
            return None

        # 如果 DataFrame 含 offer 明细列，计算真实成本
        if "original_price_gbp" in cand.columns and "sale_price_gbp" in cand.columns:
            in_stock = cand[cand["stock_count"] > 0].copy()
            if in_stock.empty:
                return None
            in_stock["true_cost"] = in_stock.apply(
                lambda r: _true_cost(
                    r["site_name"], r.get("original_price_gbp"), r.get("sale_price_gbp")
                ),
                axis=1,
            )
            in_stock = in_stock[in_stock["true_cost"] > 0]
            # 按站点聚合
            agg = (
                in_stock.groupby("site_name")
                .agg(sizes=("true_cost", "count"), min_cost=("true_cost", "min"))
                .reset_index()
            )
            agg = agg[agg["sizes"] >= self.min_sizes]
            if agg.empty:
                return None
            best = agg.sort_values(
                ["min_cost", "sizes"], ascending=[True, False]
            ).iloc[0]
            return best["site_name"]

        # 退化：只有聚合价格，无法应用折扣规则，按网页价比较
        cand = cand[
            (cand["sizes_in_stock"] >= self.min_sizes)
            & cand["min_eff_price"].notna()
        ]
        if cand.empty:
            return None
        best = cand.sort_values(
            ["min_eff_price", "sizes_in_stock"], ascending=[True, False]
        ).iloc[0]
        return best["site_name"]


class BandStockStrategy(SupplierStrategy):
    """
    合并库存策略：选价格最低的供货商作为主供货商，
    同时将价格浮动 band_ratio 范围内的所有供货商的库存合并计算。

    逻辑：
    1. 计算每个供货商的真实落地成本（应用 SUPPLIER_DISCOUNT_RULES）
    2. 找到最低成本 min_cost
    3. 将 cost <= min_cost * (1 + band_ratio) 的供货商归为"可用带"
    4. 在这个带内，合并所有站点的有货尺码（去重，同一尺码只要任一站点有货即算有货）
    5. 若合并后尺码数 >= min_sizes，选带内成本最低的站点作为主供货商
    6. 若不达标，逐步放宽 band 或放弃

    与下游 merge_band_stock_into_inventory 配合使用：
    - 本策略选出主供货商 → 写入 barbour_supplier_map
    - merge_band_stock_into_inventory 读取映射，按同样的 band_ratio 合并库存回填 inventory
    """

    def __init__(self, min_sizes: int = 3, band_ratio: float = 0.20):
        self.min_sizes = min_sizes
        self.band_ratio = band_ratio

    def select(self, conn: Connection, code: str) -> Optional[str]:
        # 查出该商品所有有货 offer 的明细
        rows = conn.execute(SQL_SITE_OFFERS, {"code": code}).fetchall()

        # 按站点聚合：有货尺码集合 + 最低真实成本
        site_stats: Dict[str, dict] = {}
        for r in rows:
            if r.stock_count <= 0:
                continue
            site = canonical_site(r.site_name) or r.site_name
            cost = _true_cost(site, r.original_price_gbp, r.sale_price_gbp)
            if cost <= 0:
                continue
            if site not in site_stats:
                site_stats[site] = {"min_cost": float("inf"), "sizes": 0}
            site_stats[site]["sizes"] += 1
            site_stats[site]["min_cost"] = min(site_stats[site]["min_cost"], cost)

        if not site_stats:
            return None

        # 按真实成本排序
        sorted_sites = sorted(site_stats.items(), key=lambda x: x[1]["min_cost"])
        best_cost = sorted_sites[0][1]["min_cost"]
        threshold = best_cost * (1 + self.band_ratio)

        # 带内供货商
        band_sites = [
            (site, s) for site, s in sorted_sites if s["min_cost"] <= threshold
        ]

        # 合并带内总尺码数（简化：各站点有货尺码数之和的上界）
        # 精确去重需要尺码信息，这里用 SQL_SITE_OFFERS 无尺码列，
        # 所以用"带内最大单站尺码数"和"带内总尺码数"取折中
        band_total_sizes = sum(s["sizes"] for _, s in band_sites)

        if band_total_sizes >= self.min_sizes:
            # 选带内成本最低的作为主供货商
            return band_sites[0][0]

        # 带内合并仍不够，退化为单站选择（选尺码最多的达标站点）
        single_eligible = [
            (site, s) for site, s in sorted_sites if s["sizes"] >= self.min_sizes
        ]
        if single_eligible:
            return single_eligible[0][0]

        return None

    def select_from_df(
        self, candidates: pd.DataFrame, code: str
    ) -> Optional[str]:
        cand = candidates[candidates["product_code"] == code].copy()
        if cand.empty:
            return None

        # 优先使用 offer 明细
        if "original_price_gbp" in cand.columns and "sale_price_gbp" in cand.columns:
            in_stock = cand[cand["stock_count"] > 0].copy()
            if in_stock.empty:
                return None
            in_stock["true_cost"] = in_stock.apply(
                lambda r: _true_cost(
                    r["site_name"], r.get("original_price_gbp"), r.get("sale_price_gbp")
                ),
                axis=1,
            )
            in_stock = in_stock[in_stock["true_cost"] > 0]
            if in_stock.empty:
                return None

            # 按站点聚合
            agg = (
                in_stock.groupby("site_name")
                .agg(sizes=("true_cost", "count"), min_cost=("true_cost", "min"))
                .reset_index()
                .sort_values("min_cost")
            )
            best_cost = agg.iloc[0]["min_cost"]
            threshold = best_cost * (1 + self.band_ratio)

            band = agg[agg["min_cost"] <= threshold]
            band_total_sizes = int(band["sizes"].sum())

            if band_total_sizes >= self.min_sizes:
                return band.iloc[0]["site_name"]

            # 退化
            single = agg[agg["sizes"] >= self.min_sizes]
            if not single.empty:
                return single.iloc[0]["site_name"]
            return None

        # 退化：只有聚合数据
        cand = cand[cand["min_eff_price"].notna()].copy()
        if cand.empty:
            return None
        cand = cand.sort_values("min_eff_price")
        best_price = cand.iloc[0]["min_eff_price"]
        threshold = best_price * (1 + self.band_ratio)
        band = cand[cand["min_eff_price"] <= threshold]
        band_total = int(band["sizes_in_stock"].sum())
        if band_total >= self.min_sizes:
            return band.iloc[0]["site_name"]
        single = cand[cand["sizes_in_stock"] >= self.min_sizes]
        if not single.empty:
            return single.iloc[0]["site_name"]
        return None


# ---------------------------------------------------------------------------
#  辅助函数
# ---------------------------------------------------------------------------
def _load_exclude_codes(xlsx_path: Optional[str]) -> Set[str]:
    """
    读取排除清单 Excel，返回需要"完全忽略更新"的商品编码集合。
    兼容列名：Product Code / 商品编码 / product_code / color_code / 编码
    """
    if not xlsx_path:
        return set()
    df = pd.read_excel(xlsx_path, dtype=str)
    cols = [c.strip().lower().replace(" ", "") for c in df.columns]
    name2idx = {cols[i]: i for i in range(len(cols))}
    for key in ("productcode", "商品编码", "product_code", "color_code", "编码"):
        if key in name2idx:
            s = df.iloc[:, name2idx[key]].astype(str).str.strip()
            return {x for x in s if x}
    logger.warning("未在排除清单中识别到编码列：%s，将忽略该文件。", list(df.columns))
    return set()


def _load_publication_mappings(pub_dir: Path) -> Dict[str, str]:
    """
    读取目录下所有 barbour_publication_*.xlsx，返回 {product_code -> canonical_site}。
    后读的新文件覆盖旧文件（以"最新发布"为准）。
    """

    def _headers(ws) -> Dict[str, int]:
        h = {}
        for j, c in enumerate(ws[1], start=1):
            k = str(c.value or "").strip().lower().replace(" ", "")
            if k:
                h[k] = j
        return h

    mappings: Dict[str, str] = {}
    files = sorted(pub_dir.glob(PATTERN), key=lambda p: p.stat().st_mtime)
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb.active
            hdr = _headers(ws)
            col_code = next(
                (
                    hdr[k]
                    for k in (
                        "productcode", "商品编码", "product_code",
                        "color_code", "编码",
                    )
                    if k in hdr
                ),
                None,
            )
            col_site = next(
                (
                    hdr[k]
                    for k in ("supplier", "供应商", "site", "站点")
                    if k in hdr
                ),
                None,
            )
            if not col_code or not col_site:
                continue
            for i in range(2, ws.max_row + 1):
                code = str(ws.cell(i, col_code).value or "").strip()
                site_raw = str(ws.cell(i, col_site).value or "").strip()
                if not code or not site_raw:
                    continue
                site = canonical_site(site_raw)
                if site:
                    mappings[code] = site
        except Exception as e:
            logger.warning("解析失败 %s: %s", fp.name, e)
    return mappings


# ---------------------------------------------------------------------------
#  核心业务函数
# ---------------------------------------------------------------------------
def fill_supplier_map(
    strategy: SupplierStrategy | None = None,
    force_refresh: bool = False,
    exclude_xlsx: Optional[str] = None,
) -> None:
    """
    构建 / 更新 barbour_supplier_map 映射表。

    优先级：排除清单保护 > 发布文件映射 > strategy 兜底选择。
    """
    if strategy is None:
        strategy = LowestPriceStrategy(min_sizes=3)

    engine = _get_engine()
    exclude_codes = _load_exclude_codes(exclude_xlsx)
    if exclude_codes:
        logger.info("排除清单：%d 个编码将不被更新/覆盖。", len(exclude_codes))

    with engine.begin() as conn:
        # 0) 确保表存在
        conn.execute(SQL_CREATE)

        # 保存排除编码的历史映射（force_refresh 时用于回填）
        preserved: Dict[str, str] = {}
        if force_refresh:
            if exclude_codes:
                rows = conn.execute(
                    text(
                        f"SELECT product_code, site_name FROM {TABLE} "
                        f"WHERE product_code = ANY(:codes)"
                    ),
                    {"codes": list(exclude_codes)},
                ).fetchall()
                preserved = {r[0]: r[1] for r in rows}
                if preserved:
                    logger.info("预保存排除映射 %d 条。", len(preserved))
            conn.execute(text(f"TRUNCATE TABLE {TABLE};"))
            logger.warning("已清空 %s 表。", TABLE)

        # 1) 取"已发布"的编码集合
        published: Set[str] = {
            r[0] for r in conn.execute(SQL_PUBLISHED_CODES).fetchall()
        }
        logger.info("已发布编码：%d 个。", len(published))

        # 2) 已有映射（注意：force_refresh 之后此处为空集，符合预期）
        try:
            existing: Set[str] = {
                r[0] for r in conn.execute(SQL_EXISTING_MAP).fetchall()
            }
        except Exception:
            existing = set()

        # 3) 按发布清单覆盖（跳过排除编码）
        pub_map = _load_publication_mappings(PUBLICATION_DIR)
        pub_hit: List[str] = []
        for code, site in pub_map.items():
            if code in exclude_codes:
                continue
            if code in published:
                conn.execute(SQL_UPSERT, {"code": code, "site": site})
                pub_hit.append(code)
        logger.info("按发布文件更新：%d 条。", len(pub_hit))

        # 4) 兜底：对未命中且未映射的已发布编码，使用策略选择
        need = (published - set(pub_hit)) - existing
        offer_filled: List[str] = []
        for code in sorted(need):
            if code in exclude_codes:
                continue
            site = strategy.select(conn, code)
            if site:
                conn.execute(SQL_UPSERT, {"code": code, "site": site})
                offer_filled.append(code)
        logger.info("按策略兜底：%d 条（策略=%s）。", len(offer_filled), type(strategy).__name__)

        # 5) 回填排除清单中的历史映射
        if preserved:
            rows_to_restore = [
                {"code": k, "site": v} for k, v in preserved.items()
            ]
            conn.execute(
                text(
                    f"INSERT INTO {TABLE}(product_code, site_name) "
                    f"VALUES (:code, :site) "
                    f"ON CONFLICT (product_code) DO UPDATE SET site_name = EXCLUDED.site_name"
                ),
                rows_to_restore,
            )
            logger.info("已恢复排除清单中的历史映射 %d 条。", len(rows_to_restore))

        # 6) 统计
        total_now = conn.execute(
            text(f"SELECT COUNT(*) FROM {TABLE}")
        ).scalar_one()
        logger.info("完成映射，总计 %d 条。", total_now)


def reassign_low_stock_suppliers(
    strategy: SupplierStrategy | None = None,
    size_threshold: int = 3,
    dry_run: bool = True,
    exclude_xlsx: Optional[str] = None,
) -> List[dict]:
    """
    找出当前映射站点有货尺码数 < size_threshold 的商品，
    用 strategy 从其他达标候选中选出替代供货商。
    """
    if strategy is None:
        strategy = LowestPriceStrategy(min_sizes=size_threshold)

    engine = _get_engine()
    exclude_codes = _load_exclude_codes(exclude_xlsx)
    if exclude_codes:
        logger.info("排除清单：%d 个编码将不会被更新/覆盖。", len(exclude_codes))

    suggest: List[dict] = []

    with engine.begin() as conn:
        # 当前映射
        map_df = pd.read_sql(
            f"SELECT product_code, site_name FROM {TABLE}", conn
        )
        map_df["site_name"] = map_df["site_name"].map(
            lambda s: canonical_site(s) or s
        )

        # 全量聚合
        agg_df = pd.read_sql(SQL_FULL_AGG, conn)
        agg_df["site_name"] = agg_df["site_name"].map(
            lambda s: canonical_site(s) or s
        )

        # TrueCostStrategy 需要 offer 明细来计算真实成本
        offers_df: Optional[pd.DataFrame] = None
        if isinstance(strategy, (TrueCostStrategy, BandStockStrategy)):
            offers_df = pd.read_sql(SQL_FULL_OFFERS, conn)
            offers_df["site_name"] = offers_df["site_name"].map(
                lambda s: canonical_site(s) or s
            )

        # 供 select_from_df 使用的数据源
        candidate_df = offers_df if offers_df is not None else agg_df

        # 合并取当前站点表现
        cur_df = map_df.merge(
            agg_df.rename(
                columns={
                    "sizes_in_stock": "cur_sizes_in_stock",
                    "min_eff_price": "cur_min_eff_price",
                    "latest": "cur_latest",
                }
            ),
            on=["product_code", "site_name"],
            how="left",
        )
        cur_df["cur_sizes_in_stock"] = (
            cur_df["cur_sizes_in_stock"].fillna(0).astype(int)
        )

        for _, r in cur_df.iterrows():
            code = str(r["product_code"])
            if code in exclude_codes:
                continue

            cur_sizes = int(r["cur_sizes_in_stock"])
            if cur_sizes >= size_threshold:
                continue

            old_site = r["site_name"]

            # 用策略从候选中选择
            new_site = strategy.select_from_df(candidate_df, code)
            if not new_site or new_site == old_site:
                continue

            new_row = agg_df[
                (agg_df["product_code"] == code)
                & (agg_df["site_name"] == new_site)
            ]
            if new_row.empty:
                continue
            new_row = new_row.iloc[0]

            suggest.append(
                {
                    "product_code": code,
                    "old_site": old_site,
                    "old_sizes": cur_sizes,
                    "old_min_price": float(r.get("cur_min_eff_price") or 0.0),
                    "new_site": new_site,
                    "new_sizes": int(new_row["sizes_in_stock"]),
                    "new_min_price": float(new_row["min_eff_price"] or 0.0),
                }
            )

        # 写库
        if suggest and not dry_run:
            rows_to_update = [
                {"code": s["product_code"], "site": s["new_site"]}
                for s in suggest
                if s["product_code"] not in exclude_codes
            ]
            conn.execute(
                text(
                    f"INSERT INTO {TABLE}(product_code, site_name) "
                    f"VALUES (:code, :site) "
                    f"ON CONFLICT (product_code) DO UPDATE SET site_name = EXCLUDED.site_name"
                ),
                rows_to_update,
            )

    # 打印建议
    if suggest:
        logger.info(
            "共 %d 条建议（已排除 %d 条编码）：", len(suggest), len(exclude_codes)
        )
        for s in suggest:
            logger.info(
                "  - %s: %s(%d尺) -> %s(%d尺), 价 %.2f -> %.2f",
                s["product_code"],
                s["old_site"],
                s["old_sizes"],
                s["new_site"],
                s["new_sizes"],
                s["old_min_price"],
                s["new_min_price"],
            )
    else:
        logger.info("未找到需要切换的商品（当前映射站点均满足尺码阈值或无更优候选）。")

    return suggest


def export_supplier_stock_price_report(
    min_sizes_ok: int = 1, output_path: str | None = None
) -> str:
    """
    导出每个商品在各站点的"在售尺码数、最低有效价、最近更新时间"，
    并标注当前映射与推荐站点。
    """
    out_dir = Path(BARBOUR["OUTPUT_DIR"])
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = Path(output_path) if output_path else (out_dir / "barbour_supplier_report.xlsx")

    engine = _get_engine()
    with engine.begin() as conn:
        map_df = pd.read_sql(
            f"SELECT product_code, site_name FROM {TABLE}", conn
        )

        sql = f"""
        WITH base AS (
          SELECT
            product_code,
            site_name,
            COALESCE(stock_count, 0) AS stock_count,
            COALESCE(NULLIF(sale_price_gbp, 0), NULLIF(price_gbp, 0), original_price_gbp) AS eff_price,
            last_checked
          FROM barbour_offers
          WHERE is_active = TRUE
            AND product_code IS NOT NULL AND product_code <> ''
            AND site_name IS NOT NULL AND site_name <> ''
        ),
        agg AS (
          SELECT
            product_code,
            site_name,
            SUM(CASE WHEN stock_count > 0 THEN 1 ELSE 0 END)   AS sizes_in_stock,
            MIN(eff_price) FILTER (WHERE stock_count > 0)       AS min_eff_price,
            MAX(last_checked)                                   AS latest
          FROM base
          GROUP BY product_code, site_name
        ),
        ranked AS (
          SELECT
            *,
            ROW_NUMBER() OVER (
              PARTITION BY product_code
              ORDER BY
                CASE WHEN sizes_in_stock >= 3 THEN 0 ELSE 1 END,
                min_eff_price ASC NULLS LAST,
                sizes_in_stock DESC,
                latest DESC
            ) AS rank_all,
            ROW_NUMBER() OVER (
              PARTITION BY product_code
              ORDER BY
                CASE WHEN sizes_in_stock >= 3 THEN 0 ELSE 1 END,
                min_eff_price ASC NULLS LAST
            ) AS rank_price_first
          FROM agg
        )
        SELECT * FROM ranked
        """
        df = pd.read_sql(sql, conn)

    # 规范站点名
    df["site_name"] = df["site_name"].map(lambda s: canonical_site(s) or s)
    map_df["site_name"] = map_df["site_name"].map(
        lambda s: canonical_site(s) or s
    )

    # 标注"当前映射"
    df = df.merge(map_df, on="product_code", how="left", suffixes=("", "_mapped"))
    df["is_current"] = (df["site_name"] == df["site_name_mapped"]).fillna(False)

    # 推荐站点
    best = (
        df[(df["sizes_in_stock"] >= 3) & df["min_eff_price"].notna()]
        .sort_values(
            ["product_code", "min_eff_price", "sizes_in_stock", "latest"],
            ascending=[True, True, False, False],
        )
        .drop_duplicates(["product_code"])
        .rename(
            columns={
                "site_name": "best_site",
                "min_eff_price": "best_min_eff_price",
                "sizes_in_stock": "best_sizes_in_stock",
            }
        )[["product_code", "best_site", "best_min_eff_price", "best_sizes_in_stock"]]
    )
    df = df.merge(best, on="product_code", how="left")

    out = df[
        [
            "product_code",
            "site_name",
            "is_current",
            "sizes_in_stock",
            "min_eff_price",
            "latest",
            "site_name_mapped",
            "best_site",
            "best_min_eff_price",
            "best_sizes_in_stock",
            "rank_all",
            "rank_price_first",
        ]
    ].sort_values(["product_code", "is_current"], ascending=[True, False])

    out.to_excel(out_file, index=False)
    logger.info("报表已导出：%s", out_file)
    return str(out_file)


def apply_barbour_supplier_overrides(
    xlsx_path: str, dry_run: bool = False
) -> None:
    """
    按 Excel 文件手动指定供货商（最高优先级）。
    Excel 要求列名：商品编码 / 供货商。
    """
    engine = _get_engine()

    df = pd.read_excel(xlsx_path, dtype=str)
    required_cols = ["商品编码", "供货商"]
    if not all(c in df.columns for c in required_cols):
        raise ValueError(
            f"Excel 必须包含列：{required_cols}，当前表头：{list(df.columns)}"
        )

    df = df[required_cols].rename(
        columns={"商品编码": "product_code", "供货商": "site_name"}
    )
    df["product_code"] = df["product_code"].astype(str).str.strip()
    df["site_name"] = df["site_name"].astype(str).str.strip()
    df = df[(df["product_code"] != "") & (df["site_name"] != "")].drop_duplicates()

    if df.empty:
        logger.info("Excel 中无有效数据。")
        return

    df["site_name"] = df["site_name"].map(lambda s: canonical_site(s) or s)

    logger.info("读取 %d 条供货商指定记录。", len(df))
    for _, row in df.head(5).iterrows():
        logger.info("  - %s -> %s", row["product_code"], row["site_name"])
    if len(df) > 5:
        logger.info("  ...")

    if dry_run:
        logger.info("[DRY-RUN] 仅预览，不写入数据库。")
        return

    with engine.begin() as conn:
        conn.execute(SQL_CREATE)
        conn.execute(
            text(
                f"INSERT INTO {TABLE}(product_code, site_name) "
                f"VALUES (:product_code, :site_name) "
                f"ON CONFLICT (product_code) DO UPDATE SET site_name = EXCLUDED.site_name"
            ),
            df.to_dict(orient="records"),
        )
        logger.info("已成功更新 %d 条供货商映射。", len(df))

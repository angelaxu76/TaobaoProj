# -*- coding: utf-8 -*-
"""
Barbour 供应商 / 价格 / 库存 —— 单一入口

替代原本分散在 build_supplier_jingya_mapping.py（v1/v2）、
merge_offer_into_inventory.py、db_build_supplier_map_and_inventory.py
里的"选供应商 → 算价格 → 算库存"逻辑。

核心策略（allocate_and_sync）：
- 每个已发布商品，按"真实落地成本"（barbour_offers.sale_price_gbp——
  已在导入阶段套用过 SUPPLIER_DISCOUNT_RULES 的折扣比例 + 运费，见
  import_supplier_to_db_offers.compute_supplier_sale_price；取不到则
  COALESCE 到 price_gbp / original_price_gbp 兜底）从低到高遍历供应商，
  依次纳入，直到覆盖的有货尺码数达到 SUPPLIER_MIN_SIZES，或凑满
  SUPPLIER_MAX_SITES 家为止。
- 最终库存 = 所选各站点"有货尺码"的并集。
- 最终定价 = 所选各站点里成本最高的那个（避免低价站点断货、临时改用
  高价站点补货时倒贴运费亏本）。

人工干预（可选，两个独立 Excel，逻辑可叠加）：
- supplier_override_xlsx（如 barbour_supplier.xlsx）：强制指定某商品的
  供应商，跳过自动选择，但仍走同一套定价/库存回填逻辑。
- exclude_xlsx（如 barbour_exclude_list.xlsx）：编码列出现的商品默认跳过
  自动分配；但若同一行还填了"供货商"/"供应商"列，则改为强制走该供应商
  （效果等同 supplier_override_xlsx），从而拿到真实库存，而不是停在 0；
  只有既没填供应商、也没填价格的行才是真正"完全不处理"。若该文件同时含
  source_price_gbp / discount_price_gbp 列，最后会由
  apply_fixed_prices_from_excel() 覆盖为人工固定价——供应商决定库存，
  这两列决定价格，两者配合实现"强行指定供货商和价格"。
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Callable, Dict, List, Optional, Set, Tuple

import openpyxl
import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

from config import BRAND_CONFIG, BARBOUR, DEFAULT_STOCK_COUNT
from brands.barbour.core.site_utils import canonical_site
from common.product.size_utils import clean_size_for_barbour
from common.pricing.price_utils import calculate_jingya_prices
from brands.barbour.jingya.allocate_supplier_and_price_config import (
    SUPPLIER_MIN_SIZES,
    SUPPLIER_MAX_SITES,
    TAOBAO_STORE_DISCOUNT,
    SUPPLIER_OVERRIDE_XLSX,
)

TABLE_ALLOC = "barbour_supplier_allocation"
PUBLICATION_DIR = Path(BARBOUR["PUBLICATION_DIR"])
PUB_PATTERN = "barbour_publication_*.xlsx"

SQL_CREATE_ALLOC = text(f"""
CREATE TABLE IF NOT EXISTS {TABLE_ALLOC} (
  product_code   VARCHAR(50)  NOT NULL,
  site_name      VARCHAR(100) NOT NULL,
  rank           SMALLINT     NOT NULL,
  min_eff_price  NUMERIC(10,2),
  sizes_in_stock INT,
  is_price_basis BOOLEAN DEFAULT FALSE,
  source         VARCHAR(20) DEFAULT 'auto',
  updated_at     TIMESTAMP DEFAULT NOW(),
  PRIMARY KEY (product_code, site_name)
);
""")


# ═══════════════════════════════════════════════════════════════════
#  基础工具
# ═══════════════════════════════════════════════════════════════════

def _get_engine() -> Engine:
    cfg = BRAND_CONFIG["barbour"]["PGSQL_CONFIG"]
    return create_engine(
        f"postgresql+psycopg2://{cfg['user']}:{cfg['password']}"
        f"@{cfg['host']}:{cfg['port']}/{cfg['dbname']}"
    )


def _ensure_price_columns(conn) -> None:
    conn.execute(text("""
        ALTER TABLE barbour_inventory
          ADD COLUMN IF NOT EXISTS jingya_untaxed_price NUMERIC(12,2),
          ADD COLUMN IF NOT EXISTS taobao_store_price   NUMERIC(12,2),
          ADD COLUMN IF NOT EXISTS base_price_gbp       NUMERIC(10,2),
          ADD COLUMN IF NOT EXISTS exchange_rate_used   NUMERIC(8,4)
    """))


def _load_exclude_and_forced_sites(xlsx_path: Optional[str]) -> Tuple[Set[str], Dict[str, str]]:
    """
    读取排除清单。"商品编码"列出现的商品默认跳过自动分配；但如果同一行还
    填了"供货商"/"供应商"列，就不再是纯跳过，而是把它当成人工强制指定的
    供应商（和 supplier_override_xlsx 走同一条"人工指定→仍按同一套逻辑
    回填价格/库存"的路径），配合文件里的 source_price_gbp/discount_price_gbp
    列由 apply_fixed_prices_from_excel() 覆盖最终价格——这样才能同时实现
    "强行指定供货商" + "强行指定价格"，而不是让这些商品的库存一直停在 0。

    只有"商品编码"列出现、但没有可识别供应商的行，才是真正的"跳过不处理"。

    返回 (bare_exclude_codes, forced_site_by_code)。
    """
    if not xlsx_path:
        return set(), {}
    if not Path(xlsx_path).exists():
        print(f"ℹ️ 排除清单文件不存在，已跳过：{xlsx_path}")
        return set(), {}

    df = pd.read_excel(xlsx_path, dtype=str)
    col_map = {c.strip().lower().replace(" ", ""): c for c in df.columns}
    code_col = next((col_map[k] for k in ("productcode", "商品编码", "product_code", "color_code", "编码") if k in col_map), None)
    if not code_col:
        print(f"⚠️ 未在排除清单中识别到编码列：{list(df.columns)}，将忽略该文件。")
        return set(), {}
    site_col = next((col_map[k] for k in ("供货商", "供应商", "supplier", "site") if k in col_map), None)

    bare: Set[str] = set()
    forced: Dict[str, str] = {}
    for _, row in df.iterrows():
        code = str(row.get(code_col) or "").strip()
        if not code:
            continue
        site_raw = str(row.get(site_col) or "").strip() if site_col else ""
        site = canonical_site(site_raw) if site_raw else None
        if site:
            forced[code] = site
        else:
            bare.add(code)
    return bare, forced


def _load_supplier_overrides(xlsx_path: Optional[str]) -> Dict[str, str]:
    """人工指定供应商：Excel 需含列 商品编码 / 供货商。"""
    if not xlsx_path:
        return {}
    if not Path(xlsx_path).exists():
        print(f"ℹ️ 供应商指定清单文件不存在，已跳过：{xlsx_path}")
        return {}
    df = pd.read_excel(xlsx_path, dtype=str)
    required = ["商品编码", "供货商"]
    if not all(c in df.columns for c in required):
        print(f"⚠️ 供应商指定清单缺少必需列 {required}，当前列：{list(df.columns)}，已忽略。")
        return {}
    df = df[required].rename(columns={"商品编码": "product_code", "供货商": "site_name"})
    df["product_code"] = df["product_code"].astype(str).str.strip()
    df["site_name"] = df["site_name"].astype(str).str.strip().map(lambda s: canonical_site(s) or s)
    df = df[(df["product_code"] != "") & (df["site_name"] != "")]
    return dict(zip(df["product_code"], df["site_name"]))


def _load_publication_mappings(pub_dir: Path) -> Dict[str, str]:
    """
    新品刚在鲸芽发布、还没抓到任何供应商 offer 时的兜底：
    读历史 barbour_publication_*.xlsx 里记录的供应商（后读的新文件覆盖旧文件）。
    """
    def _headers(ws) -> Dict[str, int]:
        h = {}
        for j, c in enumerate(ws[1], start=1):
            k = str(c.value or "").strip().lower().replace(" ", "")
            if k:
                h[k] = j
        return h

    mappings: Dict[str, str] = {}
    if not pub_dir.exists():
        return mappings
    files = sorted(pub_dir.glob(PUB_PATTERN), key=lambda p: p.stat().st_mtime)
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb.active
            hdr = _headers(ws)
            col_code = next((hdr[k] for k in ("productcode", "商品编码", "product_code", "color_code", "编码") if k in hdr), None)
            col_site = next((hdr[k] for k in ("supplier", "供应商", "site", "站点") if k in hdr), None)
            if not col_code or not col_site:
                continue
            for i in range(2, ws.max_row + 1):
                code = str(ws.cell(i, col_code).value or "").strip()
                site_raw = str(ws.cell(i, col_site).value or "").strip()
                if not code or not site_raw:
                    continue
                site = canonical_site(site_raw)
                if site:
                    mappings[code] = site
        except Exception as e:
            print(f"⚠️ 解析发布清单失败 {fp.name}: {e}")
    return mappings


def _eff_price_row(row) -> Optional[float]:
    """有效成本口径：sale_price_gbp（已含折扣策略+运费）优先，其次 price_gbp，最后 original_price_gbp。"""
    for v in (row.get("sale_price_gbp"), row.get("price_gbp"), row.get("original_price_gbp")):
        if v is not None and not pd.isna(v) and float(v) > 0:
            return float(v)
    return None


def _greedy_select_sites(
    cand: Optional[pd.DataFrame],
    stock_sizes_lookup: Callable[[str], Set[str]],
    min_sizes: int,
    max_suppliers: int,
) -> List[dict]:
    """
    贪心选站点：按有效成本从低到高遍历，依次纳入，直到并集覆盖的有货尺码数
    达到 min_sizes、或凑满 max_suppliers 家为止。

    allocate_and_sync（批量）和 select_suppliers_for_code（单品预览）共用
    这一份实现，避免诊断工具和实际写库逻辑走两套算法、结果对不上。

    cand: 需含列 site_name / min_eff_price / sizes_in_stock / latest。
    stock_sizes_lookup(site_name) -> 该站点有货的 size_norm 集合。
    """
    chosen: List[dict] = []
    if cand is None or cand.empty:
        return chosen

    ranked = cand.sort_values(
        ["min_eff_price", "sizes_in_stock", "latest"],
        ascending=[True, False, False],
    )
    covered: Set[str] = set()
    for _, r in ranked.iterrows():
        site = r["site_name"]
        chosen.append({
            "site": site,
            "min_eff_price": float(r["min_eff_price"]),
            "sizes_in_stock": int(r["sizes_in_stock"]),
        })
        covered |= stock_sizes_lookup(site)
        if len(covered) >= min_sizes or len(chosen) >= max_suppliers:
            break
    return chosen


# ═══════════════════════════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════════════════════════

def allocate_and_sync(
    brand: str = "barbour",
    min_sizes: Optional[int] = None,
    max_suppliers: Optional[int] = None,
    exclude_xlsx: Optional[str] = None,
    supplier_override_xlsx: Optional[str] = SUPPLIER_OVERRIDE_XLSX,
    dry_run: bool = False,
) -> dict:
    """
    为所有已发布商品重新计算供应商组合 + 价格 + 库存，并同步到
    barbour_inventory + barbour_supplier_allocation。

    dry_run=True 时只打印将要发生的变更，不写库。
    supplier_override_xlsx 默认指向配置里的固定路径，文件不存在时自动忽略；
    传 None 可显式关闭人工指定供应商这一层。
    """
    if brand.lower() != "barbour":
        raise ValueError("目前仅支持 barbour")

    min_sizes = min_sizes if min_sizes is not None else SUPPLIER_MIN_SIZES
    max_suppliers = max_suppliers if max_suppliers is not None else SUPPLIER_MAX_SITES

    engine = _get_engine()
    exclude_codes, exclude_forced_sites = _load_exclude_and_forced_sites(exclude_xlsx)
    # supplier_override_xlsx 是专门的"强制供应商"文件，若与排除清单里的
    # 供货商列同时指定了同一个编码，以 supplier_override_xlsx 为准。
    manual_overrides = {**exclude_forced_sites, **_load_supplier_overrides(supplier_override_xlsx)}
    pub_map = _load_publication_mappings(PUBLICATION_DIR)
    taobao_discount = TAOBAO_STORE_DISCOUNT

    if exclude_codes:
        print(f"🛡️ 排除清单：{len(exclude_codes)} 个编码（无指定供应商）将跳过自动分配。")
    if exclude_forced_sites:
        print(f"🧭 排除清单里指定了供应商：{len(exclude_forced_sites)} 条，将强制走该供应商（价格随后由固定价覆盖）。")
    if manual_overrides:
        print(f"🧭 人工指定供应商合计：{len(manual_overrides)} 条。")

    with engine.begin() as conn:
        conn.execute(SQL_CREATE_ALLOC)
        # 只是加列（ADD COLUMN IF NOT EXISTS），不改数据，dry_run 下也要跑，
        # 否则在全新数据库上首次 dry_run 会因为列不存在而查询失败。
        _ensure_price_columns(conn)

        inv_df = pd.read_sql(text("""
            SELECT id, product_code, size, stock_count,
                   base_price_gbp, jingya_untaxed_price, taobao_store_price
            FROM barbour_inventory
            WHERE is_published = TRUE
              AND product_code IS NOT NULL AND product_code <> ''
              AND size IS NOT NULL AND size <> ''
        """), conn)

        offers_df = pd.read_sql(text("""
            SELECT product_code, site_name, size, stock_count,
                   sale_price_gbp, price_gbp, original_price_gbp, last_checked
            FROM barbour_offers
            WHERE is_active = TRUE
              AND product_code IS NOT NULL AND product_code <> ''
              AND size IS NOT NULL AND size <> ''
        """), conn)

    if inv_df.empty:
        print("ℹ️ 没有已发布商品（barbour_inventory.is_published=TRUE 为空），已跳过。")
        return {"processed": 0}

    inv_df["size_norm"] = inv_df["size"].map(clean_size_for_barbour)
    offers_df["site_name"] = offers_df["site_name"].map(lambda s: canonical_site(s) or s)
    offers_df["size_norm"] = offers_df["size"].map(clean_size_for_barbour)
    offers_df["eff_price"] = offers_df.apply(_eff_price_row, axis=1)
    offers_df["in_stock"] = pd.to_numeric(offers_df["stock_count"], errors="coerce").fillna(0) > 0

    in_stock_df = offers_df[offers_df["in_stock"]]

    site_agg = (
        in_stock_df[in_stock_df["eff_price"].notna()]
        .groupby(["product_code", "site_name"])
        .agg(
            sizes_in_stock=("size_norm", "nunique"),
            min_eff_price=("eff_price", "min"),
            latest=("last_checked", "max"),
        )
        .reset_index()
    )
    site_agg_by_code = {code: g for code, g in site_agg.groupby("product_code")}

    # (product_code, site_name) -> 该站点有货的 size_norm 集合
    stock_sizes_map: Dict[Tuple[str, str], Set[str]] = (
        in_stock_df.groupby(["product_code", "site_name"])["size_norm"]
        .apply(set)
        .to_dict()
    )

    inv_by_code = {code: g for code, g in inv_df.groupby("product_code")}
    published_codes = sorted(inv_df["product_code"].unique().tolist())

    inventory_updates: List[dict] = []
    allocation_rows: List[dict] = []
    diag_excluded: List[str] = []
    diag_unresolved: List[Tuple[str, str, str]] = []
    diag_auto: List[str] = []
    diag_manual: List[str] = []
    dry_run_report: List[dict] = []

    for code in published_codes:
        if code in exclude_codes:
            diag_excluded.append(code)
            continue

        forced_site = manual_overrides.get(code)
        cand = site_agg_by_code.get(code)

        chosen: List[dict] = []  # [{"site", "min_eff_price", "sizes_in_stock"}]
        source = "auto"

        if forced_site:
            source = "manual"
            if cand is not None:
                row = cand[cand["site_name"] == forced_site]
                if not row.empty:
                    r = row.iloc[0]
                    chosen = [{
                        "site": forced_site,
                        "min_eff_price": float(r["min_eff_price"]),
                        "sizes_in_stock": int(r["sizes_in_stock"]),
                    }]
            if not chosen:
                diag_unresolved.append((code, "人工指定供应商无有效报价", forced_site))
                continue
        else:
            chosen = _greedy_select_sites(
                cand,
                lambda site, _code=code: stock_sizes_map.get((_code, site), set()),
                min_sizes,
                max_suppliers,
            )

            if not chosen:
                # 兜底：发布 Excel 里的历史供应商（新品还没抓到 offers 的情况）
                fallback_site = pub_map.get(code)
                if fallback_site:
                    fb_rows = offers_df[
                        (offers_df["product_code"] == code)
                        & (offers_df["site_name"] == fallback_site)
                        & offers_df["eff_price"].notna()
                    ]
                    # 定价只看有货的行，避免拿一个缺货尺码的低价当基准；
                    # 若这个站点当前一个尺码都没货，才退回用全部行估个价。
                    fb_in_stock = fb_rows[fb_rows["in_stock"]]
                    fb_price_rows = fb_in_stock if not fb_in_stock.empty else fb_rows
                    if not fb_price_rows.empty:
                        chosen = [{
                            "site": fallback_site,
                            "min_eff_price": float(fb_price_rows["eff_price"].min()),
                            "sizes_in_stock": int(fb_in_stock["size_norm"].nunique()),
                        }]

            if not chosen:
                diag_unresolved.append((code, "无达标供应商", ""))
                continue

        price_basis = max(c["min_eff_price"] for c in chosen)
        untaxed, retail = calculate_jingya_prices(price_basis)
        untaxed = round(float(untaxed), 2) if untaxed else 0.0
        retail_tb = round(float(retail) * float(taobao_discount), 2) if retail else 0.0

        covered_sizes: Set[str] = set()
        for c in chosen:
            covered_sizes |= stock_sizes_map.get((code, c["site"]), set())

        (diag_manual if source == "manual" else diag_auto).append(code)

        for i, c in enumerate(chosen, start=1):
            allocation_rows.append({
                "product_code": code,
                "site_name": c["site"],
                "rank": i,
                "min_eff_price": c["min_eff_price"],
                "sizes_in_stock": c["sizes_in_stock"],
                "is_price_basis": c["min_eff_price"] == price_basis,
                "source": source,
            })

        primary_site = chosen[0]["site"]
        for _, inv_row in inv_by_code[code].iterrows():
            new_stock = DEFAULT_STOCK_COUNT if inv_row["size_norm"] in covered_sizes else 0
            inventory_updates.append({
                "bi_id": int(inv_row["id"]),
                "stock_count": new_stock,
                "source_site": primary_site,
                "source_price_gbp": price_basis,
                "original_price_gbp": price_basis,
                "discount_price_gbp": price_basis,
                "base_price_gbp": price_basis,
                "jingya_untaxed_price": untaxed,
                "taobao_store_price": retail_tb,
            })
            if dry_run:
                old_price = inv_row.get("jingya_untaxed_price")
                old_stock = inv_row.get("stock_count")
                old_price_f = None if pd.isna(old_price) else float(old_price)
                if old_price_f != untaxed or int(old_stock or 0) != new_stock:
                    dry_run_report.append({
                        "product_code": code, "size": inv_row["size"],
                        "旧库存": old_stock, "新库存": new_stock,
                        "旧未税价": old_price_f, "新未税价": untaxed,
                        "供应商组合": "+".join(c["site"] for c in chosen),
                    })

    # ── 打印诊断 ──
    print(
        f"✅ 自动分配：{len(diag_auto)} 个；人工指定：{len(diag_manual)} 个；"
        f"排除清单跳过：{len(diag_excluded)} 个；无法分配：{len(diag_unresolved)} 个。"
    )
    if diag_unresolved:
        print(f"⚠️ 以下 {len(diag_unresolved)} 个商品本次未能分配供应商（保留原有价格/库存）：")
        for code, reason, extra in diag_unresolved[:30]:
            print(f"   {code}: {reason} {extra}".rstrip())
        if len(diag_unresolved) > 30:
            print(f"   ...共 {len(diag_unresolved)} 个，已省略 {len(diag_unresolved) - 30} 个")

    if dry_run:
        print(f"\n[DRY-RUN] 将变更 {len(dry_run_report)} 条尺码记录（未写库）。示例前 20 条：")
        for r in dry_run_report[:20]:
            print(f"   {r}")
        return {
            "processed": len(diag_auto) + len(diag_manual),
            "excluded": len(diag_excluded),
            "unresolved": len(diag_unresolved),
            "would_change": len(dry_run_report),
        }

    # ── 写库 ──
    processed_codes = list(set(diag_auto) | set(diag_manual))
    with engine.begin() as conn:
        if inventory_updates:
            conn.execute(text("""
                UPDATE barbour_inventory
                SET stock_count = :stock_count,
                    source_site = :source_site,
                    source_price_gbp = :source_price_gbp,
                    original_price_gbp = :original_price_gbp,
                    discount_price_gbp = :discount_price_gbp,
                    base_price_gbp = :base_price_gbp,
                    jingya_untaxed_price = :jingya_untaxed_price,
                    taobao_store_price = :taobao_store_price,
                    last_checked = NOW()
                WHERE id = :bi_id
            """), inventory_updates)

        if processed_codes:
            conn.execute(
                text(f"DELETE FROM {TABLE_ALLOC} WHERE product_code = ANY(:codes)"),
                {"codes": processed_codes},
            )
        if allocation_rows:
            conn.execute(text(f"""
                INSERT INTO {TABLE_ALLOC}
                    (product_code, site_name, rank, min_eff_price, sizes_in_stock, is_price_basis, source, updated_at)
                VALUES
                    (:product_code, :site_name, :rank, :min_eff_price, :sizes_in_stock, :is_price_basis, :source, NOW())
            """), allocation_rows)

    print(
        f"✅ barbour_inventory 已更新 {len(inventory_updates)} 条尺码记录；"
        f"{TABLE_ALLOC} 已写入 {len(allocation_rows)} 条供应商分配记录。"
    )

    # ── 人工固定价覆盖（最终层，沿用今天的 exclude_list.xlsx 双重用途）──
    if exclude_xlsx:
        apply_fixed_prices_from_excel(
            xlsx_path=exclude_xlsx, code_col="商品编码", sheet_name=0, dry_run=False
        )

    return {
        "processed": len(processed_codes),
        "excluded": len(diag_excluded),
        "unresolved": len(diag_unresolved),
        "inventory_rows_updated": len(inventory_updates),
        "allocation_rows": len(allocation_rows),
    }


# ═══════════════════════════════════════════════════════════════════
#  单商品预览（供 tool_inspect_supplier.py 等诊断脚本调用，不写库）
# ═══════════════════════════════════════════════════════════════════

def select_suppliers_for_code(
    code: str,
    min_sizes: Optional[int] = None,
    max_suppliers: Optional[int] = None,
) -> dict:
    """
    对单个商品跑一遍与 allocate_and_sync 相同的贪心选择算法，只返回结果、不写库。
    返回 {"chosen": [{"site","min_eff_price","sizes_in_stock"}, ...],
          "covered_sizes": set(size_norm), "price_basis": float | None}
    """
    min_sizes = min_sizes if min_sizes is not None else SUPPLIER_MIN_SIZES
    max_suppliers = max_suppliers if max_suppliers is not None else SUPPLIER_MAX_SITES

    engine = _get_engine()
    with engine.connect() as conn:
        offers_df = pd.read_sql(
            text("""
                SELECT site_name, size, stock_count,
                       sale_price_gbp, price_gbp, original_price_gbp, last_checked
                FROM barbour_offers
                WHERE product_code = :code AND is_active = TRUE
            """),
            conn, params={"code": code},
        )

    if offers_df.empty:
        return {"chosen": [], "covered_sizes": set(), "price_basis": None}

    offers_df["site_name"] = offers_df["site_name"].map(lambda s: canonical_site(s) or s)
    offers_df["size_norm"] = offers_df["size"].map(clean_size_for_barbour)
    offers_df["eff_price"] = offers_df.apply(_eff_price_row, axis=1)
    offers_df["in_stock"] = pd.to_numeric(offers_df["stock_count"], errors="coerce").fillna(0) > 0

    in_stock_df = offers_df[offers_df["in_stock"]]
    site_agg = (
        in_stock_df[in_stock_df["eff_price"].notna()]
        .groupby("site_name")
        .agg(
            sizes_in_stock=("size_norm", "nunique"),
            min_eff_price=("eff_price", "min"),
            latest=("last_checked", "max"),
        )
        .reset_index()
    )
    stock_sizes_map = in_stock_df.groupby("site_name")["size_norm"].apply(set).to_dict()

    chosen = _greedy_select_sites(
        site_agg,
        lambda site: stock_sizes_map.get(site, set()),
        min_sizes,
        max_suppliers,
    )
    covered: Set[str] = set()
    for c in chosen:
        covered |= stock_sizes_map.get(c["site"], set())

    price_basis = max((c["min_eff_price"] for c in chosen), default=None)
    return {"chosen": chosen, "covered_sizes": covered, "price_basis": price_basis}


# ═══════════════════════════════════════════════════════════════════
#  人工固定价覆盖（原 merge_offer_into_inventory.apply_fixed_prices_from_excel，原样迁移）
# ═══════════════════════════════════════════════════════════════════

def apply_fixed_prices_from_excel(
    xlsx_path: str,
    sheet_name: str | None = None,
    code_col: str = "product_code",
    source_price_col: str = "source_price_gbp",
    discount_price_col: str = "discount_price_gbp",
    also_set_original_price: bool = True,
    mark_source: bool = True,
    dry_run: bool = False,
):
    """
    从 Excel 读取固定价格清单，批量回填到 barbour_inventory（按 product_code 覆盖所有尺码行）。

    Excel 必需列（默认列名，可通过参数改）：
      - product_code
      - source_price_gbp
      - discount_price_gbp

    会更新的字段（默认）：
      - source_price_gbp / discount_price_gbp
      - original_price_gbp（可选：also_set_original_price=True 时设置为折扣价）
      - base_price_gbp（= COALESCE(discount_price_gbp, source_price_gbp)）
      - jingya_untaxed_price / taobao_store_price（由 calculate_jingya_prices 计算）
      - last_checked
      - （可选）source_site/source_offer_url 标记为 manual
    """
    if not xlsx_path or not os.path.exists(xlsx_path):
        print(f"ℹ️ 固定价格清单文件不存在，已跳过：{xlsx_path}")
        return

    engine = _get_engine()
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name)

    if isinstance(df, dict):
        if not df:
            raise ValueError("Excel 里没有任何 sheet。")
        df = next(iter(df.values()))

    if source_price_col not in df.columns or discount_price_col not in df.columns:
        # 该 Excel 只用于"排除自动分配"，没有价格列，属正常情况，直接跳过。
        return

    df = df.rename(columns={
        code_col: "product_code",
        source_price_col: "source_price_gbp",
        discount_price_col: "discount_price_gbp",
    })

    df["product_code"] = df["product_code"].astype(str).str.strip()
    df = df[df["product_code"].notna() & (df["product_code"] != "")]
    df["source_price_gbp"] = pd.to_numeric(df["source_price_gbp"], errors="coerce")
    df["discount_price_gbp"] = pd.to_numeric(df["discount_price_gbp"], errors="coerce")

    df["base_gbp"] = df["discount_price_gbp"].fillna(df["source_price_gbp"])
    df = df[df["base_gbp"].notna()]
    if df.empty:
        return

    discount = TAOBAO_STORE_DISCOUNT
    jy_list, tb_list = [], []
    for v in df["base_gbp"].tolist():
        untaxed, retail = calculate_jingya_prices(float(v))
        jy_list.append(round(float(untaxed), 2) if untaxed is not None else None)
        tb_list.append(round(float(retail) * float(discount), 2) if retail is not None else None)

    df["jingya_untaxed_price"] = jy_list
    df["taobao_store_price"] = tb_list

    src_tag = "manual_excel"
    offer_tag = f"excel:{os.path.basename(xlsx_path)}"
    payload = []
    for r in df.to_dict("records"):
        payload.append({
            "product_code": r["product_code"],
            "source_price_gbp": None if pd.isna(r["source_price_gbp"]) else float(r["source_price_gbp"]),
            "discount_price_gbp": None if pd.isna(r["discount_price_gbp"]) else float(r["discount_price_gbp"]),
            "original_price_gbp": None if (not also_set_original_price or pd.isna(r["discount_price_gbp"])) else float(r["discount_price_gbp"]),
            "base_price_gbp": None if pd.isna(r["base_gbp"]) else float(r["base_gbp"]),
            "jingya_untaxed_price": r["jingya_untaxed_price"],
            "taobao_store_price": r["taobao_store_price"],
            "source_site": src_tag,
            "source_offer_url": offer_tag,
        })

    if dry_run:
        print(f"[DryRun] 将覆盖 {len(payload)} 个 product_code 的 inventory 价格（所有尺码行）。示例前5行：")
        for x in payload[:5]:
            print(x)
        return

    with engine.begin() as conn:
        if mark_source:
            sql = text("""
                UPDATE barbour_inventory
                SET
                    source_price_gbp     = :source_price_gbp,
                    original_price_gbp   = COALESCE(:original_price_gbp, original_price_gbp),
                    discount_price_gbp   = :discount_price_gbp,
                    base_price_gbp       = :base_price_gbp,
                    jingya_untaxed_price = :jingya_untaxed_price,
                    taobao_store_price   = :taobao_store_price,
                    source_site          = :source_site,
                    source_offer_url     = :source_offer_url,
                    last_checked         = NOW()
                WHERE product_code = :product_code
            """)
        else:
            sql = text("""
                UPDATE barbour_inventory
                SET
                    source_price_gbp     = :source_price_gbp,
                    original_price_gbp   = COALESCE(:original_price_gbp, original_price_gbp),
                    discount_price_gbp   = :discount_price_gbp,
                    base_price_gbp       = :base_price_gbp,
                    jingya_untaxed_price = :jingya_untaxed_price,
                    taobao_store_price   = :taobao_store_price,
                    last_checked         = NOW()
                WHERE product_code = :product_code
            """)
        conn.execute(sql, payload)

    print(f"✅ 固定价格已回填到 barbour_inventory：{len(payload)} 个 product_code（覆盖所有尺码行）。")


# ═══════════════════════════════════════════════════════════════════
#  CLI：默认 dry-run，加 --apply 才真正写库
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys

    apply = "--apply" in sys.argv
    allocate_and_sync(dry_run=not apply)

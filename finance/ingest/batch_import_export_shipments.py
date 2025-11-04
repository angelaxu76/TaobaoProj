# -*- coding: utf-8 -*-
"""
batch_import_export_shipments.py
- 宽松文件匹配 + 坐标读取（Header=第15行, B:I）
- 自动建表 & 自动迁移（缺列则 ALTER TABLE ADD COLUMN）
"""

import os
import re
import datetime as dt
from typing import List, Dict, Any, Optional
import re
from PyPDF2 import PdfReader
import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
except Exception:
    openpyxl = None


# ---------------------- logging ----------------------
def _now(): return dt.datetime.now().strftime("%H:%M:%S")
def _info(msg: str): print(f"{_now()} [INFO] {msg}")
def _warn(msg: str): print(f"{_now()} [WARNING] {msg}")
def _debug(msg: str): print(f"{_now()} [DEBUG] {msg}")


# ---------------------- DB helpers -------------------
def _db_url_from_config() -> str:
    try:
        from config import DB_URL  # type: ignore
        if DB_URL:
            _info(f"[DB] Using DB URL from config: {DB_URL}")
            return DB_URL
    except Exception:
        pass

    try:
        from config import PGSQL_CONFIG  # type: ignore
        user = PGSQL_CONFIG.get("user")
        password = PGSQL_CONFIG.get("password")
        host = PGSQL_CONFIG.get("host", "localhost")
        port = PGSQL_CONFIG.get("port", 5432)
        database = PGSQL_CONFIG.get("database", PGSQL_CONFIG.get("dbname"))
        if not all([user, password, host, port, database]):
            raise KeyError("Missing keys in PGSQL_CONFIG")
        db_url = f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{database}"
        _info(f"[DB] Using DB URL from config: {db_url}")
        return db_url
    except Exception as e:
        raise RuntimeError("无法从 config 读取数据库连接信息。请在 config.py 中提供 PGSQL_CONFIG 或 DB_URL。") from e


def _get_engine() -> Engine:
    url = _db_url_from_config()
    eng = create_engine(url, pool_pre_ping=True)
    with eng.connect() as _:
        _info("[DB] Connection OK")
    return eng


# ---------------------- scan helpers -----------------
PDF_PATT = re.compile(r"(?i)\bpoe[_-]?sd\d+\.pdf$")

def _find_poe_pdf(files: List[str]) -> Optional[str]:
    for f in files:
        if PDF_PATT.search(os.path.basename(f)):
            return f
    return None

def _list_excel_files(folder: str) -> List[str]:
    out = []
    for fn in os.listdir(folder):
        if fn.startswith("~$"):
            continue
        lower = fn.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            out.append(os.path.join(folder, fn))
    return sorted(out)


# ---------------------- POE meta ---------------------
def _parse_poe_meta_from_name(pdf_path: str) -> Dict[str, Any]:
    fname = os.path.basename(pdf_path)
    m = re.search(r"(?i)(sd\d+)", fname)
    return {
        "poe_id": m.group(1).upper() if m else None,
        "poe_mrn": None,
        "poe_office": None,
        "poe_date": None,
    }

import re
from PyPDF2 import PdfReader

def _parse_poe_meta_from_pdf(pdf_path: str) -> dict:
    """
    从 POE PDF 文件中提取关键信息：
    - poe_id (SD编号)
    - poe_mrn (25GB...)
    - poe_office (GB0000..)
    - poe_date (2025-10-03)
    """
    meta = {"poe_id": None, "poe_mrn": None, "poe_office": None, "poe_date": None}
    try:
        text = ""
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            text += page.extract_text() or ""

        # 提取 POE ID
        m = re.search(r"\b(SD\d{8,})\b", text)
        if m:
            meta["poe_id"] = m.group(1)

        # 提取 MRN
        m = re.search(r"\b(25GB[A-Z0-9]{14,})\b", text)
        if m:
            meta["poe_mrn"] = m.group(1)

        # 提取 Office of Exit
        m = re.search(r"\bGB\d{6,}\b", text)
        if m:
            meta["poe_office"] = m.group(0)

        # 提取 Export Date (多种日期格式)
        m = re.search(r"(\d{2}/\d{2}/\d{4})", text)
        if m:
            try:
                d = dt.datetime.strptime(m.group(1), "%d/%m/%Y").date()
                meta["poe_date"] = d
            except Exception:
                pass

    except Exception as e:
        _warn(f"[PDF] Failed to parse {pdf_path}: {e}")

    # 若 PDF 内找不到则用文件名兜底
    if not meta["poe_id"]:
        base = os.path.basename(pdf_path)
        m = re.search(r"(?i)(sd\d+)", base)
        if m:
            meta["poe_id"] = m.group(1).upper()

    return meta

# ------------------ Invoice (by coord) ---------------
HDR_ROW_1BASED = 15           # 第15行作表头
COLS_RANGE = ("B", "I")       # B:I

EXPECTED = [
    "快递运单号\nShipment ID",
    "商品ID\nskuid",
    "LP订单号\nLP Number",
    "产品英文名\nProduct Description",
    "价值（GBP）\nValue (GBP)",
    "数量\nQuantity",
    "净重（KG）\nNet Weight (KG)",
    "欧盟税号\nHS Code",
]

CANON_KEYS = [
    "shipment_id",
    "skuid",
    "lp_number",
    "product_description",
    "value_gbp",
    "quantity",
    "net_weight_kg",
    "hs_code",
]

def _col_letter_to_idx(letter: str) -> int:
    s = 0
    for ch in letter.upper():
        s = s * 26 + (ord(ch) - 64)
    return s


def _read_invoice_by_coord(xlsx_path: str, verbose: bool = False) -> Optional[pd.DataFrame]:
    if openpyxl is None:
        _warn("[COORD] openpyxl 未安装，无法按坐标读取。")
        return None

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws: Worksheet = wb.active

    start_col = _col_letter_to_idx(COLS_RANGE[0])
    end_col = _col_letter_to_idx(COLS_RANGE[1])
    hdr_row = HDR_ROW_1BASED

    # 读表头
    headers = []
    for col in range(start_col, end_col + 1):
        v = ws.cell(row=hdr_row, column=col).value
        headers.append("" if v is None else str(v).strip())

    if verbose:
        _info(f"[INVOICE] [COORD] {os.path.basename(xlsx_path)} header row={hdr_row}, cols={COLS_RANGE[0]}:{COLS_RANGE[1]} -> {headers}")

    # 宽松校验：命中 EXPECTED >= 5 即通过
    hit = sum(1 for i, h in enumerate(headers) if i < len(EXPECTED) and h == EXPECTED[i])
    if hit < 5:
        _warn(f"[COORD] Header mismatch: {os.path.basename(xlsx_path)} 命中 {hit}/8，跳过。")
        return None

    # 读数据
    rows: List[List[Any]] = []
    empty_streak = 0
    r = hdr_row + 1
    while True:
        values = [ws.cell(row=r, column=c).value for c in range(start_col, end_col + 1)]
        norm = [(str(v).strip() if v is not None else "") for v in values]

        key_fields = [norm[0], norm[1], norm[3]]  # Shipment ID / skuid / Product Desc
        if all(k == "" for k in key_fields):
            empty_streak += 1
        else:
            empty_streak = 0

        if empty_streak >= 3:
            break

        if any(n != "" for n in norm):
            rows.append(norm)

        r += 1
        if r - hdr_row > 2000:
            _warn("[COORD] 超过 2000 行，强制停止。")
            break

    if not rows:
        _warn(f"[COORD] {os.path.basename(xlsx_path)} 表头下面没有数据行。")
        return None

    df = pd.DataFrame(rows, columns=CANON_KEYS)

    # 清洗
    for col in ["shipment_id", "skuid", "lp_number", "product_description", "hs_code"]:
        df[col] = df[col].astype(str).str.strip()

    def _to_float(x):
        try:
            return float(str(x).strip().replace(",", ""))
        except Exception:
            return None

    def _to_int(x):
        try:
            return int(float(str(x).strip().replace(",", "")))
        except Exception:
            return None

    df["value_gbp"] = df["value_gbp"].apply(_to_float)
    df["quantity"] = df["quantity"].apply(_to_int)
    df["net_weight_kg"] = df["net_weight_kg"].apply(_to_float)

    # LP 拆分
    def _split_lp(x: str):
        x = (x or "").strip()
        if not x:
            return []
        parts = re.split(r"[\s,;\/\n]+", x)
        parts = [p.strip() for p in parts if p.strip()]
        seen, out = set(), []
        for p in parts:
            if p not in seen:
                seen.add(p)
                out.append(p)
        return out

    df["lp_list"] = df["lp_number"].apply(_split_lp)
    return df


# -------------- DDL: create & migrate table ---------------
TABLE = "export_shipments"
REQUIRED_COLUMNS = {
    "id": "SERIAL PRIMARY KEY",
    "folder_name": "TEXT",
    "invoice_file": "TEXT",
    "poe_file": "TEXT",
    "poe_id": "TEXT",
    "poe_mrn": "TEXT",
    "poe_office": "TEXT",
    "poe_date": "DATE",
    "shipment_id": "TEXT",
    "skuid": "TEXT",
    "lp_number": "TEXT",
    "product_description": "TEXT",
    "value_gbp": "NUMERIC",
    "quantity": "INTEGER",
    "net_weight_kg": "NUMERIC",
    "hs_code": "TEXT",
    "created_at": "TIMESTAMP DEFAULT NOW()",
}

CREATE_SQL = f"""
CREATE TABLE IF NOT EXISTS {TABLE} (
    {', '.join([f"{k} {v}" for k, v in REQUIRED_COLUMNS.items()])}
);
"""

def _ensure_table_and_migrate(eng: Engine):
    with eng.begin() as conn:
        # 1) 若不存在则创建（新库能一次性得到所有列）
        conn.execute(text(CREATE_SQL))

        # 2) 读取现有列，缺啥补啥
        q = text("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = :tbl
              AND table_schema = COALESCE(current_schema(), 'public')
        """)
        existing = {r[0] for r in conn.execute(q, {"tbl": TABLE}).fetchall()}

        to_add = [(col, coltype) for col, coltype in REQUIRED_COLUMNS.items() if col not in existing]
        if to_add:
            _warn(f"[DDL] {TABLE} 缺少列：{[c for c, _ in to_add]}，开始自动迁移...")
            for col, coltype in to_add:
                # 注意：不能再次添加 PRIMARY KEY；如果缺的是 id，改用 bigserial 普通列也可，这里假设旧表有 id。
                if col == "id":
                    # 旧表若真没有 id，给一个可空的 bigserial
                    conn.execute(text(f'ALTER TABLE {TABLE} ADD COLUMN IF NOT EXISTS id BIGSERIAL'))
                else:
                    conn.execute(text(f'ALTER TABLE {TABLE} ADD COLUMN IF NOT EXISTS {col} {coltype}'))
            _info(f"[DDL] 已补齐列：{[c for c, _ in to_add]}")
        else:
            _debug(f"[DDL] {TABLE} 列完整，无需迁移。")


# ------------------------- main -----------------------------
def import_poe_invoice(root_folder: str, verbose: bool = False) -> None:
    eng = _get_engine()
    _ensure_table_and_migrate(eng)

    if not os.path.isdir(root_folder):
        _warn(f"[SCAN] Root not a directory: {root_folder}")
        return

    subfolders = sorted([d for d in os.listdir(root_folder) if os.path.isdir(os.path.join(root_folder, d))])
    _info(f"[SCAN] Found {len(subfolders)} subfolder(s) under {root_folder}")

    total_inserted = 0
    for sub in subfolders:
        folder_path = os.path.join(root_folder, sub)
        _info(f"\n[SCAN] Folder: {sub}")

        files = [os.path.join(folder_path, f) for f in os.listdir(folder_path)]
        excel_files = _list_excel_files(folder_path)
        poe_pdf = _find_poe_pdf(files)

        if verbose:
            _debug(f"[SCAN] excel candidates: {[os.path.basename(x) for x in excel_files]}")
            _debug(f"[SCAN] poe candidate: {os.path.basename(poe_pdf) if poe_pdf else None}")

        _info(f"[SCAN] Folder: {sub} -> {len(excel_files)} invoice(s), {1 if poe_pdf else 0} POE(s)")

        if not excel_files:
            _warn(f"[WARN] {sub} 没有发票，跳过。")
            continue

        poe_meta = _parse_poe_meta_from_pdf(poe_pdf) if poe_pdf else {"poe_id": None, "poe_mrn": None, "poe_office": None, "poe_date": None}

        batch_rows = []
        for xlsx in excel_files:
            df = _read_invoice_by_coord(xlsx, verbose=verbose)
            if df is None or df.empty:
                _warn(f"[INVOICE] {os.path.basename(xlsx)} 无法解析或没有有效数据，跳过。")
                continue

            # explode LP；若 LP 为空则保留一行
            if df["lp_list"].map(len).gt(0).any():
                exploded = df.explode("lp_list")
                exploded["lp_number"] = exploded["lp_list"].fillna("")
                exploded = exploded.drop(columns=["lp_list"])
            else:
                exploded = df.drop(columns=["lp_list"])
                exploded["lp_number"] = exploded["lp_number"].fillna("")

            exploded["folder_name"] = sub
            exploded["invoice_file"] = os.path.basename(xlsx)
            exploded["poe_file"] = os.path.basename(poe_pdf) if poe_pdf else None
            exploded["poe_id"] = poe_meta.get("poe_id")
            exploded["poe_mrn"] = poe_meta.get("poe_mrn")
            exploded["poe_office"] = poe_meta.get("poe_office")
            exploded["poe_date"] = poe_meta.get("poe_date")

            keep = [
                "folder_name", "invoice_file", "poe_file",
                "poe_id", "poe_mrn", "poe_office", "poe_date",
                "shipment_id", "skuid", "lp_number", "product_description",
                "value_gbp", "quantity", "net_weight_kg", "hs_code"
            ]
            exploded = exploded[keep]
            batch_rows.append(exploded)

        if not batch_rows:
            _warn(f"[WARN] {sub} 没有任何可插入的数据。")
            continue

        out_df = pd.concat(batch_rows, ignore_index=True)

        with eng.begin() as conn:
            # 用原生 INSERT 批量兼容 numeric/int（to_sql 也可，这里继续用 to_sql）
            out_df.to_sql(TABLE, con=conn, if_exists="append", index=False, method="multi", chunksize=1000)

        _info(f"[OK] Inserted {len(out_df)} rows for folder {sub}")
        total_inserted += len(out_df)

    _info(f"[SUMMARY] Inserted total rows: {total_inserted}")

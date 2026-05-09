"""
按 VAT 季度（或自定义日期范围）批量生成 Credit Note Excel + PDF。

查询条件：is_returned=TRUE，return_date 在指定范围内，且尚未关联 credit_note_no。

CN 编号格式：CN-YYYY-QN-NNN（例：CN-2026-Q1-001）
明细按 POE 批次分组，每张 CN 涵盖该季所有退货行。

生成后自动回填 export_shipments.credit_note_no 和 credit_note_date。

用法：修改下方 CONFIG 后运行：
    python finance/pipeline/run_generate_credit_note.py

依赖：pip install openpyxl psycopg2-binary
"""

from __future__ import annotations
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from config import PGSQL_CONFIG
from finance_config import MARGIN_RATE

# ============================================================
# CONFIG
# ============================================================

# 日期范围（VAT 季度或自定义）
DATE_FROM = "2026-01-01"
DATE_TO   = "2026-03-31"

# CN 编号（生成前确认当前最新编号，避免重复）
CN_NUMBER = "CN-2026-Q1-001"

# CN 出具日期（通常为季末或实际出具日）
CN_DATE = "2026-03-31"

# 输出目录
OUTPUT_DIR = r"C:\Users\angel\OneDrive\CrossBorderDocs_UK\02_Invoices\CreditNotes"

# True = 仅预览受影响行，不写库不生成文件；False = 正式执行
DRY_RUN = True

# ============================================================
# 公司信息
# ============================================================
ISSUER = {
    "name":       "EMINZORA TRADE LTD",
    "address":    "3rd Floor, 86-90 Paul Street, London, EC2A 4NE, UK",
    "company_no": "12345678",          # 按实际填写
    "vat_no":     "GB487486722000",
}
RECIPIENT = {
    "name":    "Alipay.com Co., Ltd.",
    "address": "2F, Building B, Huanglong Times Square, Hangzhou, 310063, CN",
}

# ============================================================


def get_conn():
    return psycopg2.connect(**PGSQL_CONFIG)


def _fetch_return_lines(date_from: str, date_to: str) -> list[dict]:
    sql = """
        SELECT
            id,
            shipment_id,
            skuid,
            product_description,
            quantity,
            purchase_unit_cost_gbp,
            poe_id,
            poe_date,
            uk_invoice_no,
            return_date
        FROM public.export_shipments
        WHERE is_returned = TRUE
          AND return_date BETWEEN %s AND %s
          AND (credit_note_no IS NULL OR credit_note_no = '')
          AND skuid IS NOT NULL AND skuid != ''
          AND purchase_unit_cost_gbp IS NOT NULL
        ORDER BY poe_id, poe_date, shipment_id
    """
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute(sql, (date_from, date_to))
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return rows


def _backfill_cn(ids: list[int], cn_no: str, cn_date: date) -> None:
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute(
        """UPDATE public.export_shipments
              SET credit_note_no  = %s,
                  credit_note_date = %s,
                  updated_at       = NOW()
            WHERE id = ANY(%s)""",
        (cn_no, cn_date, ids)
    )
    conn.commit()
    cur.close()
    conn.close()


def _build_excel(lines: list[dict], cn_no: str, cn_date: date, output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credit Note"

    # ── 样式 ──────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    hdr_font   = Font(color="FFFFFF", bold=True, size=10)
    bold_font  = Font(bold=True, size=10)
    normal_font = Font(size=10)
    grey_fill  = PatternFill("solid", fgColor="D9E1F2")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    def _cell(row, col, value, font=None, fill=None, align=None, border=None, number_format=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:          c.font          = font
        if fill:          c.fill          = fill
        if align:         c.alignment     = align
        if border:        c.border        = border
        if number_format: c.number_format = number_format
        return c

    row = 1

    # ── 标题区 ────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:I{row}")
    _cell(row, 1, "CREDIT NOTE", Font(bold=True, size=16), align=center)
    row += 1

    ws.merge_cells(f"A{row}:D{row+3}")
    issuer_text = (
        f"{ISSUER['name']}\n"
        f"{ISSUER['address']}\n"
        f"Company No: {ISSUER['company_no']}\n"
        f"VAT No: {ISSUER['vat_no']}"
    )
    c = ws.cell(row=row, column=1, value=issuer_text)
    c.font = normal_font
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:D{row+3}")

    ws.merge_cells(f"F{row}:I{row+3}")
    recip_text = f"To:\n{RECIPIENT['name']}\n{RECIPIENT['address']}"
    c = ws.cell(row=row, column=6, value=recip_text)
    c.font = normal_font
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"F{row}:I{row+3}")
    row += 4

    # CN 元数据
    for label, value in [
        ("Credit Note No:", cn_no),
        ("Date:",           cn_date.strftime("%d %B %Y")),
        ("Period:",         f"{DATE_FROM} to {DATE_TO}"),
    ]:
        _cell(row, 6, label, bold_font)
        ws.merge_cells(f"G{row}:I{row}")
        _cell(row, 7, value, normal_font)
        row += 1
    row += 1

    # ── 明细表头 ──────────────────────────────────────────────
    headers = [
        "POE ID", "POE Date", "Invoice No",
        "Shipment ID", "SKU", "Description",
        "Qty", "Unit Cost (GBP)", "Credit Amount (GBP)",
    ]
    for col, h in enumerate(headers, 1):
        _cell(row, col, h, hdr_font, hdr_fill, center, thin_border)
    ws.row_dimensions[row].height = 20
    row += 1

    # ── 明细行 ────────────────────────────────────────────────
    total_credit = 0.0
    prev_poe = None

    for line in lines:
        poe_id    = line["poe_id"] or ""
        poe_date  = line["poe_date"]
        inv_no    = line["uk_invoice_no"] or ""
        sid       = line["shipment_id"] or ""
        skuid     = line["skuid"] or ""
        desc      = line["product_description"] or ""
        qty       = line["quantity"] or 1
        cost      = float(line["purchase_unit_cost_gbp"] or 0)
        # 按与 CI 相同的加成计算卖价（credit 按原卖价冲销）
        unit_sell = round(cost * (1 + MARGIN_RATE), 2)
        credit    = round(unit_sell * qty, 2)
        total_credit += credit

        fill = grey_fill if poe_id != prev_poe and prev_poe is not None else None
        prev_poe = poe_id

        vals = [
            poe_id,
            poe_date.strftime("%Y-%m-%d") if poe_date else "",
            inv_no, sid, skuid, desc, qty,
            unit_sell, credit,
        ]
        for col, v in enumerate(vals, 1):
            fmt = '#,##0.00' if col in (8, 9) else None
            _cell(row, col, v, normal_font, fill, right if col >= 7 else None,
                  thin_border, fmt)
        row += 1

    # ── 合计行 ────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    _cell(row, 1, "TOTAL CREDIT (GBP)", bold_font, align=right)
    _cell(row, 8, "", bold_font, fill=hdr_fill)
    _cell(row, 9, round(total_credit, 2), bold_font, hdr_fill, right,
          number_format="#,##0.00")
    ws.row_dimensions[row].height = 18
    row += 2

    # ── 备注 ──────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:I{row}")
    _cell(row, 1,
          "This credit note is issued in respect of returned goods "
          "originally exported under the referenced POE declarations.",
          Font(italic=True, size=9))

    # ── 列宽 ──────────────────────────────────────────────────
    col_widths = [18, 12, 16, 22, 16, 30, 6, 16, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def main():
    cn_date   = datetime.strptime(CN_DATE, "%Y-%m-%d").date()
    lines     = _fetch_return_lines(DATE_FROM, DATE_TO)

    if not lines:
        print(f"日期范围 {DATE_FROM} ~ {DATE_TO} 内没有待处理的退货行（is_returned=TRUE，无CN编号，有成本数据）。")
        return

    # 统计
    poes        = sorted(set(r["poe_id"] for r in lines))
    total_lines = len(lines)
    total_value = sum(
        round(float(r["purchase_unit_cost_gbp"] or 0) * (1 + MARGIN_RATE) * (r["quantity"] or 1), 2)
        for r in lines
    )

    print(f"{'=' * 60}")
    print(f"CN 编号  : {CN_NUMBER}")
    print(f"CN 日期  : {CN_DATE}")
    print(f"日期范围 : {DATE_FROM} ~ {DATE_TO}")
    print(f"涉及 POE : {len(poes)} 个  {poes}")
    print(f"退货行数 : {total_lines}")
    print(f"冲销金额 : £{total_value:,.2f}")
    print(f"模式     : {'DRY RUN（仅预览）' if DRY_RUN else '正式执行'}")
    print(f"{'=' * 60}\n")

    for line in lines:
        print(f"  {line['poe_id']}  {line['shipment_id']}  {line['skuid']}  "
              f"qty={line['quantity']}  cost={line['purchase_unit_cost_gbp']}  "
              f"return={line['return_date']}")

    if DRY_RUN:
        print(f"\n→ 确认无误后将 DRY_RUN 改为 False 再次运行。")
        return

    # 生成 Excel
    out_path = Path(OUTPUT_DIR) / f"{CN_NUMBER}.xlsx"
    _build_excel(lines, CN_NUMBER, cn_date, out_path)
    print(f"\n✓ Credit Note 已生成: {out_path}")

    # 回填数据库
    ids = [r["id"] for r in lines]
    _backfill_cn(ids, CN_NUMBER, cn_date)
    print(f"✓ 已回填 credit_note_no={CN_NUMBER} 到 {len(ids)} 行")


if __name__ == "__main__":
    main()

"""
自动填充 *_costs.xlsx 中的 supplier_order_no 字段。

支持两种淘宝订单导出文件格式：
  CSV（3个月前订单）：物流单号 格式为 No:78939466185575，编码 gb18030
  XLSX（3个月内订单）：物流单号 格式为纯数字 79001585231145

两种格式中，商家备注 均映射到 supplier_order_no（仅覆盖空单元格）。

用法：按需修改下方 CONFIG 区域后运行：
    python finance/pipeline/run_fill_supplier_order_no.py
"""
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")

# ============================================================
# CONFIG（按需修改）
# ============================================================
COSTS_DIR    = r"G:\temp\taobao\poe"   # *_costs.xlsx 所在目录
ORDER_DIR    = r"G:\temp\taobao\tx"    # ExportOrderList*.csv / *.xlsx 所在目录
CSV_ENCODING = "gb18030"
# ============================================================


def _extract_tracking_no(raw: str) -> str:
    """
    将物流单号原始字符串统一提取为纯数字字符串。
    CSV 格式：'No:78939466185575' → '78939466185575'
    XLSX 格式：'79001585231145'   → '79001585231145'
    """
    raw = raw.strip()
    m = re.match(r"^No:(\d+)$", raw)
    if m:
        return m.group(1)
    if re.match(r"^\d+$", raw):
        return raw
    return ""


def _load_order_lookup(order_dir: str) -> dict[str, str]:
    """
    扫描 order_dir 中所有 ExportOrderList*.csv 和 ExportOrderList*.xlsx，
    返回 {物流单号: 商家备注} 查找字典。
    CSV 和 XLSX 均处理，后扫描的文件不覆盖先扫描的已有记录。
    """
    lookup: dict[str, str] = {}
    root = Path(order_dir)

    csv_files  = sorted(root.glob("ExportOrderList*.csv"))
    xlsx_files = sorted(root.glob("ExportOrderList*.xlsx"))

    if not csv_files and not xlsx_files:
        print(f"警告：在 {order_dir} 未找到 ExportOrderList*.csv / *.xlsx 文件")
        return lookup

    # --- CSV（3个月前，物流单号格式：No:XXXXXX）---
    for path in csv_files:
        df = pd.read_csv(path, encoding=CSV_ENCODING, dtype=str)
        df.columns = [c.strip() for c in df.columns]
        if "物流单号" not in df.columns or "商家备注" not in df.columns:
            print(f"  [跳过] {path.name}：缺少必需列（物流单号 / 商家备注）")
            continue
        count_before = len(lookup)
        for _, row in df.iterrows():
            tracking = _extract_tracking_no(str(row["物流单号"]) if pd.notna(row["物流单号"]) else "")
            note = str(row["商家备注"]).strip() if pd.notna(row["商家备注"]) else ""
            if tracking and note and tracking not in lookup:
                lookup[tracking] = note
        print(f"  [CSV ] {path.name}：新增 {len(lookup) - count_before} 条（合计 {len(lookup)} 条）")

    # --- XLSX（3个月内，物流单号格式：纯数字）---
    for path in xlsx_files:
        df = pd.read_excel(path, dtype=str)
        df.columns = [c.strip() for c in df.columns]
        if "物流单号" not in df.columns or "商家备注" not in df.columns:
            print(f"  [跳过] {path.name}：缺少必需列（物流单号 / 商家备注）")
            continue
        count_before = len(lookup)
        for _, row in df.iterrows():
            tracking = _extract_tracking_no(str(row["物流单号"]) if pd.notna(row["物流单号"]) else "")
            note = str(row["商家备注"]).strip() if pd.notna(row["商家备注"]) else ""
            # xlsx 商家备注中可能含换行符，替换为空格
            note = re.sub(r"\s*\n\s*", " ", note)
            if tracking and note and tracking not in lookup:
                lookup[tracking] = note
        print(f"  [XLSX] {path.name}：新增 {len(lookup) - count_before} 条（合计 {len(lookup)} 条）")

    return lookup


def _fill_costs_xlsx(xlsx_path: Path, lookup: dict[str, str]) -> tuple[int, int]:
    """
    更新单个 *_costs.xlsx 的 supplier_order_no 列。
    返回 (填充行数, 因已有值跳过行数)。
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1] if cell.value is not None}

    if "shipment_id" not in headers or "supplier_order_no" not in headers:
        print(f"  [跳过] {xlsx_path.name}：缺少必需列 shipment_id / supplier_order_no")
        return 0, 0

    sid_col = headers["shipment_id"]
    sno_col = headers["supplier_order_no"]

    filled = skipped = 0

    for row in ws.iter_rows(min_row=2):
        sid_val = row[sid_col - 1].value
        if sid_val is None:
            continue

        try:
            tracking_no = str(int(float(str(sid_val))))
        except (ValueError, OverflowError):
            tracking_no = str(sid_val).strip()

        sno_cell = row[sno_col - 1]
        existing = sno_cell.value
        if existing is not None and str(existing).strip():
            skipped += 1
            continue

        if tracking_no in lookup:
            sno_cell.value = lookup[tracking_no]
            filled += 1

    wb.save(xlsx_path)
    return filled, skipped


def main():
    print("=== 自动填充 supplier_order_no ===\n")

    print("步骤 1：加载订单查找表（CSV + XLSX）")
    lookup = _load_order_lookup(ORDER_DIR)
    if not lookup:
        print("查找表为空，退出。")
        return
    print(f"查找表共 {len(lookup)} 条\n")

    print("步骤 2：处理 costs.xlsx 文件")
    costs_root = Path(COSTS_DIR)
    xlsx_files = sorted(p for p in costs_root.glob("*_costs.xlsx") if not p.name.startswith("~"))

    if not xlsx_files:
        print(f"在 {COSTS_DIR} 未找到 *_costs.xlsx 文件")
        return

    total_filled = 0
    for xlsx_path in xlsx_files:
        print(f"  {xlsx_path.name} ... ", end="")
        filled, skipped = _fill_costs_xlsx(xlsx_path, lookup)
        print(f"填充 {filled} 行，已有值跳过 {skipped} 行")
        total_filled += filled

    print(f"\n完成：共填充 {total_filled} 个 supplier_order_no")


if __name__ == "__main__":
    main()

"""
扫描 06_Export_Proofs 下所有 SD 文件夹，提取 CommercialInvoice_PoE DOCX
中的日期和总金额，写入 Intercompany_Ledger.xlsx 的 Ledger 页。

- 已存在的 CI 行（按 File Name 去重）自动跳过，可安全重复运行
- CI 金额写为负数（减少 HK 预付余额）
- 写入后按日期重新排序，重算 Running Balance，更新 Summary 页汇总数

用法：直接运行，无需修改参数（路径已固定）
    python finance/pipeline/run_update_intercompany_ledger.py
"""
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
from docx import Document

sys.stdout.reconfigure(encoding="utf-8")

# ============================================================
# CONFIG
# ============================================================
EXPORT_PROOFS_DIR = r"C:\Users\angel\OneDrive\CrossBorderDocs_UK\06_Export_Proofs"
LEDGER_PATH       = r"C:\Users\angel\OneDrive\CrossBorderDocs_UK\04_PaymentProofs\Intercompany_Ledger\Intercompany_Ledger.xlsx"
LEDGER_SHEET      = "Ledger"
SUMMARY_SHEET     = "Summary"
# ============================================================

_TOTAL_RE = re.compile(r"Total Amount Payable \(GBP\):\s*([\d,]+\.?\d*)", re.IGNORECASE)
_DATE_RE  = re.compile(r"Invoice Date:\s*(\d{4}-\d{2}-\d{2})", re.IGNORECASE)


def _parse_ci_docx(docx_path: Path) -> tuple[str, float] | None:
    """
    从 CommercialInvoice_PoE DOCX 提取 (invoice_date, total_gbp)。
    解析失败返回 None。
    """
    try:
        doc = Document(str(docx_path))
    except Exception as e:
        print(f"    [错误] 无法读取 {docx_path.name}: {e}")
        return None

    invoice_date = total_gbp = None

    for para in doc.paragraphs:
        text = para.text.strip()
        if not invoice_date:
            m = _DATE_RE.search(text)
            if m:
                invoice_date = m.group(1)
        if not total_gbp:
            m = _TOTAL_RE.search(text)
            if m:
                total_gbp = float(m.group(1).replace(",", ""))
        if invoice_date and total_gbp:
            break

    if not invoice_date or not total_gbp:
        print(f"    [警告] {docx_path.name}：无法提取日期或金额（date={invoice_date}, total={total_gbp}）")
        return None

    return invoice_date, total_gbp


def _collect_ci_records(proofs_dir: str) -> list[dict]:
    """
    扫描所有 SD 文件夹，返回 CI 记录列表。
    每条：{date, amount, file_name, notes}
    """
    records = []
    root = Path(proofs_dir)

    for sd_dir in sorted(root.iterdir()):
        if not sd_dir.is_dir() or not sd_dir.name.startswith("SD"):
            continue
        docx_files = sorted(sd_dir.glob("CommercialInvoice_PoE_*.docx"))
        if not docx_files:
            continue
        docx_path = docx_files[0]
        result = _parse_ci_docx(docx_path)
        if result is None:
            continue
        invoice_date, total_gbp = result
        records.append({
            "date":      invoice_date,
            "amount":    -total_gbp,          # 负数：减少 HK 预付余额
            "file_name": docx_path.name,
            "notes":     f"CI  {sd_dir.name.split('_')[0]}",
        })

    return records


def _update_ledger(ledger_path: str, new_records: list[dict]) -> tuple[int, int]:
    """
    将新 CI 记录追加进 Ledger 页，按日期排序，重算 Running Balance。
    返回 (新增行数, 跳过行数)。
    """
    wb = openpyxl.load_workbook(ledger_path)

    # ── 读取现有 Ledger ──────────────────────────────────────
    ws_ledger = wb[LEDGER_SHEET]
    headers = [cell.value for cell in ws_ledger[1]]

    existing_rows = []
    for row in ws_ledger.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            existing_rows.append(dict(zip(headers, row)))

    existing_filenames = {
        str(r.get("File Name", "") or "").strip()
        for r in existing_rows
        if r.get("Type") == "CI"
    }

    # ── 过滤出真正新增的记录 ─────────────────────────────────
    added = skipped = 0
    for rec in new_records:
        if rec["file_name"] in existing_filenames:
            skipped += 1
            continue
        existing_rows.append({
            "Date":            rec["date"],
            "Type":            "CI",
            "Amount (GBP)":    rec["amount"],
            "File Name":       rec["file_name"],
            "Notes":           rec["notes"],
            "Running Balance": None,          # 后面统一重算
        })
        added += 1

    if added == 0:
        wb.close()
        return added, skipped

    # ── 按日期排序 ───────────────────────────────────────────
    def _sort_key(r):
        d = r.get("Date", "")
        if hasattr(d, "strftime"):
            return str(d.date())
        return str(d) if d else ""

    existing_rows.sort(key=_sort_key)

    # ── 重算 Running Balance ─────────────────────────────────
    balance = 0.0
    for r in existing_rows:
        amt = r.get("Amount (GBP)") or 0
        try:
            balance = round(balance + float(amt), 2)
        except (TypeError, ValueError):
            pass
        r["Running Balance"] = balance

    # ── 清空旧数据行，重写 ────────────────────────────────────
    for row in ws_ledger.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    for i, r in enumerate(existing_rows, start=2):
        ws_ledger.cell(i, 1).value = r.get("Date")
        ws_ledger.cell(i, 2).value = r.get("Type")
        ws_ledger.cell(i, 3).value = r.get("Amount (GBP)")
        ws_ledger.cell(i, 4).value = r.get("File Name")
        ws_ledger.cell(i, 5).value = r.get("Notes")
        ws_ledger.cell(i, 6).value = r.get("Running Balance")

    # ── 更新 Summary 页 ──────────────────────────────────────
    if SUMMARY_SHEET in wb.sheetnames:
        total_payments = sum(
            float(r.get("Amount (GBP)") or 0)
            for r in existing_rows if r.get("Type") == "Payment"
        )
        total_ci = sum(
            abs(float(r.get("Amount (GBP)") or 0))
            for r in existing_rows if r.get("Type") == "CI"
        )
        running_balance = round(total_payments - total_ci, 2)
        today_str = str(date.today())

        ws_sum = wb[SUMMARY_SHEET]
        # Summary 固定布局：逐行扫描找关键词后更新右侧单元格
        _summary_updates = {
            "Total HK Payments":              total_payments,
            "Total UK Invoices (CI)":         total_ci,
            "Running Balance (HK prepayment balance)": running_balance,
            "Last Update Date":               today_str,
        }
        for row in ws_sum.iter_rows():
            for cell in row:
                label = str(cell.value or "").strip()
                if label in _summary_updates:
                    # 更新同行右边第一个有值的单元格（或紧邻列）
                    target = ws_sum.cell(cell.row, cell.column + 1)
                    target.value = _summary_updates[label]

    wb.save(ledger_path)
    wb.close()
    return added, skipped


def main():
    print("=== 更新 Intercompany Ledger（CI 条目）===\n")

    print("步骤 1：扫描 CommercialInvoice_PoE DOCX")
    records = _collect_ci_records(EXPORT_PROOFS_DIR)
    print(f"共找到 {len(records)} 个 CI 文件\n")

    if not records:
        print("未找到任何 CI 文件，退出。")
        return

    print("步骤 2：写入 Ledger")
    added, skipped = _update_ledger(LEDGER_PATH, records)
    print(f"\n完成：新增 {added} 行 CI，跳过（已存在）{skipped} 行")


if __name__ == "__main__":
    main()
